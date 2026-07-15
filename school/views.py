from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.http import require_http_methods
from django.db.models import Q
from django.core.paginator import Paginator
from userauths.models import User
from .models import Grade, School, Parent,StaffProfile,Student, ScanLog, SmartID,Scholarship
import logging
from django.db.models import Q, Count, Avg, Case, When, Value, FloatField, ExpressionWrapper, F, DurationField
from django.db import IntegrityError
import os
from django.db.models import Q, Value, CharField
from django.db.models.functions import Concat

from django.db.models import Q
from collections import defaultdict
from django.db.models import Count
from django.views.decorators.http import require_POST
import pandas as pd
from django.http import JsonResponse
from django.conf import settings
from django.template.loader import render_to_string
from django.utils.safestring import mark_safe
from django.core.mail import send_mail
from itertools import cycle
import random
from django.db.models.functions import Lower
import string
from django.utils.dateparse import parse_date
import datetime
import threading
from datetime import timedelta,date
from .services.timetable_generator import generate_for_stream
from collections import defaultdict
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Count, Avg, Case, When, Value, FloatField, F,Func, ExpressionWrapper
from django.db.models.functions import Coalesce
from django.db.models.functions import Extract
from django.db.models.functions import Coalesce
from django.http import JsonResponse, HttpResponseForbidden
from collections import defaultdict
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.core.paginator import Paginator
from django.core.exceptions import ValidationError
from django.http import JsonResponse, HttpResponse
from django.conf import settings
from django.template.loader import render_to_string
from django.utils.safestring import mark_safe
from django.core.mail import send_mail
from django.db import transaction
from django.utils import timezone
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from django.db.models import Count, Q, OuterRef, Exists,Prefetch
from django.utils.dateparse import parse_date

import csv
from io import StringIO
import logging
from decimal import Decimal
from django.utils.timezone import localdate
import uuid
from userauths.models import User
from .models import (
    Grade, School, Parent, StaffProfile, Student, Subject, Enrollment, Timetable, Lesson, PolicymakerProfile, AttendanceAlert,
    Session, Attendance, DisciplineRecord, SummaryReport, Notification, SmartID, ScanLog, TeacherStreamAssignment,
    ClassTeacherAssignment,
    Payment, Assignment, Submission, Role, Invoice, SchoolSubscription, SubscriptionPlan, UploadedFile, County, Constituency, Ward,
    ContactMessage, MpesaStkPushRequestResponse, MpesaPayment,GradeAttendance, Streams,Term, TimeSlot, AcademicYear,
    SubjectEnrollment,AcademicYear,SubCounty,Pathway,Upload,SubjectCatalog,
    Announcement, FeeInvoice,
    ExamSession, ExamResult, cbc_grade_band, kcse_grade,
    FeeStructure, FeeType, FEE_TYPE_CHOICES,
)
from .forms import (
    # Assuming forms exist or need to be created; placeholders for now
    GradeForm,ParentCreationForm, ParentUpdateForm,StaffCreationForm, StaffUpdateForm,
    StudentCreationForm, StudentUpdateForm,SmartIDForm,GenerateTimetableForm,
    SubjectForm, EnrollmentForm, TimetableForm, LessonForm, SessionForm,GradeUploadForm,
    AttendanceForm, DisciplineRecordForm, NotificationForm, PaymentForm,
    AssignmentForm, SubmissionForm, RoleForm, InvoiceForm, SchoolSubscriptionForm,
    ContactMessageForm,ParentStudentCreationForm,TermForm,TimeSlotForm,AssignParentStudentForm,
    SubjectCatalogForm,SubjectActivationForm,ComplaintForm,BulkSubjectUploadForm,
    BulkNotificationForm,
    ExamSessionForm, ExamResultForm, ExamUploadForm,
    FeeStructureForm, BulkInvoiceGenerateForm, FeePaymentUploadForm, FeeInvoiceForm,
)
from .models import AuditLog, Complaint
from kiswate_digital_app.forms import StreamForm

# Create your views here.
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
import logging
import requests
logger = logging.getLogger(__name__)


import csv
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from django.db.models import Min, Max
from school.models import Attendance, Lesson

INCIDENT_TYPE_CHOICES = [
    ('late', 'Late Arrival'),
    ('tardy', 'Tardy'),
    ('absence', 'Absence'),
    ('misconduct', 'Misconduct'),
    ('other', 'Other'),
]

SEVERITY_CHOICES = [
    ('minor', 'Minor'),
    ('major', 'Major'),
    ('critical', 'Critical'),
]


# =========================
# SMS + EMAIL HELPERS
# =========================

import time
import requests
from datetime import datetime
from django.utils.dateparse import parse_time

def safe_parse_time(value):
    if not value:
        return None

    value = str(value).strip()

    # Try Django parser first (24h format)
    t = parse_time(value)
    if t:
        return t

    # Fallback: AM/PM formats
    for fmt in ("%I:%M %p", "%I %p"):
        try:
            return datetime.strptime(value.upper(), fmt).time()
        except ValueError:
            pass

    return None
def _send_sms_via_eujim(to_phone_number: str, message: str, retries=3, delay=5) -> bool:
    if not to_phone_number:
        return False

    phone = str(to_phone_number)
    if phone.startswith("0"):
        phone = "254" + phone[1:]
    elif phone.startswith("+254"):
        phone = phone[1:]
    elif not phone.startswith("254"):
        phone = "254" + phone

    payload = {
        "apikey": settings.SMS_API_KEY,
        "partnerID": settings.SMS_PARTNERID,
        "shortcode": settings.SMS_SHORTCODE,
        "message": message,
        "mobile": phone,
    }

    for attempt in range(1, retries + 1):
        try:
            response = requests.post(
                'https://quicksms.advantasms.com/api/services/sendsms/',
                json=payload,
                timeout=15
            )
            if response.status_code == 200:
                data = response.json()
                r = data.get("responses", [{}])[0]
                return r.get("response-code") == 200
        except requests.exceptions.RequestException:
            pass
        if attempt < retries:
            time.sleep(delay)  # wait before retrying
    return False



def send_email(to_email: str, subject: str, message: str) -> bool:
    if not to_email:
        return False
    try:
        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [to_email],
            fail_silently=False,
        )
        return True
    except Exception as e:
        return False

# --------------------------------------------------
# SLOT HELPERS
# --------------------------------------------------

def is_first_slot(lesson):
    school = lesson.timetable.school if lesson.timetable_id else None
    qs = TimeSlot.objects.filter(school=school) if school else TimeSlot.objects.all()
    return lesson.time_slot == qs.order_by('start_time').first()


def is_last_slot(lesson):
    school = lesson.timetable.school if lesson.timetable_id else None
    qs = TimeSlot.objects.filter(school=school) if school else TimeSlot.objects.all()
    return lesson.time_slot == qs.order_by('-end_time').first()


# --------------------------------------------------
# MESSAGE BUILDERS
# --------------------------------------------------

def build_single_lesson_message(student, subject, status, lesson_date):
    return (
        f"Dear Parent, {student.user.get_full_name()} "
        f"was marked {status} for {subject.name} "
        f"on {lesson_date}."
    )

def build_daily_summary(student, date):
    stats = (
        Attendance.objects
        .filter(enrollment__student=student, date=date)
        .values('status')
        .annotate(total=Count('id'))
    )

    summary = ", ".join(f"{s['status']}: {s['total']}" for s in stats)
    return f"Daily Attendance Summary ({date}): {summary}"


# --------------------------------------------------
# SENDERS
# --------------------------------------------------

def notify_first_lesson(attendance):
    student = attendance.enrollment.student
    parents = student.parents.all()
    message = build_single_lesson_message(
        student,
        attendance.enrollment.lesson.subject,   # Subject object
        attendance.get_status_display(),
        attendance.date                   # Lesson date
    )
    for parent in parents:
        # _send_sms_via_eujim(parent.phone, message)
        if parent.user.email:
            send_mail("Lesson Attendance Notification", message, None, [parent.user.email], fail_silently=True)


def notify_last_lesson(student, date):
    parents = student.parents.all()
    message = build_daily_summary(student, date)

    for parent in parents:
        # _send_sms_via_eujim(parent.phone, message)
        if parent.user.email:
            send_mail(
                "Daily Attendance Summary",
                message,
                None,
                [parent.user.email],
                fail_silently=True
            )
@login_required
def export_attendance_csv(request):
    school = get_user_school(request.user)
    if not school:
        return HttpResponse("Access denied.", status=403)
    qs = Attendance.objects.filter(
        enrollment__school=school
    ).select_related(
        'enrollment__student__user',
        'enrollment__student__grade_level',
        'enrollment__student__stream',
        'enrollment__subject',
        'term', 'academic_year',
    ).order_by('date')

    grade_id = request.GET.get('grade')
    term_id  = request.GET.get('term')
    start    = request.GET.get('start')
    end      = request.GET.get('end')
    if grade_id:
        qs = qs.filter(enrollment__student__grade_level_id=grade_id)
    if term_id:
        qs = qs.filter(term_id=term_id)
    if start:
        qs = qs.filter(date__gte=start)
    if end:
        qs = qs.filter(date__lte=end)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="attendance_report.csv"'
    writer = csv.writer(response)
    writer.writerow(['Student', 'Grade', 'Stream', 'Subject', 'Date', 'Status', 'Term', 'Year'])
    STATUS_LABELS = {'P': 'Present', 'ET': 'Excused Tardy', 'UT': 'Unexcused Tardy',
                     'EA': 'Excused Absent', 'UA': 'Unexcused Absent', 'IB': 'In Building',
                     '18': 'Suspended', '20': 'Expelled'}
    for a in qs:
        try:
            writer.writerow([
                a.enrollment.student.user.get_full_name(),
                a.enrollment.student.grade_level.name,
                getattr(a.enrollment.student.stream, 'name', ''),
                a.enrollment.subject.name,
                a.date,
                STATUS_LABELS.get(a.status, a.status),
                a.term.name if a.term else '',
                a.academic_year.name if a.academic_year else '',
            ])
        except Exception:
            continue
    return response

@login_required
def export_attendance_pdf(request):
    school = get_user_school(request.user)
    if not school:
        return HttpResponse("Access denied.", status=403)
    qs = Attendance.objects.filter(
        enrollment__school=school
    ).select_related(
        'enrollment__student__user',
        'enrollment__student__grade_level',
        'enrollment__student__stream',
        'enrollment__subject',
        'term', 'academic_year',
    ).order_by('date')

    # Optional filters
    grade_id = request.GET.get('grade')
    term_id  = request.GET.get('term')
    start    = request.GET.get('start')
    end      = request.GET.get('end')
    if grade_id:
        qs = qs.filter(enrollment__student__grade_level_id=grade_id)
    if term_id:
        qs = qs.filter(term_id=term_id)
    if start:
        qs = qs.filter(date__gte=start)
    if end:
        qs = qs.filter(date__lte=end)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="attendance_report.pdf"'
    doc = SimpleDocTemplate(response, pagesize=letter, topMargin=36, bottomMargin=36, leftMargin=36, rightMargin=36)
    styles = getSampleStyleSheet()
    story = []
    from reportlab.platypus import Spacer
    story.append(Paragraph(f"<b>{school.name}</b>", styles['Title']))
    story.append(Paragraph("Attendance Report", styles['Heading2']))
    story.append(Paragraph(f"Generated: {timezone.now().strftime('%d %B %Y %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 12))

    header = ['Student', 'Grade', 'Stream', 'Subject', 'Date', 'Status', 'Term']
    rows = [header]
    STATUS_LABELS = {'P': 'Present', 'ET': 'Exc. Tardy', 'UT': 'Unexc. Tardy',
                     'EA': 'Exc. Absent', 'UA': 'Unexc. Absent', 'IB': 'In Building',
                     '18': 'Suspended', '20': 'Expelled'}
    for a in qs:
        try:
            rows.append([
                a.enrollment.student.user.get_full_name(),
                a.enrollment.student.grade_level.name,
                getattr(a.enrollment.student.stream, 'name', '—'),
                a.enrollment.subject.name,
                str(a.date),
                STATUS_LABELS.get(a.status, a.status),
                a.term.name if a.term else '—',
            ])
        except Exception:
            continue

    t = Table(rows, colWidths=[110, 55, 55, 80, 60, 65, 55])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#dee2e6')),
    ]))
    story.append(t)
    doc.build(story)
    return response


_EXCEL_MIME_TYPES = {
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
    'application/vnd.ms-excel',                                            # .xls
    'application/octet-stream',  # Some browsers report this for xlsx
}
_EXCEL_MAX_BYTES = 10 * 1024 * 1024  # 10 MB


def _validate_excel_upload(uploaded_file):
    """
    Raise ValueError if the uploaded file is too large or is not an Excel file.
    Returns None on success.
    """
    if uploaded_file.size > _EXCEL_MAX_BYTES:
        raise ValueError(f"File too large (max {_EXCEL_MAX_BYTES // 1024 // 1024} MB).")
    # Extension check
    name = (uploaded_file.name or '').lower()
    if not (name.endswith('.xlsx') or name.endswith('.xls')):
        raise ValueError("Only Excel files (.xlsx / .xls) are accepted.")
    # MIME check (browser-reported, not authoritative, but adds one layer)
    content_type = getattr(uploaded_file, 'content_type', '') or ''
    if content_type and content_type not in _EXCEL_MIME_TYPES:
        raise ValueError(f"Invalid file type '{content_type}'. Upload an Excel file.")


def _safe_next_redirect(request, default_url_name):
    """Return a redirect to `next` only if it's a safe same-host relative URL."""
    from django.utils.http import url_has_allowed_host_and_scheme
    next_url = request.POST.get('next') or request.GET.get('next', '')
    if next_url and url_has_allowed_host_and_scheme(next_url, allowed_hosts={request.get_host()}):
        return redirect(next_url)
    return redirect(default_url_name)


def get_user_school(user):
    # Admins/principals: try school_admin_profile (School.school_admin FK) first
    if user.is_admin or user.is_principal:
        school = getattr(user, "school_admin_profile", None)
        if school:
            return school

    # Any staff with a StaffProfile (teachers, deputies, finance, HOD, etc.)
    try:
        return user.staffprofile.school
    except Exception:
        pass

    return None


def _can_access_finance(user):
    """True for admins, principals, deputies, and finance-position staff."""
    if user.is_admin or user.is_principal or user.is_deputy_principal:
        return True
    try:
        return user.staffprofile.position in ('finance', 'administrator', 'clerks')
    except Exception:
        return False

@login_required
def dashboard(request):
    import json as _json
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)
    if not school:
        return render(request, "school/error.html", {"message": "No school profile found."})

    if (user.is_principal or user.is_deputy_principal) and not Subject.objects.filter(school=school).exists():
        messages.info(
            request,
            "Welcome! Please select the subjects your school offers before continuing."
        )
        return redirect('school:subject-activate-catalog')

    today = timezone.localdate()

    # ── Active term ────────────────────────────────────────────────────────────
    active_term = Term.objects.filter(school=school, is_active=True).first()
    term_filter = Q(date__range=(active_term.start_date, active_term.end_date)) if active_term else Q()

    if active_term:
        term_total_days = max((active_term.end_date - active_term.start_date).days, 1)
        term_elapsed    = max(min((today - active_term.start_date).days, term_total_days), 0)
        term_progress   = round(term_elapsed / term_total_days * 100)
        term_days_left  = max((active_term.end_date - today).days, 0)
    else:
        term_total_days = term_elapsed = term_progress = term_days_left = 0

    # ── Week navigation ────────────────────────────────────────────────────────
    date_str = request.GET.get("date")
    try:
        focus_date = timezone.datetime.strptime(date_str, "%Y-%m-%d").date() if date_str else today
    except ValueError:
        focus_date = today
    week_start = focus_date - timedelta(days=focus_date.weekday())
    week_end   = week_start + timedelta(days=6)
    week_days  = [week_start + timedelta(days=i) for i in range(7)]
    prev_week  = week_start - timedelta(days=7)
    next_week  = week_start + timedelta(days=7)
    is_current_week = week_start == today - timedelta(days=today.weekday())

    # ── Core counts ───────────────────────────────────────────────────────────
    total_teachers = StaffProfile.objects.filter(school=school, position='teacher').count()
    total_students = Student.objects.filter(school=school, is_active=True).count()
    total_parents  = Parent.objects.filter(school=school).count()
    total_grades   = Grade.objects.filter(school=school).count()
    total_streams  = Streams.objects.filter(school=school).count()

    # ── Today's attendance rate ────────────────────────────────────────────────
    today_att = Attendance.objects.filter(enrollment__school=school, date=today)
    today_total   = today_att.count()
    today_present = today_att.filter(status='P').count()
    today_rate    = round(today_present / today_total * 100) if today_total else 0

    # ── Term attendance rate ───────────────────────────────────────────────────
    term_att     = Attendance.objects.filter(enrollment__school=school).filter(term_filter)
    term_total   = term_att.count()
    term_present = term_att.filter(status='P').count()
    term_rate    = round(term_present / term_total * 100) if term_total else 0
    term_absent  = term_att.filter(status__in=['UA', 'EA']).count()
    term_tardy   = term_att.filter(status__in=['UT', 'ET']).count()

    # ── Fee collection (term) ─────────────────────────────────────────────────
    from django.db.models import Sum
    fee_qs    = Payment.objects.filter(school=school)
    fee_paid  = fee_qs.filter(status='completed').aggregate(t=Sum('amount'))['t'] or 0
    fee_total = fee_qs.aggregate(t=Sum('amount'))['t'] or 0
    fee_rate  = round(fee_paid / fee_total * 100) if fee_total else 0
    fee_pending_count = fee_qs.filter(status='pending').count()

    # ── Discipline (this term) ─────────────────────────────────────────────────
    discipline_count = DisciplineRecord.objects.filter(school=school).filter(
        Q(date__range=(active_term.start_date, active_term.end_date)) if active_term else Q()
    ).count()
    discipline_recent = DisciplineRecord.objects.filter(school=school).select_related(
        'student__user'
    ).order_by('-date')[:5]

    # ── Complaints ────────────────────────────────────────────────────────────
    open_complaints = Complaint.objects.filter(school=school, status__in=['open','in_review']).count()

    # ── Today's lessons ────────────────────────────────────────────────────────
    todays_lessons = Lesson.objects.filter(
        timetable__school=school,
        lesson_date=today,
    ).select_related(
        'subject', 'teacher__user', 'stream__grade', 'time_slot'
    ).order_by('time_slot__start_time')[:8]

    # ── Weekly attendance chart data ───────────────────────────────────────────
    week_counts = {s: [] for s in ['P', 'ET', 'UT', 'EA', 'UA']}
    for status in week_counts:
        day_dict = {
            d['date']: d['count']
            for d in Attendance.objects.filter(
                status=status,
                date__range=(week_start, week_end),
                enrollment__school=school
            ).values('date').annotate(count=Count('id'))
        }
        for day in week_days:
            week_counts[status].append(day_dict.get(day, 0))

    chart_labels  = [d.strftime('%a %-d') for d in week_days]
    chart_present = week_counts['P']
    chart_absent  = [week_counts['UA'][i] + week_counts['EA'][i] for i in range(len(week_days))]
    chart_tardy   = [week_counts['UT'][i] + week_counts['ET'][i] for i in range(len(week_days))]

    # ── Attendance donut (today) ───────────────────────────────────────────────
    donut_data = [
        today_present,
        today_att.filter(status__in=['UA','EA']).count(),
        today_att.filter(status__in=['UT','ET']).count(),
        today_att.filter(status__in=['18','20','IB']).count(),
    ]

    # ── Teachers with missed lessons this week ─────────────────────────────────
    teachers_missed = list(
        StaffProfile.objects.filter(position='teacher', school=school).annotate(
            missed_count=Count(
                'lessons_taught__l_enrollments__attendances',
                filter=Q(
                    lessons_taught__l_enrollments__attendances__status__in=['UT','UA'],
                    lessons_taught__l_enrollments__attendances__date__range=(week_start, week_end),
                )
            )
        ).filter(missed_count__gt=0).select_related('user').order_by('-missed_count')[:8]
    )

    # ── Students missed lessons by stream this week ────────────────────────────
    student_missed_by_stream = []
    for stream in Streams.objects.filter(school=school).select_related('grade'):
        qs = Student.objects.filter(
            enrollments__lesson__stream=stream,
            enrollments__attendances__status__in=['UT','UA'],
            enrollments__attendances__date__range=(week_start, week_end),
        ).annotate(
            missed_count=Count(
                'enrollments__attendances',
                filter=Q(
                    enrollments__attendances__status__in=['UT','UA'],
                    enrollments__attendances__date__range=(week_start, week_end),
                )
            )
        ).distinct().order_by('-missed_count')[:10]
        if qs.exists():
            student_missed_by_stream.append({'stream': stream, 'students': qs})

    # ── Suspensions / Expulsions this week ────────────────────────────────────
    suspensions = Attendance.objects.filter(
        status='18', enrollment__school=school,
        date__range=(week_start, week_end),
    ).select_related('enrollment__student__user', 'enrollment__lesson__subject').order_by('-date')[:10]

    expulsions = Attendance.objects.filter(
        status='20', enrollment__school=school,
        date__range=(week_start, week_end),
    ).select_related('enrollment__student__user', 'enrollment__lesson__subject').order_by('-date')[:10]

    # ── Grade attendance breakdown (term) ─────────────────────────────────────
    grade_att = (
        Attendance.objects.filter(enrollment__school=school)
        .filter(term_filter)
        .values(grade_name=F('enrollment__lesson__stream__grade__name'))
        .annotate(
            total=Count('id'),
            present=Count('id', filter=Q(status='P')),
        )
        .annotate(rate=ExpressionWrapper(
            F('present') * 100.0 / F('total'), output_field=FloatField()
        ))
        .order_by('grade_name')
    )

    context = {
        # Identity
        "school":       school,
        "active_term":  active_term,
        "today":        today,
        # Term progress
        "term_progress":   term_progress,
        "term_days_left":  term_days_left,
        # KPI tiles
        "total_teachers": total_teachers,
        "total_students": total_students,
        "total_parents":  total_parents,
        "total_grades":   total_grades,
        "total_streams":  total_streams,
        # Today
        "today_total":   today_total,
        "today_present": today_present,
        "today_rate":    today_rate,
        # Term attendance
        "term_rate":    term_rate,
        "term_total":   term_total,
        "term_present": term_present,
        "term_absent":  term_absent,
        "term_tardy":   term_tardy,
        # Fees
        "fee_paid":          fee_paid,
        "fee_total":         fee_total,
        "fee_rate":          fee_rate,
        "fee_pending_count": fee_pending_count,
        # Discipline / complaints
        "discipline_count":  discipline_count,
        "discipline_recent": discipline_recent,
        "open_complaints":   open_complaints,
        # Today's schedule
        "todays_lessons": todays_lessons,
        # Week navigation
        "week_days":        week_days,
        "week_start":       week_start,
        "week_end":         week_end,
        "prev_week":        prev_week,
        "next_week":        next_week,
        "is_current_week":  is_current_week,
        # Chart data – raw Python lists; json_script in the template handles serialization
        "chart_labels_json":  chart_labels,
        "chart_present_json": chart_present,
        "chart_absent_json":  chart_absent,
        "chart_tardy_json":   chart_tardy,
        "donut_data_json":    donut_data,
        "donut_data":         donut_data,
        # Alerts
        "teachers_missed":         teachers_missed,
        "student_missed_by_stream": student_missed_by_stream,
        "suspensions":             suspensions,
        "expulsions":              expulsions,
        "grade_att":               grade_att,
    }

    return render(request, "school/dashboard.html", context)
@login_required
def time_slot_list(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    slots = TimeSlot.objects.filter(school=school)
    form = TimeSlotForm(school=school)  # Empty form for modal
    return render(request, "school/time_slots.html", {
        "slots": slots,
        "school": school,
        "form": form,  # Enables {{ form.field }} in template
    })

@login_required
def time_slot_create(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:time-slot-list')

    # Optional: Role check (from previous code)
    # user_role = get_user_role(request.user)
    # if user_role not in ['admin']:
    #     messages.error(request, "Insufficient permissions.")
    #     return redirect("school:time-slot-list")

    form = TimeSlotForm(request.POST or None, school=school)
    if request.method == "POST" and form.is_valid():
        try:
            slot = form.save(commit=False)
            slot.school = school
            try:
                slot.created_by = request.user.staffprofile
                slot.updated_by = request.user.staffprofile
            except ObjectDoesNotExist:
                messages.error(request, "Missing staff profile for audit trail.")
                return redirect("school:time-slot-list")
            slot.save()
            messages.success(request, "Time slot created successfully.")
            return redirect("school:time-slot-list")
        except IntegrityError as e:
            messages.error(request, f"Save failed: Time slot may overlap or violate constraints. Details: {str(e)}")
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
    return render(request, "school/time_slots.html", {
        "form": form, "school": school, "mode": "create"
    })

@login_required
def time_slot_edit(request, pk):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access to time slots.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:time-slot-list')
    # Optional role check
    # user_role = get_user_role(request.user)
    # if user_role not in ['admin']:
    #     messages.error(request, "Insufficient permissions.")
    #     return redirect("school:time-slot-list")

    slot = get_object_or_404(TimeSlot, pk=pk, school=school)
    form = TimeSlotForm(request.POST or None, instance=slot, school=slot.school)
    if request.method == "POST" and form.is_valid():
        try:
            slot = form.save(commit=False)
            try:
                slot.updated_by = request.user.staffprofile
            except ObjectDoesNotExist:
                messages.error(request, "Missing staff profile for audit trail.")
                return redirect("school:time-slot-list")
            slot.save()
            messages.success(request, "Time slot updated successfully.")
            return redirect("school:time-slot-list")
        except IntegrityError as e:
            messages.error(request, f"Save failed: Time slot may overlap or violate constraints. Details: {str(e)}")
        except Exception as e:
            messages.error(request, f"An error occurred: {str(e)}")
    return render(request, "school/time_slots.html", {
        "form": form, "school": slot.school, "mode": "edit", "slot": slot
    })

@login_required
def time_slot_delete(request, pk):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to create Time slots.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:time-slot-list')


    slot = get_object_or_404(TimeSlot, pk=pk, school=school)
    if request.method == "POST":
        slot.delete()
        messages.success(request, "Time slot deleted successfully.")
        return redirect("school:time-slot-list")

    return render(request, "school/time_slots.html", {"slot": slot})



def term_list(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access Term List.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    terms = Term.objects.filter(school=school)
    form = TermForm()
    return render(request, "school/terms.html", {"school": school, "terms": terms, "form": form})


@login_required
def create_term(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to create Term.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    form = TermForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        term = form.save(commit=False)
        term.school = school
        term.is_active = True
        term.save()

        # If new term is active → deactivate all other terms
        if term.is_active:
            Term.objects.filter(school=school).exclude(id=term.id).update(is_active=False)

        messages.success(request, "New term created successfully.")
        return redirect("school:term-list")

    return render(request, "school/terms/create_term.html", {
        "form": form,
        "school": school
    })

@login_required
def edit_term(request, term_id):
    term = get_object_or_404(Term, id=term_id)
    school = term.school

    form = TermForm(request.POST or None, instance=term)

    if request.method == "POST" and form.is_valid():
        updated_term = form.save(commit=False)
        updated_term.school = school
        updated_term.save()

        # If edited term is now active → deactivate all other terms
        if updated_term.is_active:
            Term.objects.filter(school=school).exclude(id=updated_term.id).update(is_active=False)

        messages.success(request, "Term updated successfully.")
        return redirect("school:term-list")

    return render(request, "school/terms/edit_term.html", {
        "form": form,
        "term": term,
        "school": school,
    })


@login_required
def delete_term(request, term_id):
    term = get_object_or_404(Term, id=term_id)
    school_id = term.school.id

    if request.method == "POST":
        term.delete()
        messages.success(request, "Term deleted successfully.")
        return redirect("school:term-list")

    return render(request, "school/terms/delete_term_confirm.html", {
        "term": term
    })


from django.core.paginator import Paginator


def paginate(request, queryset, per_page=10, page_param='page'):
    """Paginate a queryset with a custom GET parameter."""
    paginator = Paginator(queryset, per_page)
    page_number = request.GET.get(page_param)
    return paginator.get_page(page_number)

@login_required
def school_users(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this page.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)
    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    tab = request.GET.get('tab', 'staff')
    query = request.GET.get('q', '').strip()

    staff_qs = StaffProfile.objects.filter(school=school).select_related('user').order_by('user__first_name')
    students_qs = Student.objects.filter(school=school).select_related('user', 'grade_level', 'stream').order_by('user__first_name')
    parents_qs = Parent.objects.filter(school=school).select_related('user').order_by('user__first_name')

    if query:
        staff_qs = staff_qs.filter(
            Q(user__first_name__icontains=query) |
            Q(user__last_name__icontains=query) |
            Q(user__email__icontains=query) |
            Q(staff_id__icontains=query) |
            Q(position__icontains=query)
        )
        students_qs = students_qs.filter(
            Q(user__first_name__icontains=query) |
            Q(user__last_name__icontains=query) |
            Q(user__email__icontains=query) |
            Q(student_id__icontains=query)
        )
        parents_qs = parents_qs.filter(
            Q(user__first_name__icontains=query) |
            Q(user__last_name__icontains=query) |
            Q(user__email__icontains=query) |
            Q(parent_id__icontains=query)
        )

    staff_count = StaffProfile.objects.filter(school=school).count()
    students_count = Student.objects.filter(school=school).count()
    parents_count = Parent.objects.filter(school=school).count()

    staff_list = paginate(request, staff_qs, per_page=25, page_param='sp')
    students_list = paginate(request, students_qs, per_page=25, page_param='stp')
    parents_list = paginate(request, parents_qs, per_page=25, page_param='pp')

    staff_form = StaffCreationForm(school=school)
    student_form = StudentCreationForm(school=school)
    parent_form = ParentCreationForm(school=school)

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'create_staff':
            staff_form = StaffCreationForm(request.POST, request.FILES, school=school)
            if staff_form.is_valid():
                staff_obj, pwd = staff_form.save()
                messages.success(
                    request,
                    f'Staff "{staff_obj.user.get_full_name()}" created. Temp password: {pwd}'
                )
                return redirect(f"{request.path}?tab=staff")
            else:
                tab = 'staff'
        elif action == 'create_student':
            student_form = StudentCreationForm(request.POST, request.FILES, school=school)
            if student_form.is_valid():
                student_obj, pwd = student_form.save()
                messages.success(request, f'Student "{student_obj.user.get_full_name()}" created.')
                return redirect(f"{request.path}?tab=students")
            else:
                tab = 'students'
        elif action == 'create_parent':
            parent_form = ParentCreationForm(request.POST, request.FILES, school=school)
            if parent_form.is_valid():
                parent_obj, pwd = parent_form.save()
                messages.success(request, f'Parent "{parent_obj.user.get_full_name()}" created.')
                return redirect(f"{request.path}?tab=parents")
            else:
                tab = 'parents'

    context = {
        'school': school,
        'tab': tab,
        'query': query,
        'staff_list': staff_list,
        'students_list': students_list,
        'parents_list': parents_list,
        'staff_count': staff_count,
        'students_count': students_count,
        'parents_count': parents_count,
        'staff_form': staff_form,
        'student_form': student_form,
        'parent_form': parent_form,
    }

    return render(request, 'school/staff.html', context)


@login_required
def update_student(request, pk):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:school-users')

    student = get_object_or_404(Student, pk=pk, school=school)
    if request.method == 'POST':
        form = StudentUpdateForm(request.POST, request.FILES, instance=student, school=school)  # Fixed: school=school
        if form.is_valid():
            try:
                with transaction.atomic():
                    form.save()
                messages.success(request, f'Successfully updated Student "{student.user.get_full_name()}".')
                return redirect('school:school-users')
            except Exception as e:
                messages.error(request, f"Failed to update: {str(e)}")
        else:
            messages.error(request, "Please correct the form errors below.")
    else:
        form = StudentUpdateForm(instance=student, school=school)
    context = {'form': form, 'student': student, 'school': school}
    return render(request, 'school/student_edit.html', context)

# Similar for Parent Update
@login_required
def update_parent(request, pk):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:school-users')
    parent = get_object_or_404(Parent, pk=pk, school=school)
    if request.method == 'POST':
        form = ParentUpdateForm(request.POST, request.FILES, instance=parent, school=school)
        if form.is_valid():
            try:
                with transaction.atomic():
                    form.save()
                messages.success(request, f'Successfully updated Parent "{parent.user.get_full_name()}".')
                return redirect('school:school-users')
            except Exception as e:
                messages.error(request, f"Failed to update: {str(e)}")
        else:
            messages.error(request, "Please correct the form errors below.")
    else:
        form = ParentUpdateForm(instance=parent, school=school)
    context = {'form': form, 'parent': parent, 'school': school}
    return render(request, 'school/parent_edit.html', context)

# Similar for Staff Update
@login_required
def update_staff(request, pk):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:school-users')
    staff = get_object_or_404(StaffProfile, pk=pk, school=school)
    if request.method == 'POST':
        form = StaffUpdateForm(request.POST, request.FILES, instance=staff, school=school)
        if form.is_valid():
            try:
                with transaction.atomic():
                    form.save()
                messages.success(request, f'Successfully updated Staff "{staff.user.get_full_name()}".')
                return redirect('school:school-users')
            except Exception as e:
                messages.error(request, f"Failed to update: {str(e)}")
        else:
            messages.error(request, "Please correct the form errors below.")
    else:
        form = StaffUpdateForm(instance=staff, school=school)
    context = {'form': form, 'staff': staff, 'school': school}
    return render(request, 'school/staff_edit.html', context)

# Delete Views
@login_required
def delete_student(request, pk):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:school-users')
    
    student = get_object_or_404(Student, pk=pk, school=school)
    try:
        with transaction.atomic():
            student.user.delete()  # Cascade or handle as per model
        messages.success(request, f'Successfully deleted Student "{student.user.get_full_name()}".')
    except Exception as e:
        messages.error(request, f"Failed to delete: {str(e)}")
    return redirect('school:school-users')

@login_required
def delete_parent(request, pk):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:school-users')
    
    parent = get_object_or_404(Parent, pk=pk, school=school)
    try:
        with transaction.atomic():
            parent.user.delete()
        messages.success(request, f'Successfully deleted Parent "{parent.user.get_full_name()}".')
    except Exception as e:
        messages.error(request, f"Failed to delete: {str(e)}")
    return redirect('school:school-users')

@login_required
def delete_staff(request, pk):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:school-users')
    
    staff = get_object_or_404(StaffProfile, pk=pk, school=school)
    try:
        with transaction.atomic():
            staff.user.delete()
        messages.success(request, f'Successfully deleted Staff "{staff.user.get_full_name()}".')
    except Exception as e:
        messages.error(request, f"Failed to delete: {str(e)}")
    return redirect('school:school-users')

def is_school_admin(user):
    # adapt to your project's admin flag - either is_superuser or custom flag on user
    return user.is_active and (user.is_superuser or getattr(user, 'is_admin', False))


@login_required
@user_passes_test(is_school_admin)
def get_streams_for_grade(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)    

    grade_id = request.GET.get('grade_id')

    if not school or not grade_id:
        return JsonResponse({'streams': []})

    streams = Streams.objects.filter(grade_id=grade_id, grade__school=school, is_active=True)
    streams_data = [{'id': s.id, 'name': s.name} for s in streams]
    return JsonResponse({'streams': streams_data})

@login_required
@user_passes_test(is_school_admin)
def generate_timetable_view(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    if request.method == 'POST':
        form = GenerateTimetableForm(request.POST, school=school)
        if form.is_valid():
            scope = form.cleaned_data['scope']
            grade = form.cleaned_data.get('grade')
            stream = form.cleaned_data.get('stream')
            overwrite = form.cleaned_data.get('overwrite')

            term = Term.objects.filter(school=school, is_active=True).first()
            if not term:
                messages.error(request, "No active term defined for your school.")
                return redirect(request.META.get('HTTP_REFERER', '/'))

            # Determine streams to generate
            streams_qs = Streams.objects.filter(grade__school=school)
            if scope == 'grade':
                if not grade:
                    messages.error(request, "Please select a grade.")
                    return redirect(request.META.get('HTTP_REFERER', '/'))
                streams_qs = streams_qs.filter(grade=grade)
            elif scope == 'stream':
                if not stream:
                    messages.error(request, "Please select a stream.")
                    return redirect(request.META.get('HTTP_REFERER', '/'))
                streams_qs = streams_qs.filter(id=stream.id)

            created_total = 0
            errors = []
            for st in streams_qs:
                tt, created = Timetable.objects.get_or_create(
                    school=school,
                    grade=st.grade,
                    stream=st,
                    term=term,
                    year=term.start_date.year,
                    defaults={'start_date': term.start_date, 'end_date': term.end_date}
                )
                try:
                    lessons_created = generate_for_stream(tt, overwrite=overwrite)
                    created_total += len(lessons_created)
                except Exception as e:
                    errors.append(str(e))

            if created_total:
                messages.success(request, f"Timetable generated — {created_total} lessons created.")
            if errors:
                messages.warning(request, "Some errors: " + "; ".join(errors))

            return redirect(request.META.get('HTTP_REFERER', '/'))

    else:
        form = GenerateTimetableForm(school=request.user.school_admin_profile)

    # Fetch all timetables for the school
    timetables = Timetable.objects.filter(school=school).order_by('-term__start_date')

    return render(request, 'school/generate_timetable.html', {
        'form': form,
        'timetables': timetables,
        'school': school,
    })

@login_required
def view_timetable_week(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    # ── Focus date ───────────────────────────────────────────────────────────
    focus_date_str = request.GET.get("date")
    try:
        focus_date = datetime.datetime.strptime(focus_date_str, "%Y-%m-%d").date() if focus_date_str else timezone.localdate()
    except (ValueError, TypeError):
        focus_date = timezone.localdate()

    # ── Week range (Monday → Sunday) ─────────────────────────────────────────
    week_start = focus_date - timedelta(days=focus_date.weekday())
    week_end = week_start + timedelta(days=6)
    week_days = [week_start + timedelta(days=i) for i in range(7)]

    # ── Time slots ───────────────────────────────────────────────────────────
    time_slots = TimeSlot.objects.filter(school=school).order_by('start_time')

    # ── Lessons ──────────────────────────────────────────────────────────────
    lessons_qs = Lesson.objects.filter(
        timetable__school=school,
        timetable__term__is_active=True,
        lesson_date__range=(week_start, week_end),
    ).select_related(
        'subject',
        'teacher__user',
        'stream',
        'timetable',
    )

    # ── Streams ──────────────────────────────────────────────────────────────
    streams = Streams.objects.filter(school=school)
    streams_dict = {s.id: s for s in streams}

    # ── Build calendar safely ────────────────────────────────────────────────
    calendar = defaultdict(lambda: defaultdict(lambda: defaultdict(list)))

    for lesson in lessons_qs:
        date_str = lesson.lesson_date.strftime("%Y-%m-%d")

        # Skip lessons without time_slot or stream
        if not lesson.time_slot or not lesson.stream:
            continue

        stream_id = lesson.stream.id
        # Precompute stream name safely
        lesson.stream_name = streams_dict.get(stream_id).name if streams_dict.get(stream_id) else "(stream missing)"

        # Append lesson to calendar
        calendar[lesson.time_slot.id][date_str][stream_id].append(lesson)

    # Convert defaultdict → plain dict
    calendar_plain = {
        slot_id: {
            date_str: dict(stream_dict)
            for date_str, stream_dict in date_dict.items()
        }
        for slot_id, date_dict in calendar.items()
    }

    # ── Subjects & colors ────────────────────────────────────────────────────
    subjects = Subject.objects.filter(school=school).distinct()
    subject_colors = {s.id: getattr(s, "color", "#eee") for s in subjects}
    subjects_dict = {s.id: s for s in subjects}

    # ── Conflicts placeholder ────────────────────────────────────────────────
    conflict_lessons = set()

    # ── Context ──────────────────────────────────────────────────────────────
    context = {
        "week_days": week_days,
        "time_slots": time_slots,
        "calendar": calendar_plain,
        "subject_colors": subject_colors,
        "streams_dict": streams_dict,
        "focus_date": focus_date.strftime("%Y-%m-%d"),
        "conflict_lessons": conflict_lessons,
        "subjects": subjects_dict,
    }

    return render(request, "school/timetable_week.html", context)

@login_required
def teacher_timetable_view(request):
    if not hasattr(request.user, 'staffprofile'):
        return HttpResponseForbidden()
    teacher = request.user.staffprofile
    time_slots =  TimeSlot.objects.filter(school=teacher.school).order_by("start_time")
    date_str = request.GET.get('date')
    if date_str:
        focus_date = date.fromisoformat(date_str)
    else:
        focus_date = timezone.localdate()
    monday = focus_date - timedelta(days=focus_date.weekday())
    week_days = [monday + timedelta(days=i) for i in range(7)]

    lessons_qs = Lesson.objects.filter(teacher=teacher, date__range=(week_days[0], week_days[-1])).order_by('lesson_date')

    # Group by date
    lessons_by_date = {d: lessons_qs.filter(lesson_date=d) for d in week_days}

    context = {
        'week_days': week_days,
        'lessons_by_date': lessons_by_date,
        'time_slots': time_slots,
        'focus_date': focus_date,
    }
    return render(request, 'school/teacher_timetable.html', context)


@login_required
def student_details(request, pk):
    user = request.user
    school = get_user_school(user)
    student = get_object_or_404(Student, id=pk, school=school)
    if user.is_principal or user.is_deputy_principal or user.is_admin:
        base_tpl = 'school/base.html'
    elif user.is_teacher:
        base_tpl = 'school/teacher/base.html'
    else:
        base_tpl = 'school/base.html'
    return render(request, "school/student_details.html", {"student": student, "base_template": base_tpl})

@login_required
def edit_staff(request, staff_id):
    school = get_user_school(request.user)
    staff = get_object_or_404(StaffProfile, id=staff_id, school=school)
    form = StaffCreationForm(request.POST or None, request.FILES or None, instance=staff, school=school)

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Staff updated successfully.")
        return redirect('school:school-users')

    return render(request, "school/forms/edit_staff.html", {"form": form})

@login_required
def edit_student(request, student_id):
    school = get_user_school(request.user)
    student = get_object_or_404(Student, id=student_id, school=school)
    form = StudentCreationForm(
        request.POST or None,
        request.FILES or None,
        instance=student,
        school=school,
    )

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Student updated successfully.")
        return redirect('school:school-users')

    return render(request, "school/forms/edit_student.html", {"form": form})

@login_required
def edit_parent(request, parent_id):
    school = get_user_school(request.user)
    parent = get_object_or_404(Parent, id=parent_id, school=school)
    form = ParentCreationForm(
        request.POST or None,
        request.FILES or None,
        instance=parent,
        school=school,
    )

    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Parent updated successfully.")
        return redirect('school:school-users')

    return render(request, "school/forms/edit_parent.html", {"form": form})

@login_required
@require_POST
def delete_staff(request, staff_id):
    school = get_user_school(request.user)
    staff = get_object_or_404(StaffProfile, id=staff_id, school=school)
    staff.delete()
    messages.success(request, "Staff deleted.")
    return redirect('school:school-users')

@login_required
@require_POST
def delete_student(request, student_id):
    school = get_user_school(request.user)
    student = get_object_or_404(Student, id=student_id, school=school)
    student.delete()
    messages.success(request, "Student deleted.")
    return redirect('school:school-users')

@login_required
@require_POST
def delete_parent(request, parent_id):
    school = get_user_school(request.user)
    parent = get_object_or_404(Parent, id=parent_id, school=school)
    parent.delete()
    messages.success(request, "Parent deleted.")
    return redirect('school:school-users')

@login_required
def school_grades(request):
    # Ensure user is a school admin
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    form = GradeForm(school=school)
    stream_form = StreamForm()
    upload_form = GradeUploadForm()

    query = request.GET.get('q', '').strip()

    grades_qs = Grade.objects.filter(
        school=school,
        is_active=True
    ).order_by('name')  # ✅ sorted

    if query:
        grades_qs = grades_qs.filter(
            Q(name__icontains=query) |
            Q(code__icontains=query)
        )

    # ✅ Pagination
    paginator = Paginator(grades_qs, 10)  # 10 grades per page
    page_number = request.GET.get('page')
    grades = paginator.get_page(page_number)

    streams = Streams.objects.filter(
        school=school,
        is_active=True
    ).order_by('name')

    context = {
        'grades': grades,
        'form': form,
        'stream_form': stream_form,
        'upload_form': upload_form,
        'school': school,
        'streams': streams,
        'query': query,
    }
    return render(request, "school/grade.html", context)

@login_required
def smartid_list(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    query = request.GET.get('q', '')

    smartids = SmartID.objects.filter(
        school=school,
        is_active=True
    ).select_related(
        'profile',
        'profile__student',
        'profile__staffprofile'
    )

    if query:
        smartids = smartids.filter(
            Q(profile__first_name__icontains=query) |
            Q(profile__last_name__icontains=query) |
            Q(card_id__icontains=query) |
            Q(user_f18_id__icontains=query)
        )

    form = SmartIDForm(school=school)

    return render(request, 'school/smartid_list.html', {
        'smartids': smartids,
        'form': form,
        'school': school,
        'query': query,
    })


@login_required
def smartid_create(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    if request.method != 'POST':
        return redirect('school:smartid-list')
    
    form = SmartIDForm(request.POST, school=school)
    if form.is_valid():
        smartid = form.save(commit=False)
        smartid.school = school
        smartid.save()
        messages.success(request, f'Smart ID "{smartid.card_id}" created successfully.')
        logger.info(f"Created Smart ID {smartid.card_id} for school {school.name} by {request.user.email}.")
        return redirect('school:smartid-list')
    else:
        # Add form errors to messages
        for field, error_list in form.errors.items():
            if field == '__all__':
                for error in error_list:
                    messages.error(request, error)
            else:
                label = form.fields[field].label
                for error in error_list:
                    messages.error(request, f"{label}: {error}")
        return redirect('school:smartid-list')

@login_required
def smartid_edit(request, pk):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    if request.method != 'POST':
        return redirect('school:smartid-list')
    
    smartid = get_object_or_404(SmartID, pk=pk, school=school)
    form = SmartIDForm(request.POST, instance=smartid, school=school)
    if form.is_valid():
        form.save()
        messages.success(request, f'Smart ID "{smartid.card_id}" updated successfully.')
        logger.info(f"Updated Smart ID {smartid.card_id} for school {school.name} by {request.user.email}.")
        return redirect('school:smartid-list')
    else:
        # Add form errors to messages
        for field, error_list in form.errors.items():
            if field == '__all__':
                for error in error_list:
                    messages.error(request, error)
            else:
                label = form.fields[field].label
                for error in error_list:
                    messages.error(request, f"{label}: {error}")
        return redirect('school:smartid-list')

@login_required
def smartid_delete(request, pk):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    if request.method != 'POST':
        messages.warning(request, "Use the delete button in the list to confirm.")
        return redirect('school:smartid-list')
    
    smartid = get_object_or_404(SmartID, pk=pk, school=school)
    smartid.is_active = False  # Soft delete
    smartid.save()
    messages.success(request, f'Smart ID "{smartid.card_id}" deactivated.')
    logger.info(f"Deactivated Smart ID {smartid.card_id} for school {school.name} by {request.user.email}.")
    return redirect('school:smartid-list')

# -------------------------------SCAN LOGS VIEW-------------------------------
@login_required
def scan_logs_view(request):
    """
    Display all scan logs with pagination and simple search filters.
    """
    search = request.GET.get('q', '')
    school = getattr(request.user, 'school', None)

    logs = ScanLog.objects.all()
    # Optional: limit logs by user's school
    if school:
        logs = logs.filter(school=school)

    if search:
        logs = logs.filter(school=school)

    paginator = Paginator(logs, 25)  # Show 25 logs per page
    page = request.GET.get('page')
    logs_page = paginator.get_page(page)

    context = {
        'logs': logs_page,
        'search': search,
    }
    return render(request, 'school/scan_logs.html', context)

@login_required
def grade_create(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    if request.method != 'POST':
        return redirect('school:school-grades')

    form = GradeForm(request.POST, school=school)
    if form.is_valid():
        grade = form.save(commit=False)
        grade.school = school
        grade.is_active = True
        grade.save()
        messages.success(request, f'Grade "{grade.name}" created successfully.')
        logger.info(f"Created grade {grade.name} for school {school.name} by {request.user.email}.")
        return redirect('school:school-grades')
    else:
        # Add form errors to messages
        for field, error_list in form.errors.items():
            if field == '__all__':
                for error in error_list:
                    messages.error(request, error)
            else:
                label = form.fields[field].label
                for error in error_list:
                    messages.error(request, f"{label}: {error}")
        return redirect('school:school-grades')


@login_required
def upload_grade_file(request):
    # Ensure user has a school
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    if request.method == 'POST':
        form = GradeUploadForm(request.POST, request.FILES)
        if form.is_valid():
            upload = form.save(commit=False)
            upload.upload_file_category = 'grade'
            upload.uploaded_by = request.user
            upload.school = school
            upload.save()

            # Process Excel file
            file_path = upload.file.path
            ext = os.path.splitext(file_path)[1].lower()
            if ext not in ['.xls', '.xlsx']:
                messages.error(request, "Invalid file type. Only Excel files allowed.")
                upload.delete()
                return redirect('upload_grade_file')

            try:
                df = pd.read_excel(file_path)  # expects columns: name, description, code, capacity
            except Exception as e:
                messages.error(request, f"Failed to read Excel file: {str(e)}")
                upload.delete()
                return redirect('upload_grade_file')

            created_count = 0
            for index, row in df.iterrows():
                # Skip rows without name or code
                if pd.isna(row.get('name')) or pd.isna(row.get('code')):
                    continue

                grade, created = Grade.objects.get_or_create(
                    school=school,
                    code=row.get('code'),
                    defaults={
                        'name': row.get('name'),
                        'description': row.get('description', ''),
                        'capacity': int(row.get('capacity', 50)),
                    }
                )
                if created:
                    created_count += 1

            messages.success(request, f"File uploaded successfully. {created_count} grades created.")
            return redirect('school:school-grades')

        else:
            messages.error(request, "Please fix the errors below.")
    else:
        form = GradeUploadForm()

    uploaded_files = UploadedFile.objects.filter(
        upload_file_category='grade',
        school=school
    ).order_by('-uploaded_at')

    context = {
        'form': form,
        'uploaded_files': uploaded_files,
    }
    return redirect('school:school-grades')



@login_required
def grade_edit(request, pk):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    if request.method != 'POST':
        return redirect('school:school-grades')

    grade = get_object_or_404(Grade, pk=pk, school=school)
    form = GradeForm(request.POST, instance=grade, school=school)
    if form.is_valid():
        form.save()
        messages.success(request, f'Grade "{grade.name}" updated successfully.')
        logger.info(f"Updated grade {grade.name} for school {school.name} by {request.user.email}.")
        return redirect('school:school-grades')
    else:
        # Add form errors to messages
        for field, error_list in form.errors.items():
            if field == '__all__':
                for error in error_list:
                    messages.error(request, error)
            else:
                label = form.fields[field].label
                for error in error_list:
                    messages.error(request, f"{label}: {error}")
        return redirect('school:school-grades')


@login_required
def grade_delete(request, pk):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    if request.method != 'POST':
        messages.warning(request, "Use the delete button in the list to confirm.")
        return redirect('school:school-grades')

    grade = get_object_or_404(Grade, pk=pk, school=school)
    grade.is_active = False  # Soft delete
    grade.save()
    messages.success(request, f'Grade "{grade.name}" deactivated.')
    logger.info(f"Deactivated grade {grade.name} for school {school.name} by {request.user.email}.")
    return redirect('school:school-grades')



@login_required
def parent_list_create(request):
    """
    List parents with search/filter.
    Handles creation via POST (from add modal).
    """
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    if request.method == 'POST':
        form = ParentCreationForm(request.POST, request.FILES, school=school)
        if form.is_valid():
            parent, password = form.save()
            messages.success(
                request,
                f'Successfully created parent "{parent.user.get_full_name()}" ({parent.parent_id}). '
                f'Temporary password: <strong>{password}</strong>. '
                f'Email sent: {"Yes" if form.cleaned_data["send_email"] else "No"}.'
            )
            logger.info(f"Created parent {parent.parent_id} for school {school.name}.")
            return redirect('school:parent_list_create')  # Back to list
        # Errors: Re-render list with bound form
    else:
        form = ParentCreationForm(school=school)

    # List parents
    query = request.GET.get('q', '')
    parents = Parent.objects.filter(school=school).select_related('user').order_by('-user__created_at')
    if query:
        parents = parents.filter(
            Q(user__first_name__icontains=query) |
            Q(user__last_name__icontains=query) |
            Q(parent_id__icontains=query) |
            Q(phone__icontains=query) |
            Q(user__email__icontains=query)
        )

    context = {
        'parents': parents,
        'form': form,  # For add modal
        'school': school,
        'query': query,
    }
    return render(request, 'school/parents.html', context)


@login_required
def parent_edit(request, pk):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    

    parent = get_object_or_404(Parent, pk=pk, school=school)

    if request.method == 'POST':
        form = ParentEditForm(request.POST, request.FILES, instance=parent)
        if form.is_valid():
            form.save()
            messages.success(
                request,
                f'Successfully updated parent "{parent.user.get_full_name()}".'
            )
            logger.info(f"Updated parent {parent.parent_id}.")
            return redirect('school:parent_list_create')
        else:
            messages.error(request, "Please correct the form errors.")
            # Still redirect back to list (modal-based editing)
            return redirect('school:parent_list_create')

    # If GET, editing is via modal only, so redirect to list
    messages.info(request, "Use the edit button in the list to modify.")
    return redirect('school:parent_list_create')




@login_required
def parent_delete(request, pk):
    """
    Delete via modal form POST.
    Redirects back to list on success.
    """
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    parent = get_object_or_404(Parent, pk=pk, school=school)

    if request.method == 'POST':
        user = parent.user
        user.is_active = False
        user.save()
        parent.delete()
        messages.success(request, f'Parent "{parent.user.get_full_name()}" ({parent.parent_id}) has been deleted.')
        logger.info(f"Deleted parent {parent.parent_id}.")
        return redirect('school:parent_list_create')  # Back to list (modal closes)
    
    # For GET (direct access): Redirect to list
    messages.warning(request, "Use the delete button in the list to confirm.")
    return redirect('school:parent_list_create')

@login_required
def staff_list_create(request):
    """
    List staff with search/filter.
    Handles creation via POST (from add modal).
    """
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    if request.method == 'POST':
        form = StaffCreationForm(request.POST, request.FILES, school=school)
        if form.is_valid():
            staff, password = form.save()
            messages.success(
                request,
                f'Successfully created staff "{staff.user.get_full_name()}" ({staff.staff_id}). '
                f'Position: {staff.get_position_display()}. '
                f'Temporary password: <strong>{password}</strong>. '
                f'Email sent: {"Yes" if form.cleaned_data["send_email"] else "No"}.'
            )
            logger.info(f"Created staff {staff.staff_id} for school {school.name}.")
            return redirect('school:staff_list_create')
        else:
            messages.error(request, "Please correct the form errors below.")
    else:
        form = StaffCreationForm(school=school)

    # List staff
    query = request.GET.get('q', '')
    staff_list = StaffProfile.objects.filter(school=school).select_related('user').order_by('-user__created_at')
    if query:
        staff_list = staff_list.filter(
            Q(user__first_name__icontains=query) |
            Q(user__last_name__icontains=query) |
            Q(staff_id__icontains=query) |
            Q(phone__icontains=query) |
            Q(user__email__icontains=query) |
            Q(position__icontains=query)
        )

    context = {
        'staff_list': staff_list,
        'form': form,
        'school': school,
        'query': query,
    }
    return render(request, 'school/staff.html', context)


@login_required
def staff_edit(request, pk):
    """
    Edit via modal form POST.
    Redirects back to list on success/error.
    """
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    staff = get_object_or_404(StaffProfile, pk=pk, school=school)

    if request.method == 'POST':
        form = StaffEditForm(request.POST, request.FILES, instance=staff)
        if form.is_valid():
            form.save()
            messages.success(
                request,
                f'Successfully updated staff "{staff.user.get_full_name()}". '
                f'Email sent: {"Yes" if form.cleaned_data["send_email"] else "No"}. '
                f'Password reset: {"Yes" if form.cleaned_data["reset_password"] else "No"}.'
            )
            logger.info(f"Updated staff {staff.staff_id}.")
            return redirect('school:staff_list_create')
        else:
            messages.error(request, "Please correct the form errors.")
    # For GET: Redirect to list (modal-driven)
    messages.info(request, "Use the edit button in the list to modify.")
    return redirect('school:staff_list_create')


@login_required
def staff_delete(request, pk):
    """
    Delete via modal form POST.
    Redirects back to list on success.
    """
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    staff = get_object_or_404(StaffProfile, pk=pk, school=school)

    if request.method == 'POST':
        user = staff.user
        user.is_active = False
        user.save()
        staff.delete()
        messages.success(request, f'Staff "{staff.user.get_full_name()}" ({staff.staff_id}) has been deleted.')
        logger.info(f"Deleted staff {staff.staff_id}.")
        return redirect('school:staff_list_create')
    
    # For GET: Redirect to list
    messages.warning(request, "Use the delete button in the list to confirm.")
    return redirect('school:staff_list_create')


@login_required
def student_list_create(request):
    """
    List students with search/filter.
    Handles creation via POST (from add modal).
    """
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    if request.method == 'POST':
        form = StudentCreationForm(request.POST, request.FILES, school=school)
        if form.is_valid():
            student, password = form.save(commit=False)
            student.school = school
            student.save()
            messages.success(
                request,
                f'Successfully created student "{student.user.get_full_name()}" ({student.student_id}). '
                f'Grade: {student.grade_level}. '
                f'Temporary password: <strong>{password}</strong>. '
                f'Email sent: {"Yes" if form.cleaned_data["send_email"] else "No"}.'
            )
            logger.info(f"Created student {student.student_id} for school {school.name}.")
            return redirect('school:student_list_create')
        else:
            messages.error(request, "Please correct the form errors below.")
    else:
        form = StudentCreationForm(school=school)

    # List students
    query = request.GET.get('q', '')
    students = Student.objects.filter(school=school).select_related('user', 'grade_level').order_by('-user__created_at')
    if query:
        students = students.filter(
            Q(user__first_name__icontains=query) |
            Q(user__last_name__icontains=query) |
            Q(student_id__icontains=query) |
            Q(user__phone_number__icontains=query) |
            Q(user__email__icontains=query) |
            Q(grade_level__name__icontains=query)
        )

    context = {
        'students': students,
        'form': form,
        'school': school,
        'query': query,
    }
    return render(request, 'school/students.html', context)


@login_required
def student_edit(request, pk):
    """
    Edit via modal form POST.
    Redirects back to list on success/error.
    """
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    student = get_object_or_404(Student, pk=pk, school=school)

    if request.method == 'POST':
        form = StudentEditForm(request.POST, request.FILES, instance=student)
        if form.is_valid():
            form.save()
            messages.success(
                request,
                f'Successfully updated student "{student.user.get_full_name()}". '
                f'Email sent: {"Yes" if form.cleaned_data["send_email"] else "No"}. '
                f'Password reset: {"Yes" if form.cleaned_data["reset_password"] else "No"}.'
            )
            logger.info(f"Updated student {student.student_id}.")
            return redirect('school:student_list_create')
        else:
            messages.error(request, "Please correct the form errors.")
    # For GET: Redirect to list (modal-driven)
    messages.info(request, "Use the edit button in the list to modify.")
    return redirect('school:student_list_create')


@login_required
def student_delete(request, pk):
    """
    Delete via modal form POST.
    Redirects back to list on success.
    """
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    student = get_object_or_404(Student, pk=pk, school=school)

    if request.method == 'POST':
        user = student.user
        user.is_active = False
        user.save()
        student.delete()
        messages.success(request, f'Student "{student.user.get_full_name()}" ({student.student_id}) has been deleted.')
        logger.info(f"Deleted student {student.student_id}.")
        return redirect('school:student_list_create')
    
    # For GET: Redirect to list
    messages.warning(request, "Use the delete button in the list to confirm.")
    return redirect('school:student_list_create')




@login_required
def school_exams(request):
    """List exam sessions for the school."""
    user = request.user
    school = get_user_school(user)
    if not school:
        messages.error(request, "School not found.")
        return redirect('school:dashboard')

    sessions = ExamSession.objects.filter(school=school).select_related('grade', 'term').order_by('-year', '-created_at')
    paginator = Paginator(sessions, 15)
    page = paginator.get_page(request.GET.get('page'))

    if user.is_principal or user.is_deputy_principal or user.is_admin:
        base_tpl = 'school/base.html'
    elif user.is_teacher:
        base_tpl = 'school/teacher/base.html'
    else:
        base_tpl = 'school/base.html'

    return render(request, 'school/exams/session_list.html', {
        'school': school, 'page_obj': page, 'base_template': base_tpl,
    })


@login_required
def exam_session_create(request):
    user = request.user
    school = get_user_school(user)
    if not school:
        return redirect('school:dashboard')
    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "Only principals, deputy principals, or administrators can create exam sessions.")
        return redirect('school:school-exams')

    form = ExamSessionForm(school, request.POST or None)
    if request.method == 'POST' and form.is_valid():
        session = form.save(commit=False)
        session.school = school
        try:
            session.created_by = user.staffprofile
        except Exception:
            pass
        session.save()
        messages.success(request, f'Exam session "{session.name}" created.')
        return redirect('school:exam-session-detail', pk=session.pk)
    return render(request, 'school/exams/session_form.html', {'form': form, 'school': school})


@login_required
def exam_session_detail(request, pk):
    user = request.user
    school = get_user_school(user)
    base_template = 'school/teacher/base.html' if user.is_teacher else 'school/base.html'
    session = get_object_or_404(ExamSession, pk=pk, school=school)
    streams = Streams.objects.filter(grade=session.grade, school=school)
    subjects = Subject.objects.filter(school=school, grade=session.grade, is_active=True)
    return render(request, 'school/exams/session_detail.html', {
        'session': session, 'school': school, 'streams': streams, 'subjects': subjects,
        'base_template': base_template,
    })


@login_required
def exam_result_entry(request, session_pk, stream_pk, subject_pk):
    """Teacher enters scores for all students in a stream for one subject."""
    user = request.user
    school = get_user_school(user)
    session = get_object_or_404(ExamSession, pk=session_pk, school=school)
    stream = get_object_or_404(Streams, pk=stream_pk, school=school, grade=session.grade)
    subject = get_object_or_404(Subject, pk=subject_pk, school=school)

    students = Student.objects.filter(school=school, stream=stream, is_active=True).select_related('user').order_by('user__last_name', 'user__first_name')

    # Build or fetch existing ExamResult objects for each student
    existing = {
        r.student_id: r
        for r in ExamResult.objects.filter(session=session, subject=subject, stream=stream)
    }

    if request.method == 'POST':
        errors = []
        saved = 0
        try:
            staff = user.staffprofile
        except Exception:
            staff = None

        for student in students:
            prefix = f'student_{student.pk}'

            def _val(key):
                raw = request.POST.get(f'{prefix}_{key}', '').strip()
                return float(raw) if raw else None

            cat = _val('cat')
            assignment = _val('assignment')
            assessment = _val('assessment')
            exam = _val('exam')

            # Validate before touching the database
            checks = [
                (cat, session.cat_out_of, 'CAT'),
                (assignment, session.assignment_out_of, 'Assignment'),
                (assessment, session.assessment_out_of, 'Assessment'),
                (exam, session.exam_out_of, 'Exam'),
            ]
            ok = True
            for score, max_score, label in checks:
                if score is not None and score > max_score:
                    errors.append(f"{student.user.get_full_name()}: {label} score {score} exceeds max {max_score}.")
                    ok = False

            if not ok:
                continue

            result, _ = ExamResult.objects.get_or_create(
                session=session, student=student, subject=subject,
                defaults={'stream': stream, 'school': school, 'entered_by': staff},
            )
            result.cat_score = cat
            result.assignment_score = assignment
            result.assessment_score = assessment
            result.exam_score = exam
            result.stream = stream
            result.school = school
            if staff:
                result.entered_by = staff
            result.save()
            saved += 1

        if errors:
            for e in errors:
                messages.error(request, e)
        if saved:
            messages.success(request, f'{saved} result(s) saved.')
        return redirect('school:exam-result-entry', session_pk=session_pk, stream_pk=stream_pk, subject_pk=subject_pk)

    rows = [{'student': s, 'result': existing.get(s.pk)} for s in students]
    return render(request, 'school/exams/result_entry.html', {
        'session': session, 'stream': stream, 'subject': subject,
        'rows': rows, 'school': school,
    })


@login_required
def _run_exam_upload_job(job_id):
    """Background thread: process an ExamUploadJob."""
    import openpyxl
    import threading
    from django.db import connection as _conn
    # Each thread needs its own DB connection
    _conn.close()

    from .models import ExamUploadJob, ExamResult, Student
    try:
        job = ExamUploadJob.objects.get(pk=job_id)
        job.status = ExamUploadJob.STATUS_PROCESSING
        job.save(update_fields=['status'])

        wb = openpyxl.load_workbook(job.file_path, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]

        def col(name):
            return headers.index(name) if name in headers else None

        idx_id   = col('student_id')
        idx_cat  = col('cat')
        idx_asgn = col('assignment')
        idx_asmt = col('assessment')
        idx_exam = col('exam')

        if idx_id is None:
            job.status = ExamUploadJob.STATUS_FAILED
            job.error  = 'Column "student_id" not found in file.'
            job.finished_at = timezone.now()
            job.save(update_fields=['status', 'error', 'finished_at'])
            return

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        data_rows = [r for r in rows if r[idx_id]]
        job.total_rows = len(data_rows)
        job.save(update_fields=['total_rows'])

        try:
            staff = job.uploaded_by.staffprofile
        except Exception:
            staff = None

        saved = skipped = 0
        for i, row in enumerate(data_rows):
            sid = str(row[idx_id]).strip()
            try:
                student = Student.objects.get(student_id=sid, school=job.school)
            except Student.DoesNotExist:
                skipped += 1
            else:
                def _v(idx, _row=row):
                    if idx is None or _row[idx] is None:
                        return None
                    try:
                        return float(_row[idx])
                    except (ValueError, TypeError):
                        return None

                result, _ = ExamResult.objects.get_or_create(
                    session=job.session, student=student, subject=job.subject,
                    defaults={'stream': job.stream, 'school': job.school, 'entered_by': staff},
                )
                result.cat_score        = _v(idx_cat)
                result.assignment_score = _v(idx_asgn)
                result.assessment_score = _v(idx_asmt)
                result.exam_score       = _v(idx_exam)
                result.stream    = job.stream
                result.school    = job.school
                if staff:
                    result.entered_by = staff
                result.save()
                saved += 1

            job.processed = i + 1
            job.saved     = saved
            job.skipped   = skipped
            job.save(update_fields=['processed', 'saved', 'skipped'])

        job.status      = ExamUploadJob.STATUS_DONE
        job.finished_at = timezone.now()
        job.save(update_fields=['status', 'finished_at'])

    except Exception as exc:
        try:
            job = ExamUploadJob.objects.get(pk=job_id)
            job.status      = ExamUploadJob.STATUS_FAILED
            job.error       = str(exc)
            job.finished_at = timezone.now()
            job.save(update_fields=['status', 'error', 'finished_at'])
        except Exception:
            pass
    finally:
        try:
            import os
            if os.path.exists(job.file_path):
                os.remove(job.file_path)
        except Exception:
            pass
        _conn.close()


def exam_result_upload(request, session_pk):
    """Bulk upload scores from Excel — processed in a background thread."""
    import threading, os, tempfile
    user = request.user
    school = get_user_school(user)
    session = get_object_or_404(ExamSession, pk=session_pk, school=school)
    form = ExamUploadForm(session, request.POST or None, request.FILES or None)

    if request.method == 'POST' and form.is_valid():
        stream  = form.cleaned_data['stream']
        subject = form.cleaned_data['subject']
        f       = form.cleaned_data['file']

        try:
            _validate_excel_upload(f)
        except ValueError as ve:
            messages.error(request, str(ve))
            return redirect('school:exam-result-upload', session_pk=session_pk)

        # Save file to a temp path the background thread can read
        suffix = os.path.splitext(f.name)[1] or '.xlsx'
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
        for chunk in f.chunks():
            tmp.write(chunk)
        tmp.close()

        from .models import ExamUploadJob
        job = ExamUploadJob.objects.create(
            session     = session,
            school      = school,
            stream      = stream,
            subject     = subject,
            uploaded_by = user,
            file_path   = tmp.name,
        )

        t = threading.Thread(target=_run_exam_upload_job, args=(str(job.pk),), daemon=True)
        t.start()

        return redirect('school:exam-upload-progress', job_pk=job.pk)

    base_template = 'school/teacher/base.html' if request.user.is_teacher else 'school/base.html'
    return render(request, 'school/exams/upload.html', {
        'form': form, 'session': session, 'school': school, 'base_template': base_template,
    })


@login_required
def exam_upload_progress(request, job_pk):
    """Page that polls the upload job status."""
    from .models import ExamUploadJob
    school = get_user_school(request.user)
    job = get_object_or_404(ExamUploadJob, pk=job_pk, school=school)
    return render(request, 'school/exams/upload_progress.html', {'job': job})


@login_required
def exam_upload_status(request, job_pk):
    """JSON endpoint polled by the progress page."""
    import json as _json
    from .models import ExamUploadJob
    from django.http import JsonResponse
    school = get_user_school(request.user)
    job = get_object_or_404(ExamUploadJob, pk=job_pk, school=school)
    return JsonResponse({
        'status':      job.status,
        'total_rows':  job.total_rows,
        'processed':   job.processed,
        'saved':       job.saved,
        'skipped':     job.skipped,
        'progress':    job.progress_pct,
        'error':       job.error,
        'session_pk':  job.session_id,
    })


def _compute_student_totals(session, students, stream=None):
    """Return list of dicts {student, total, subjects_count} sorted by total desc."""
    qs = ExamResult.objects.filter(session=session, student__in=students)
    if stream:
        qs = qs.filter(stream=stream)
    from collections import defaultdict
    totals = defaultdict(float)
    counts = defaultdict(int)
    for r in qs:
        t = r.total
        if t is not None:
            totals[r.student_id] += t
            counts[r.student_id] += 1
    result = []
    for student in students:
        result.append({
            'student': student,
            'total': round(totals[student.pk], 2) if student.pk in totals else None,
            'subjects_count': counts[student.pk],
        })
    result.sort(key=lambda x: (x['total'] is None, -(x['total'] or 0)))
    for i, row in enumerate(result):
        row['position'] = i + 1 if row['total'] is not None else '–'
    return result


@login_required
def exam_ranking_stream(request, session_pk, stream_pk):
    """Stream-level student ranking."""
    user = request.user
    school = get_user_school(user)
    base_template = 'school/teacher/base.html' if user.is_teacher else 'school/base.html'
    session = get_object_or_404(ExamSession, pk=session_pk, school=school)
    stream = get_object_or_404(Streams, pk=stream_pk, school=school)
    students = Student.objects.filter(school=school, stream=stream, is_active=True).select_related('user')
    ranking = _compute_student_totals(session, students, stream=stream)
    subjects = Subject.objects.filter(school=school, grade=session.grade, is_active=True)
    return render(request, 'school/exams/ranking_stream.html', {
        'session': session, 'stream': stream, 'ranking': ranking, 'subjects': subjects, 'school': school,
        'base_template': base_template,
    })


@login_required
def exam_ranking_grade(request, session_pk):
    """Grade-level student ranking across all streams."""
    user = request.user
    school = get_user_school(user)
    base_template = 'school/teacher/base.html' if user.is_teacher else 'school/base.html'
    session = get_object_or_404(ExamSession, pk=session_pk, school=school)
    students = Student.objects.filter(school=school, grade_level=session.grade, is_active=True).select_related('user', 'stream')
    ranking = _compute_student_totals(session, students)
    streams = Streams.objects.filter(grade=session.grade, school=school)
    return render(request, 'school/exams/ranking_grade.html', {
        'session': session, 'ranking': ranking, 'streams': streams, 'school': school,
        'base_template': base_template,
    })


@login_required
def exam_subject_performance(request, session_pk):
    """Subject-level performance report: avg score, pass rate per subject."""
    user = request.user
    school = get_user_school(user)
    base_template = 'school/teacher/base.html' if user.is_teacher else 'school/base.html'
    session = get_object_or_404(ExamSession, pk=session_pk, school=school)
    subjects = Subject.objects.filter(school=school, grade=session.grade, is_active=True)

    data = []
    for subject in subjects:
        results = ExamResult.objects.filter(session=session, subject=subject)
        totals = [r.total for r in results if r.total is not None]
        if not totals:
            data.append({'subject': subject, 'count': 0, 'avg': None, 'highest': None, 'lowest': None, 'pass_rate': None})
            continue
        max_t = session.total_marks
        percentages = [(t / max_t * 100) for t in totals]
        pass_count = sum(1 for p in percentages if p >= 50)
        data.append({
            'subject': subject,
            'count': len(totals),
            'avg': round(sum(percentages) / len(percentages), 1),
            'highest': round(max(percentages), 1),
            'lowest': round(min(percentages), 1),
            'pass_rate': round(pass_count / len(totals) * 100, 1),
        })
    data.sort(key=lambda x: (x['avg'] is None, -(x['avg'] or 0)))

    return render(request, 'school/exams/subject_performance.html', {
        'session': session, 'data': data, 'school': school,
        'base_template': base_template,
    })


@login_required
def exam_publish(request, session_pk):
    user = request.user
    school = get_user_school(user)
    if not (user.is_admin or user.is_principal):
        messages.error(request, "Only principals can publish results.")
        return redirect('school:exam-session-detail', pk=session_pk)
    session = get_object_or_404(ExamSession, pk=session_pk, school=school)
    session.is_published = not session.is_published
    session.save()
    status = 'published' if session.is_published else 'unpublished'
    messages.success(request, f'Results {status}.')
    return redirect('school:exam-session-detail', pk=session_pk)


@login_required
def report_slip_html(request, session_pk, student_pk):
    """HTML preview of a student's report slip."""
    user = request.user

    # Resolve student (no school filter yet — access check below determines school scope)
    student = get_object_or_404(Student, pk=student_pk)
    school = student.school

    # Access control: staff/admin must belong to the same school
    is_staff_or_admin = False
    if user.is_admin or user.is_principal or user.is_deputy_principal or getattr(user, 'is_teacher', False):
        user_school = get_user_school(user)
        is_staff_or_admin = (user_school is not None and user_school.pk == school.pk)

    is_parent_of = False
    if getattr(user, 'is_parent', False):
        try:
            is_parent_of = user.parent.children.filter(pk=student_pk).exists()
        except Exception:
            pass
    is_own = (getattr(user, 'is_student', False) and hasattr(user, 'student') and user.student.pk == student_pk)

    if not (is_staff_or_admin or is_parent_of or is_own):
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    session = get_object_or_404(ExamSession, pk=session_pk, school=school)

    results = ExamResult.objects.filter(session=session, student=student).select_related('subject').order_by('subject__name')

    # Stream ranking
    stream_students = Student.objects.filter(school=school, stream=student.stream, is_active=True).select_related('user')
    stream_ranking = _compute_student_totals(session, stream_students, stream=student.stream)
    stream_pos = next((r['position'] for r in stream_ranking if r['student'].pk == student_pk), '–')
    stream_total_students = len([r for r in stream_ranking if r['total'] is not None])

    # Grade ranking
    grade_students = Student.objects.filter(school=school, grade_level=session.grade, is_active=True).select_related('user')
    grade_ranking = _compute_student_totals(session, grade_students)
    grade_pos = next((r['position'] for r in grade_ranking if r['student'].pk == student_pk), '–')
    grade_total_students = len([r for r in grade_ranking if r['total'] is not None])

    overall_total = sum(r.total for r in results if r.total is not None)
    overall_pct = round((overall_total / (session.total_marks * len(results))) * 100, 1) if results and session.total_marks else None
    overall_band = cbc_grade_band(overall_pct) if overall_pct is not None else '–'

    return render(request, 'school/exams/report_slip.html', {
        'session': session, 'student': student, 'results': results,
        'school': school, 'stream_pos': stream_pos, 'grade_pos': grade_pos,
        'stream_total': stream_total_students, 'grade_total': grade_total_students,
        'overall_total': round(overall_total, 2), 'overall_pct': overall_pct, 'overall_band': overall_band,
    })


@login_required
def report_slip_pdf(request, session_pk, student_pk):
    """Download PDF report slip."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    user = request.user
    student = get_object_or_404(Student, pk=student_pk)
    school = student.school

    # Access control: same-school staff, the student's parent, or the student
    is_staff = False
    if user.is_admin or user.is_principal or user.is_deputy_principal or getattr(user, 'is_teacher', False):
        user_school = get_user_school(user)
        is_staff = (user_school is not None and user_school.pk == school.pk)
    is_parent_of = False
    if getattr(user, 'is_parent', False):
        try:
            is_parent_of = user.parent.children.filter(pk=student_pk).exists()
        except Exception:
            pass
    is_own = (getattr(user, 'is_student', False) and hasattr(user, 'student') and user.student.pk == student_pk)
    if not (is_staff or is_parent_of or is_own):
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    session = get_object_or_404(ExamSession, pk=session_pk, school=school)
    results = ExamResult.objects.filter(session=session, student=student).select_related('subject').order_by('subject__name')

    stream_students = Student.objects.filter(school=school, stream=student.stream, is_active=True).select_related('user')
    stream_ranking = _compute_student_totals(session, stream_students, stream=student.stream)
    stream_pos = next((r['position'] for r in stream_ranking if r['student'].pk == student_pk), '–')
    stream_total = len([r for r in stream_ranking if r['total'] is not None])

    grade_students = Student.objects.filter(school=school, grade_level=session.grade, is_active=True).select_related('user')
    grade_ranking = _compute_student_totals(session, grade_students)
    grade_pos = next((r['position'] for r in grade_ranking if r['student'].pk == student_pk), '–')
    grade_total = len([r for r in grade_ranking if r['total'] is not None])

    overall_total = sum(r.total for r in results if r.total is not None)
    overall_pct = round((overall_total / (session.total_marks * len(results))) * 100, 1) if results and session.total_marks else None
    overall_band = cbc_grade_band(overall_pct) if overall_pct is not None else '–'

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()

    green = colors.HexColor('#0b7a2a')
    red = colors.HexColor('#bb0a21')
    dark = colors.HexColor('#111111')
    light_green = colors.HexColor('#e8f5e9')

    h1 = ParagraphStyle('H1', fontSize=16, fontName='Helvetica-Bold', textColor=green, alignment=TA_CENTER, spaceAfter=2)
    h2 = ParagraphStyle('H2', fontSize=11, fontName='Helvetica-Bold', textColor=dark, alignment=TA_CENTER, spaceAfter=2)
    sub = ParagraphStyle('Sub', fontSize=9, fontName='Helvetica', textColor=colors.grey, alignment=TA_CENTER, spaceAfter=6)
    normal = ParagraphStyle('N', fontSize=9, fontName='Helvetica')
    bold9 = ParagraphStyle('B9', fontSize=9, fontName='Helvetica-Bold')
    center9 = ParagraphStyle('C9', fontSize=9, fontName='Helvetica', alignment=TA_CENTER)
    right9 = ParagraphStyle('R9', fontSize=9, fontName='Helvetica', alignment=TA_RIGHT)

    elements = []

    # Header
    elements.append(Paragraph(school.name.upper(), h1))
    elements.append(Paragraph(f"Code: {school.code}  |  {school.address}", sub))
    elements.append(Paragraph(f"Tel: {school.contact_phone}  |  Email: {school.contact_email}", sub))
    elements.append(HRFlowable(width='100%', thickness=2, color=green, spaceAfter=6))
    elements.append(Paragraph(f"STUDENT REPORT SLIP – {session.name.upper()}", h2))
    elements.append(HRFlowable(width='100%', thickness=1, color=red, spaceAfter=8))

    # Student info
    info_data = [
        ['Name:', student.user.get_full_name(), 'Admission No:', student.student_id],
        ['Grade:', session.grade.name, 'Stream:', student.stream.name if student.stream else '–'],
        ['Term:', session.term.name if session.term else '–', 'Year:', str(session.year)],
    ]
    info_table = Table(info_data, colWidths=[2.5*cm, 6*cm, 3*cm, 6*cm])
    info_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 8))

    # Scores table
    hdr = ['Subject', f'CAT\n/{session.cat_out_of:.0f}', f'Asgn\n/{session.assignment_out_of:.0f}',
           f'Asmt\n/{session.assessment_out_of:.0f}', f'Exam\n/{session.exam_out_of:.0f}',
           f'Total\n/{session.total_marks:.0f}', '%', 'Grade']
    table_data = [hdr]
    for r in results:
        band = r.grade_band
        band_colors = {'EE': '#1b5e20', 'ME': '#2e7d32', 'AE': '#f57f17', 'BE': '#b71c1c'}
        table_data.append([
            r.subject.name,
            f'{r.cat_score:.1f}' if r.cat_score is not None else '–',
            f'{r.assignment_score:.1f}' if r.assignment_score is not None else '–',
            f'{r.assessment_score:.1f}' if r.assessment_score is not None else '–',
            f'{r.exam_score:.1f}' if r.exam_score is not None else '–',
            f'{r.total:.1f}' if r.total is not None else '–',
            f'{r.percentage:.1f}%' if r.percentage is not None else '–',
            band,
        ])

    col_widths = [4.5*cm, 1.6*cm, 1.6*cm, 1.6*cm, 1.6*cm, 2*cm, 1.8*cm, 1.8*cm]
    score_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), green),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, light_green]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ])
    score_table.setStyle(style)
    elements.append(score_table)
    elements.append(Spacer(1, 10))

    # Summary
    summary_data = [
        ['Overall Total:', f'{overall_total:.1f} / {session.total_marks * len(results):.0f}',
         'Overall %:', f'{overall_pct:.1f}%' if overall_pct else '–',
         'Grade Band:', overall_band],
        ['Stream Position:', f'{stream_pos} / {stream_total}',
         'Grade Position:', f'{grade_pos} / {grade_total}', '', ''],
    ]
    sum_table = Table(summary_data, colWidths=[3.5*cm, 4*cm, 3*cm, 3*cm, 2.5*cm, 2.5*cm])
    sum_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTNAME', (4, 0), (4, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 0), (-1, -1), light_green),
        ('BOX', (0, 0), (-1, -1), 1, green),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.lightgrey),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    elements.append(sum_table)
    elements.append(Spacer(1, 16))

    # CBC grade key
    key_data = [['CBC Grade Key:  EE = Exceeding Expectations (75–100%)  |  ME = Meeting Expectations (50–74%)  |  AE = Approaching Expectations (25–49%)  |  BE = Below Expectations (0–24%)']]
    key_table = Table(key_data, colWidths=[18.5*cm])
    key_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Oblique'),
        ('FONTSIZE', (0, 0), (-1, -1), 7),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.grey),
    ]))
    elements.append(key_table)
    elements.append(Spacer(1, 20))

    # Signature line
    sig_data = [['Class Teacher: ___________________________', 'Principal: ___________________________', f'Date: {date.today().strftime("%d/%m/%Y")}']]
    sig_table = Table(sig_data, colWidths=[7*cm, 7*cm, 4.5*cm])
    sig_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (2, 0), (2, 0), 'RIGHT'),
    ]))
    elements.append(sig_table)

    doc.build(elements)
    buffer.seek(0)
    filename = f"report_slip_{student.student_id}_{session.year}.pdf"
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def school_subscriptions(request):
    return render(request, "school/subscription.html", {})


def school_notifications(request):
    return render(request, "school/notification.html", {})


def school_books(request):
    return render(request, "school/book.html", {})

def school_book_chapters(request):
    return render(request, "school/book_chapter.html", {})

def school_library_access(request):
    return render(request, "school/library.html", {})

def school_calendar(request):
    return render(request, "school/calendar.html", {})

def school_events(request):
    return render(request, "school/event.html", {})

def school_fees(request):
    return render(request, "school/fee.html", {})

def school_certificates(request):
    return render(request, "school/certificate.html", {})

def scholarships(request):
    return render(request, "school/scholarship.html", {})

@login_required
def school_teacher_subjects(request, staff_id):
    """
    View for school admin to manage subjects assigned to a specific teacher.
    """
    try:
        school = request.user.staffprofile.school
    except AttributeError:
        messages.error(request, "Access denied: Admin privileges required.")
        return redirect('userauths:teacher-dashboard')
    
    teacher = get_object_or_404(StaffProfile, pk=staff_id, school=school)
    
    query = request.GET.get('q', '')
    subjects = Subject.objects.filter(school=school, teacher=teacher).select_related('grade')
    if query:
        subjects = subjects.filter(
            Q(name__icontains=query) | Q(code__icontains=query)
        )
    form = SubjectForm(school=school)

    context = {
        'subjects': subjects,
        'teacher': teacher,
        'form': form,
        'school': school,
        'query': query,
    }
    return render(request, 'school/teacher/subjects.html', context)

@login_required
def subject_teacher_create(request, staff_id):
    """
    Assign a subject to a teacher.
    """
    try:
        school = request.user.staffprofile.school
    except AttributeError:
        messages.error(request, "Access denied: Admin privileges required.")
        return redirect('school:school-subjects', staff_id=staff_id)
    
    teacher = get_object_or_404(StaffProfile, pk=staff_id, school=school)
    
    if request.method == 'POST':
        subject_id = request.POST.get('subject_id')
        subject = get_object_or_404(Subject, pk=subject_id, school=school)
        
        if subject.teacher == teacher:
            messages.warning(request, f'Subject "{subject.name}" is already assigned to {teacher.user.get_full_name()}.')
        else:
            subject.teacher = teacher
            subject.save()
            messages.success(request, f'Subject "{subject.name}" assigned to {teacher.user.get_full_name()}.')
            logger.info(f"Assigned subject {subject.name} to teacher {teacher.user.get_full_name()} for school {school.name}.")
        
        return redirect('school:school-teacher-subjects', staff_id=staff_id)
    
    return redirect('school:school-teacher-subjects', staff_id=staff_id)

@login_required
def subject_teacher_edit(request, staff_id, pk):
    """
    Edit the subject assigned to a teacher.
    """
    try:
        school = request.user.staffprofile.school
    except AttributeError:
        messages.error(request, "Access denied: Admin privileges required.")
        return redirect('school:school-subjects')
    
    teacher = get_object_or_404(StaffProfile, pk=staff_id, school=school)
    subject = get_object_or_404(Subject, pk=pk, school=school)
    
    if request.method == 'POST':
        form = SubjectForm(request.POST, instance=subject, school=school)
        if form.is_valid():
            form.save()
            messages.success(request, f'Subject "{subject.name}" updated successfully.')
            return redirect('school:school-teacher-subjects', staff_id=staff_id)
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{form.fields[field].label}: {error}")
    
    return redirect('school:school-teacher-subjects', staff_id=staff_id)

@login_required
def subject_teacher_delete(request, staff_id, pk):
    """
    Remove a subject from a teacher.
    """
    try:
        school = request.user.staffprofile.school
    except AttributeError:
        messages.error(request, "Access denied: Admin privileges required.")
        return redirect('school:school-subjects')
    
    teacher = get_object_or_404(StaffProfile, pk=staff_id, school=school)
    subject = get_object_or_404(Subject, pk=pk, school=school)
    
    if request.method == 'POST':
        subject.teacher = None
        subject.save()
        messages.success(request, f'Subject "{subject.name}" removed from {teacher.user.get_full_name()}.')
        return redirect('school:school-teacher-subjects', staff_id=staff_id)
    
    return redirect('school:school-teacher-subjects', staff_id=staff_id)


@login_required
def school_subjects(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    query = request.GET.get('q', '')

    subject_qs = (
        Subject.objects
        .filter(school=school)
        .select_related('school', 'pathway')
        .prefetch_related('grade')
        .order_by('name')
    )

    if query:
        subject_qs = subject_qs.filter(
            Q(name__icontains=query) |
            Q(code__icontains=query) |
            Q(grade__name__icontains=query)
        ).distinct()  # 👈 important for M2M searches

    paginator = Paginator(subject_qs, 10)
    page_number = request.GET.get('page')
    subjects = paginator.get_page(page_number)

    form = SubjectForm(school=school)

    return render(
        request,
        'school/subject.html',
        {
            'subjects': subjects,
            'form': form,
            'school': school,
            'query': query,
        }
    )


@login_required
def subject_create(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    if request.method == 'POST':
        form = SubjectForm(request.POST, school=school)
        if form.is_valid():
            subject = form.save(commit=False)
            subject.school = school
            subject.save()
            messages.success(request, f'Subject "{subject.name}" created successfully.')
            logger.info(f"Created subject {subject.name} for school {school.name}.")
            return redirect('school:school-subjects')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{form.fields[field].label}: {error}")
    return redirect('school:school-subjects')

@login_required
def subject_edit(request, pk):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    subject = get_object_or_404(Subject, pk=pk, school=school)
    if request.method == 'POST':
        form = SubjectForm(request.POST, instance=subject, school=school)
        if form.is_valid():
            form.save()
            messages.success(request, f'Subject "{subject.name}" updated successfully.')
            return redirect('school:school-subjects')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{form.fields[field].label}: {error}")
    return redirect('school:school-subjects')

@login_required
def subject_delete(request, pk):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    subject = get_object_or_404(Subject, pk=pk, school=school)
    if request.method == 'POST':
        subject.is_active = False  # Soft delete
        subject.save()
        messages.success(request, f'Subject "{subject.name}" deactivated.')
        return redirect('school:school-subjects')
    return redirect('school:school-subjects')

@login_required
def school_enrollment(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    query = request.GET.get('q', '')
    enrollments_qs = (
        Enrollment.objects
        .filter(school=school)
        .select_related('student__user', 'student__grade_level', 'lesson__subject')
        .order_by('-enrolled_at')
    )
    if query:
        enrollments_qs = enrollments_qs.filter(
            Q(student__user__first_name__icontains=query)
            | Q(student__user__last_name__icontains=query)
            | Q(lesson__subject__name__icontains=query)
        )

    page = Paginator(enrollments_qs, 50).get_page(request.GET.get('page'))
    form = EnrollmentForm(school=school)
    context = {
        'enrollments': page,
        'form': form,
        'school': school,
        'query': query,
    }
    return render(request, 'school/enrollment.html', context)

@login_required
def enrollment_create(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    if request.method == 'POST':
        form = EnrollmentForm(request.POST, school=school)
        if form.is_valid():
            enrollment = form.save(commit=False)
            enrollment.school = school
            enrollment.save()
            messages.success(request, f'Enrollment for {enrollment.student} created.')
            return redirect('school:school-enrollment')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{form.fields[field].label}: {error}")
    return redirect('school:school-enrollment')


@login_required
def enrollment_edit(request, pk):
    """
    Edit existing enrollment (e.g., change subject).
    """
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    enrollment = get_object_or_404(Enrollment, pk=pk, school=school)
    if request.method == 'POST':
        form = EnrollmentForm(request.POST, instance=enrollment, school=school)
        if form.is_valid():
            form.save()
            messages.success(request, f'Enrollment for {enrollment.student} updated successfully.')
            return redirect('school:school-enrollment')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{form.fields[field].label}: {error}")
    return redirect('school:school-enrollment')

@login_required
def enrollment_delete(request, pk):
    """
    Soft-delete enrollment (set inactive or remove).
    """
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    enrollment = get_object_or_404(Enrollment, pk=pk, school=school)
    if request.method == 'POST':
        # Soft delete: Mark inactive or cascade to lessons if needed
        enrollment.is_active = False
        enrollment.save()
        messages.success(request, f'Enrollment for {enrollment.student} deactivated.')
        return redirect('school:school-enrollment')
    return redirect('school:school-enrollment')


@login_required
def school_timetable(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "No permission.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)
    if not school:
        return redirect('school:dashboard')

    # ---------------- WEEK NAV ----------------
    today = timezone.now().date()
    week_offset = int(request.GET.get("week", 0))

    start_of_week = today - timedelta(days=today.weekday()) + timedelta(weeks=week_offset)
    end_of_week = start_of_week + timedelta(days=4)

    # ---------------- LESSONS ----------------
    lessons = Lesson.objects.filter(
        timetable__school=school,
        lesson_date__range=[start_of_week, end_of_week]
    ).select_related(
        'subject', 'teacher__user', 'stream', 'time_slot', 'timetable__grade'
    )

    # ---------------- GROUP ----------------
    grid = defaultdict(lambda: defaultdict(list))
    time_slots = set()

    for lesson in lessons:
        day = lesson.lesson_date.strftime('%A')
        grid[lesson.time_slot][day].append(lesson)
        time_slots.add(lesson.time_slot)

    time_slots = sorted(time_slots, key=lambda x: x.start_time)
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    return render(request, "school/timetable.html", {
        "grid": grid,
        "time_slots": time_slots,
        "start_of_week": start_of_week,
        "end_of_week": end_of_week,
        "week_offset": week_offset,
        "days": days,
    })

# views.py
WEEKDAY_CHOICES = [
    ('monday', 'Monday'),
    ('tuesday', 'Tuesday'),
    ('wednesday', 'Wednesday'),
    ('thursday', 'Thursday'),
    ('friday', 'Friday'),
    ('saturday', 'Saturday'),
    ('sunday', 'Sunday'),
]

@login_required
@require_http_methods(["GET", "POST"])
def lesson_edit(request, lesson_id):
    lesson = get_object_or_404(Lesson, id=lesson_id)

    if request.method == "POST":
        # Store old position
        old_day = lesson.day_of_week
        old_slot = lesson.time_slot_id

        # Update lesson
        lesson.subject_id = request.POST.get("subject")
        lesson.stream_id = request.POST.get("stream")
        lesson.teacher_id = request.POST.get("teacher")
        lesson.day_of_week = request.POST.get("day_of_week")
        lesson.time_slot_id = request.POST.get("time_slot")
        lesson.room = request.POST.get("room")
        lesson.lesson_date = request.POST.get("lesson_date") or None
        lesson.notes = request.POST.get("notes")
        lesson.is_canceled = request.POST.get("is_canceled") == "on"
        lesson.save()

        # New position
        new_day = lesson.day_of_week
        new_slot = lesson.time_slot_id

        # Send SMS
        # try:
        #     send_sms_to_teacher(lesson)
        # except Exception as e:
        #     print("SMS error:", e)

        # Return updated HTML
        cell_html = render_to_string(
            "school/partials/lesson_cell_content.html",
            {"lesson": lesson},
            request=request
        )

        return JsonResponse({
            "success": True,
            "lesson_id": lesson.id,
            "cell_html": cell_html,
            "old_day": old_day,
            "old_slot": old_slot,
            "new_day": new_day,
            "new_slot": new_slot,
        })

    # GET → load form
    context = {
        "lesson": lesson,
        "subjects": Subject.objects.all(),
        "streams": Streams.objects.all(),
        "teachers": StaffProfile.objects.select_related("user"),
        "timeslots": TimeSlot.objects.all(),
        "weekday_choices": WEEKDAY_CHOICES,
    }

    return render(request, "school/partials/lesson_form.html", context)

@login_required
def lesson_attendance(request, lesson_id):
    lesson = get_object_or_404(Lesson, id=lesson_id)

    # Only active enrollments
    enrollments = Enrollment.objects.filter(
        lesson=lesson,
        status='active'
    ).select_related('student')

    # ---------------- POST REQUEST ----------------
    if request.method == "POST":
        with transaction.atomic():
            for e in enrollments:
                status = request.POST.get(f"status_{e.id}")   # <-- use enrollment ID
                remarks = request.POST.get(f"remarks_{e.id}")

                if not status:
                    continue  # skip empty

                # Only principals/deputies can suspend (18) or expel (20)
                if status in ('18', '20') and not (
                    request.user.is_principal or request.user.is_deputy_principal or request.user.is_admin
                ):
                    continue

                # Update or create attendance
                att, created = Attendance.objects.update_or_create(
                    enrollment=e,
                    date=lesson.lesson_date,
                    defaults={
                        "status": status,
                        "remarks": remarks,
                        "marked_by": getattr(request.user, "staffprofile", None)
                    }
                )

                # Update student suspension/expulsion flags
                student = e.student
                if status == '18':  # Suspension
                    student.suspended = True
                    student.expelled = False
                    student.is_active = True
                    student.save(update_fields=['suspended', 'expelled', 'is_active'])
                elif status == '20':  # Expulsion
                    student.expelled = True
                    student.suspended = False
                    student.is_active = False
                    student.save(update_fields=['expelled', 'suspended', 'is_active'])
                else:
                    if student.suspended or student.expelled:
                        student.suspended = False
                        student.expelled = False
                        student.is_active = True
                        student.save(update_fields=['suspended', 'expelled', 'is_active'])

        # Reload fresh attendance for modal display
        updated_attendance = Attendance.objects.filter(
            enrollment__lesson=lesson,
            date=lesson.lesson_date
        ).select_related("enrollment")

        attendance_map = {a.enrollment_id: a for a in updated_attendance}

        html = render_to_string(
            "school/partials/attendance_modal.html",
            {
                "lesson": lesson,
                "enrollments": enrollments,
                "attendance_map": attendance_map,
                "status_choices": Attendance.ATTENDANCE_STATUS_CHOICES,
            },
            request=request
        )

        return JsonResponse({"success": True, "html": html})

    # ---------------- GET REQUEST ----------------
    attendance_map = {}
    existing_attendance = Attendance.objects.filter(
        enrollment__lesson=lesson,
        date=lesson.lesson_date
    ).select_related("enrollment")

    # Populate existing attendance
    for a in existing_attendance:
        attendance_map[a.enrollment_id] = a

    # Pre-populate suspended/expelled students
    for e in enrollments:
        student = e.student
        if student.suspended:
            if e.id not in attendance_map:
                attendance_map[e.id] = Attendance(
                    enrollment=e,
                    date=lesson.lesson_date,
                    status='18',
                    remarks='Suspended'
                )
        elif student.expelled:
            if e.id not in attendance_map:
                attendance_map[e.id] = Attendance(
                    enrollment=e,
                    date=lesson.lesson_date,
                    status='20',
                    remarks='Expelled'
                )

    return render(request, "school/partials/attendance_modal.html", {
        "lesson": lesson,
        "enrollments": enrollments,
        "attendance_map": attendance_map,
        "status_choices": Attendance.ATTENDANCE_STATUS_CHOICES,
    })


# ──────────────────────────────────────────────────────────────────────────────
# STUDENT STATUS ACTIONS  (suspend / expel / reinstate)
# Only principals and deputy principals may perform these actions.
# ──────────────────────────────────────────────────────────────────────────────

@login_required
@require_POST
def student_suspend(request, student_id):
    user = request.user
    if not (user.is_principal or user.is_deputy_principal or user.is_admin):
        messages.error(request, "Only principals and deputy principals can suspend students.")
        return redirect('school:dashboard')
    school = get_user_school(user)
    student = get_object_or_404(Student, id=student_id, school=school)
    reason = request.POST.get('reason', '').strip()
    student.suspended = True
    student.expelled = False
    student.is_active = True
    student.save(update_fields=['suspended', 'expelled', 'is_active'])
    DisciplineRecord.objects.create(
        school=school,
        student=student,
        reported_by=getattr(user, 'staffprofile', None),
        incident_type='suspension',
        description=reason or 'Suspended by principal.',
        severity='critical',
        action_taken='Student suspended.',
    )
    messages.success(request, f"{student.user.get_full_name()} has been suspended.")
    return _safe_next_redirect(request, 'school:school-users')


@login_required
@require_POST
def student_expel(request, student_id):
    user = request.user
    if not (user.is_principal or user.is_deputy_principal or user.is_admin):
        messages.error(request, "Only principals and deputy principals can expel students.")
        return redirect('school:dashboard')
    school = get_user_school(user)
    student = get_object_or_404(Student, id=student_id, school=school)
    reason = request.POST.get('reason', '').strip()
    student.expelled = True
    student.suspended = False
    student.is_active = False
    student.save(update_fields=['expelled', 'suspended', 'is_active'])
    DisciplineRecord.objects.create(
        school=school,
        student=student,
        reported_by=getattr(user, 'staffprofile', None),
        incident_type='expulsion',
        description=reason or 'Expelled by principal.',
        severity='critical',
        action_taken='Student expelled.',
    )
    messages.success(request, f"{student.user.get_full_name()} has been expelled.")
    return _safe_next_redirect(request, 'school:school-users')


@login_required
@require_POST
def student_reinstate(request, student_id):
    user = request.user
    if not (user.is_principal or user.is_deputy_principal or user.is_admin):
        messages.error(request, "Only principals and deputy principals can reinstate students.")
        return redirect('school:dashboard')
    school = get_user_school(user)
    student = get_object_or_404(Student, id=student_id, school=school)
    student.suspended = False
    student.expelled = False
    student.is_active = True
    student.save(update_fields=['suspended', 'expelled', 'is_active'])
    messages.success(request, f"{student.user.get_full_name()} has been reinstated.")
    return _safe_next_redirect(request, 'school:school-users')


@login_required
def timetable_create(request):
    """
    Create a new timetable for a grade/term/year.
    Validates unique_together and date range.
    """
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    if request.method == 'POST':
        form = TimetableForm(request.POST, school=school)
        if form.is_valid():
            timetable = form.save(commit=False)
            timetable.school = school
            timetable.save()
            messages.success(request, f'Timetable for {timetable.grade.name} - Term {timetable.term} {timetable.year} created successfully.')
            logger.info(f"Created timetable {timetable.id} for {school.name}.")
            return redirect('school:school-timetable')
        else:
            # Render errors in messages
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{form.fields[field].label if field != '__all__' else 'Form'}: {error}")
    else:
        form = TimetableForm(school=school)
    
    # On GET error or initial, redirect to list
    return redirect('school:school-timetable')
# Similar edit/delete...


@login_required
def timetable_edit(request, pk):
    """
    Edit existing timetable (e.g., adjust dates).
    """
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    timetable = get_object_or_404(Timetable, pk=pk, school=school)
    if request.method == 'POST':
        form = TimetableForm(request.POST, instance=timetable, school=school)
        if form.is_valid():
            form.save()
            messages.success(request, f'Timetable for {timetable.grade.name} updated successfully.')
            return redirect('school:school-timetable')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{form.fields[field].label if field != '__all__' else 'Form'}: {error}")
    else:
        form = TimetableForm(instance=timetable, school=school)
    
    context = {'form': form, 'timetable': timetable}
    return render(request, 'school/timetable_form.html', context)  # Or modal-redirect


@login_required
def timetable_delete(request, pk):
    """
    Soft-delete timetable (set inactive or remove).
    """
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    timetable = get_object_or_404(Timetable, pk=pk, school=school)
    if request.method == 'POST':
        # Soft delete: Mark inactive or cascade to lessons if needed
        timetable.delete()  # Or timetable.is_active = False; timetable.save()
        messages.success(request, f'Timetable for {timetable.grade.name} deleted.')
        return redirect('school:school-timetable')
    messages.warning(request, "Confirm deletion.")
    return redirect('school:school-timetable')


# List lessons via @login_required
@login_required
def teacher_lessons(request, staff_id):
    """
    List all lessons for a specific teacher.
    """
    try:
        school = request.user.staffprofile.school
    except AttributeError:
        messages.error(request, "Access denied: Admin privileges required.")
        return redirect('userauths:teacher-dashboard')
    
    teacher = get_object_or_404(StaffProfile, pk=staff_id, school=school)
    lessons = Lesson.objects.filter(teacher=teacher).select_related(
        'subject', 'timetable__grade', 'time_slot', 'stream'
    ).order_by('day_of_week', 'lesson_date', 'time_slot__start_time')

    query = request.GET.get('q', '')
    if query:
        lessons = lessons.filter(
            Q(subject__name__icontains=query) |
            Q(timetable__grade__name__icontains=query) |
            Q(stream__name__icontains=query) |
            Q(day_of_week__icontains=query)
        )
    
    paginator = Paginator(lessons, 10)
    page_number = request.GET.get('page')
    lessons_page = paginator.get_page(page_number)
    form = LessonForm(school=school)
    from django.utils import timezone as tz
    context = {
        'teacher': teacher,
        'lessons': lessons_page,
        'school': school,
        'form': form,
        'query': query,
        'today': tz.localdate(),
    }
    return render(request, 'school/teacher/lessons.html', context)

@login_required
def teacher_lesson_create(request):
    """
    Create a new lesson for a specific teacher.
    """
    try:
        school = request.user.staffprofile.school
    except AttributeError:
        messages.error(request, "Access denied: Admin privileges required.")
        return redirect('userauths:teacher-dashboard')
    
    if request.method == 'POST':
        form = LessonForm(request.POST, school=school)
        if form.is_valid():
            lesson = form.save(commit=False)
            lesson.teacher = request.user.staffprofile
            lesson.save()
            messages.success(request, f'Lesson for {lesson.subject} on {lesson.date} created.')
            return redirect('school:teacher-lessons', staff_id=request.user.staffprofile.id)
    else:
        form = LessonForm(school=school)
    
    context = {
        'form': form,
        'school': school,
        'teacher': request.user.staffprofile,
    }
    return render(request, 'school/teacher/lesson_form.html', context)  # Or modal-redirect


@login_required
def teacher_lesson_edit(request, lesson_id):
    """
    Edit an existing lesson for a specific teacher.
    """
    try:
        school = request.user.staffprofile.school
    except AttributeError:
        messages.error(request, "Access denied: Admin privileges required.")
        return redirect('userauths:teacher-dashboard')
    
    lesson = get_object_or_404(Lesson, pk=lesson_id, teacher=request.user.staffprofile)
    if request.method == 'POST':
        form = LessonForm(request.POST, instance=lesson, school=school)
        if form.is_valid():
            form.save()
            messages.success(request, f'Lesson for {lesson.subject} on {lesson.date} updated.')
            return redirect('school:teacher-lessons', staff_id=request.user.staffprofile.id)
    else:
        form = LessonForm(instance=lesson, school=school)
    
    context = {
        'form': form,
        'lesson': lesson,
        'school': school,
        'teacher': request.user.staffprofile,
    }
    return render(request, 'school/teacher/lesson_form.html', context)  # Or modal-redirect


@login_required
def teacher_lesson_delete(request, lesson_id):
    """
    Delete a lesson for a specific teacher.
    """
    try:
        school = request.user.staffprofile.school
    except AttributeError:
        messages.error(request, "Access denied: Admin privileges required.")
        return redirect('userauths:teacher-dashboard')
    
    lesson = get_object_or_404(Lesson, pk=lesson_id, teacher=request.user.staffprofile)
    if request.method == 'POST':
        lesson.delete()
        messages.success(request, f'Lesson for {lesson.subject} on {lesson.date} deleted.')
        return redirect('school:teacher-lessons', staff_id=request.user.staffprofile.id)
    messages.warning(request, "Confirm deletion.")
    return redirect('school:teacher-lessons', staff_id=request.user.staffprofile.id)

def lesson_list(request, timetable_id):
    """
    List all lessons for a specific timetable
    """
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    timetable = get_object_or_404(Timetable, id=timetable_id, school=school)
    lessons = timetable.lessons.select_related(
        'subject', 'teacher__user', 'time_slot', 'stream'
    ).order_by('day_of_week', 'lesson_date', 'time_slot__start_time')

    query = request.GET.get('q', '')
    if query:
        lessons = lessons.filter(
            Q(subject__name__icontains=query) |
            Q(teacher__user__first_name__icontains=query) |
            Q(teacher__user__last_name__icontains=query) |
            Q(stream__name__icontains=query) |
            Q(room__icontains=query)
        )
    
    paginator = Paginator(lessons, 10)
    page_number = request.GET.get('page')
    lessons_page = paginator.get_page(page_number)
    
    form = LessonForm(school=school, initial={'timetable': timetable})
    context = {
        'timetable': timetable,
        'lessons': lessons_page,
        'form': form,
        'query': query,
        'school': school,
    }
    return render(request, 'school/lesson.html', context)


@login_required
def lesson_create(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    if request.method == 'POST':
        form = LessonForm(request.POST, school=school)
        if form.is_valid():
            lesson = form.save(commit=False)
            lesson.save()
            messages.success(request, f'Lesson for {lesson.subject} on {lesson.date} created.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{form.fields[field].label}: {error}")
    return redirect('school:lesson-list', timetable_id=request.POST.get('timetable'))


# @login_required
# def lesson_edit(request, lesson_id):
#     user = request.user

#     if not (user.is_admin or user.is_principal or user.is_deputy_principal):
#         messages.error(request, "You do not have permission to access this dashboard.")
#         return redirect("userauths:sign-in")

#     school = get_user_school(user)

#     if not school:
#         messages.error(request, "Access denied.")
#         return redirect('school:dashboard')
    
#     lesson = get_object_or_404(Lesson, id=lesson_id, timetable__school=school)
    
#     if request.method == 'POST':
#         form = LessonForm(request.POST, instance=lesson, school=school)
#         if form.is_valid():
#             form.save()
#             messages.success(request, f'Lesson for {lesson.subject} on {lesson.lesson_date} updated.')
#         else:
#             for field, errors in form.errors.items():
#                 for error in errors:
#                     if field == '__all__':
#                         messages.error(request, error)
#                     else:
#                         messages.error(request, f"{form.fields[field].label}: {error}")
#     return redirect('school:school-timetable')


@login_required
def lesson_delete(request, lesson_id):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    lesson = get_object_or_404(Lesson, id=lesson_id, timetable__school=school)
    timetable_id = lesson.timetable.id
    lesson.delete()
    messages.success(request, "Lesson deleted successfully.")
    return redirect('school:lesson-list', timetable_id=timetable_id)

@login_required
def school_virtual_classes(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    query = request.GET.get('q', '')
    sessions = Session.objects.filter(school=school).select_related('subject', 'teacher')
    if query:
        sessions = sessions.filter(Q(title__icontains=query) | Q(subject__name__icontains=query))
    
    form = SessionForm(school=school)
    context = {
        'sessions': sessions,
        'form': form,
        'school': school,
        'query': query,
    }
    return render(request, 'school/session.html', context)


@login_required
def session_create(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    if request.method == 'POST':
        form = SessionForm(request.POST, school=school)
        if form.is_valid():
            session = form.save(commit=False)
            session.school = school
            session.teacher = request.user.staffprofile
            session.save()
            messages.success(request, f"Session '{session.title}' created successfully.")
        else:
            messages.error(request, "Error creating session. Check the form.")
    return redirect('school:school-sessions')


@login_required
def session_edit(request, session_id):
    session = get_object_or_404(Session, id=session_id)
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    if request.method == 'POST':
        form = SessionForm(request.POST, instance=session, school=school)
        if form.is_valid():
            form.save()
            messages.success(request, f"Session '{session.title}' updated successfully.")
        else:
            messages.error(request, "Error updating session. Check the form.")
    return redirect('school:school-sessions')


@login_required
def session_delete(request, session_id):
    session = get_object_or_404(Session, id=session_id)
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    if request.method == 'POST':
        session.delete()
        messages.success(request, f"Session '{session.title}' deleted successfully.")
    return redirect('school:school-sessions')



def auto_mark_attendance_from_scan(student, stream, scan_log=None):
    """
    Auto-mark attendance using the given scan log.
    Only creates GradeAttendance if the scan came from a device ID starting with 'grade'.
    If scan_log is None, use the latest ScanLog for the student.
    """
    if not scan_log:
        scan_log = ScanLog.objects.filter(
            smart_id__profile=student.user,
            smart_id__school=stream.school
        ).order_by('-scanned_at').first()

    if not scan_log:
        return

    # Only mark if device_id starts with 'grade'
    if scan_log.device_id.lower().startswith('grade'):
        GradeAttendance.objects.create(
            student=student,
            stream=stream,
            status='P',
            scan_log=scan_log
        )


@login_required
def teacher_class_attendance_report(request, stream_id):
    school = request.user.staffprofile.school
    stream = get_object_or_404(Streams, id=stream_id, school=school)
    grade = stream.grade
    students = stream.students.all()
    attendance_summary = []

    # Handle POST save
    if request.method == 'POST':
        for student in students:
            status_key = f"status_{student.id}"
            status = request.POST.get(status_key)

            if status:
                # Update latest attendance OR create new if none exists
                latest = GradeAttendance.objects.filter(student=student, stream=stream).order_by('-recorded_at').first()
                if latest:
                    latest.status = status
                    latest.save()
                else:
                    GradeAttendance.objects.create(student=student, stream=stream, status=status)
        
        messages.success(request, "Attendance saved successfully!")
        return redirect(request.path)

    # GET → Show UI
    for student in students:

        # Auto-mark using ScanLog (creates new records for each scan)
        auto_mark_attendance_from_scan(student, stream)

        # Get latest attendance for display
        latest = GradeAttendance.objects.filter(student=student, stream=stream).order_by('-recorded_at').first()

        attendance_summary.append({
            'student': student,
            'existing': latest
        })

    paginator = Paginator(attendance_summary, 50)
    page = request.GET.get('page')
    attendance_summary = paginator.get_page(page)

    context = {
        'stream': stream,
        'grade': grade,
        'attendance_summary': attendance_summary,
        'school': school,
    }
    return render(request, 'school/teacher/class_attendance_report.html', context)

from django.db.models import Count, Q, F
from django.core.paginator import Paginator
@login_required
def attendance_dashboard(request):
    user = request.user

    # ───── PERMISSIONS ─────
    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)
    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    # ───── BASE QUERYSET ─────
    qs = Attendance.objects.select_related(
        'enrollment__student__user',
        'enrollment__student__grade_level',
        'enrollment__student__stream',
        'enrollment__lesson__subject',
        'enrollment__lesson__teacher__user',
        'term',
        'academic_year',
        'marked_by__user'
    ).filter(enrollment__school=school)

    # ───── FILTERS ─────
    filters = {
        'academic_year': request.GET.get('academic_year', ''),
        'term': request.GET.get('term', ''),
        'grade': request.GET.get('grade', ''),
        'stream': request.GET.get('stream', ''),
        'subject': request.GET.get('subject', ''),
        'student': request.GET.get('student', ''),
        'date_from': request.GET.get('date_from', ''),
        'date_to': request.GET.get('date_to', ''),
    }

    if filters['academic_year'].isdigit():
        qs = qs.filter(academic_year_id=filters['academic_year'])

    if filters['term'].isdigit():
        qs = qs.filter(term_id=filters['term'])

    if filters['grade'].isdigit():
        qs = qs.filter(enrollment__student__grade_level_id=filters['grade'])

    if filters['stream'].isdigit():
        qs = qs.filter(enrollment__student__stream_id=filters['stream'])

    if filters['subject'].isdigit():
        qs = qs.filter(enrollment__lesson__subject_id=filters['subject'])

    if filters['student'].strip():
        s = filters['student'].strip()
        qs = qs.filter(
            Q(enrollment__student__student_id__icontains=s) |
            Q(enrollment__student__user__first_name__icontains=s) |
            Q(enrollment__student__user__last_name__icontains=s)
        )

    # ───── DATE RANGE ─────
    date_from = parse_date(filters['date_from']) if filters['date_from'] else None
    date_to = parse_date(filters['date_to']) if filters['date_to'] else None

    if date_from and date_to:
        qs = qs.filter(date__range=(date_from, date_to))
    elif date_from:
        qs = qs.filter(date__gte=date_from)
    elif date_to:
        qs = qs.filter(date__lte=date_to)

    # ───── STATUS CARDS ─────
    stats = qs.aggregate(
        P=Count('id', filter=Q(status='P')),
        ET=Count('id', filter=Q(status='ET')),
        UT=Count('id', filter=Q(status='UT')),
        EA=Count('id', filter=Q(status='EA')),
        UA=Count('id', filter=Q(status='UA')),
    )

    stats = {k: v or 0 for k, v in stats.items()}

    status_cards = [
        {'label': 'Present', 'color': 'success', 'count': stats['P']},
        {'label': 'Tardy', 'color': 'warning', 'count': stats['ET'] + stats['UT']},
        {'label': 'Absent', 'color': 'danger', 'count': stats['EA'] + stats['UA']},
    ]

    # ───── FULL STATUS TREND ─────
    status_trend = qs.values('date').annotate(
        P=Count('id', filter=Q(status='P')),
        ET=Count('id', filter=Q(status='ET')),
        UT=Count('id', filter=Q(status='UT')),
        EA=Count('id', filter=Q(status='EA')),
        UA=Count('id', filter=Q(status='UA')),
        IB=Count('id', filter=Q(status='IB')),
        S18=Count('id', filter=Q(status='18')),
        S20=Count('id', filter=Q(status='20')),
    ).order_by('date')

    # ───── MISSING ATTENDANCE ─────
    lesson_qs = Lesson.objects.filter(
        timetable__school=school,
        lesson_date__lt=localdate()
    )

    attendance_exists = Attendance.objects.filter(
        enrollment__lesson=OuterRef('pk')
    )

    missing_lessons = lesson_qs.annotate(
        has_attendance=Exists(attendance_exists)
    ).filter(has_attendance=False)

    missing_lessons_count = missing_lessons.count()

    # ───── TEACHER MISSED LESSONS ─────
    teacher_missed_trends = missing_lessons.values(
        'teacher__user__first_name',
        'teacher__user__last_name',
        'teacher__pk',
    ).annotate(
        missed=Count('id')
    ).order_by('-missed')[:10]

    # Per-teacher missed lesson details (for hover tooltip)
    import json as _json
    teacher_missed_details = {}
    for t in teacher_missed_trends:
        tid = t.get('teacher__pk')
        if tid:
            details = list(
                missing_lessons.filter(teacher_id=tid).select_related(
                    'subject', 'stream', 'timetable__grade', 'time_slot'
                ).values(
                    'lesson_date',
                    'time_slot__start_time',
                    'time_slot__end_time',
                    'subject__name',
                    'stream__name',
                    'timetable__grade__name',
                )[:15]
            )
            teacher_name = f"{t['teacher__user__first_name']} {t['teacher__user__last_name']}"
            teacher_missed_details[teacher_name] = [
                {
                    'date': str(d['lesson_date']),
                    'time': f"{d['time_slot__start_time']} - {d['time_slot__end_time']}" if d['time_slot__start_time'] else '—',
                    'subject': d['subject__name'] or '—',
                    'class': f"{d['timetable__grade__name']} {d['stream__name']}",
                }
                for d in details
            ]
    teacher_missed_details_json = _json.dumps(teacher_missed_details)

    # ───── DISCIPLINE BASE ─────
    discipline_qs = DisciplineRecord.objects.filter(
        school=school,
        linked_attendance__in=qs
    )

    # ───── DISCIPLINE STATS ─────
    discipline_stats = discipline_qs.aggregate(
        total=Count('id'),
        unresolved=Count('id', filter=Q(resolved=False)),
        severe=Count('id', filter=Q(severity='severe')),
    )

    discipline_stats = {k: v or 0 for k, v in discipline_stats.items()}

    # ───── DISCIPLINE DAILY TREND (FIXED ✅ uses date) ─────
    discipline_daily = discipline_qs.values('date').annotate(
        total=Count('id'),
        severe=Count('id', filter=Q(severity='severe'))
    ).order_by('date')

    # ───── DISCIPLINE PER GRADE ─────
    discipline_by_grade = discipline_qs.values(
        'linked_attendance__enrollment__student__grade_level__name'
    ).annotate(
        total=Count('id'),
        severe=Count('id', filter=Q(severity='severe')),
        repeats=Sum('frequency_count')
    )

    # ───── INCIDENT TYPE TREND (NEW 🔥) ─────
    discipline_by_type = discipline_qs.values('incident_type').annotate(
        total=Count('id')
    ).order_by('-total')

    # ───── TOP REPEAT OFFENDERS (NEW 🔥) ─────
    repeat_offenders = discipline_qs.values(
        'student__user__first_name',
        'student__user__last_name'
    ).annotate(
        total_cases=Count('id'),
        repeat_score=Sum('frequency_count')
    ).order_by('-repeat_score')[:10]

    # ───── DEFAULT TREND ─────
    trend_qs = qs
    if not (date_from or date_to):
        today = localdate()
        trend_qs = qs.filter(date__gte=today - timedelta(days=30))

    daily_trend = trend_qs.values('date').annotate(
        present=Count('id', filter=Q(status='P')),
        absent=Count('id', filter=Q(status__in=['EA', 'UA'])),
    ).order_by('date')

    # ───── PAGINATION ─────
    paginator = Paginator(qs.order_by('-date', '-id'), 25)
    page_obj = paginator.get_page(request.GET.get('page'))

    # ───── FINAL CONTEXT ─────
    return render(request, 'school/attendance_dashboard.html', {
        'attendances': page_obj,
        'status_cards': status_cards,
        'daily_trend': list(daily_trend),

        'status_trend': list(status_trend),
        'teacher_missed_trends': list(teacher_missed_trends),
        'teacher_missed_details_json': teacher_missed_details_json,
        'discipline_daily': list(discipline_daily),
        'discipline_by_grade': list(discipline_by_grade),
        'discipline_by_type': list(discipline_by_type),
        'repeat_offenders': list(repeat_offenders),
        'missing_lessons_count': missing_lessons_count,

        'discipline_stats': discipline_stats,
        'filters': filters,
        'academic_years': AcademicYear.objects.filter(school=school),
        'terms': Term.objects.filter(school=school),
        'grades': Grade.objects.filter(school=school),
        'streams': Streams.objects.filter(school=school),
        'subjects': Subject.objects.filter(school=school, is_active=True),
        'school': school,
        'alerts': [],
    })
@login_required
def get_streams_by_grade(request):
    grade_id = request.GET.get('grade')
    streams = []
    if grade_id:
        streams_qs = Streams.objects.filter(grade_id=grade_id).order_by('name')
        streams = [{'id': s.id, 'name': s.name} for s in streams_qs]
    return JsonResponse({'streams': streams})

# === EXPORT CSV ===
def export_attendance_csv(request):
    qs = Attendance.objects.all()

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="attendance_report.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Student', 'Grade', 'Stream', 'Subject', 'Date', 'Status'
    ])

    for a in qs:
        writer.writerow([
            a.enrollment.student.user.get_full_name(),
            a.enrollment.student.grade_level.name,
            a.enrollment.student.stream.name,
            a.enrollment.subject.name,
            a.date,
            a.status
        ])

    return response


@login_required
def attendance_mark(request, lesson_id):
    # --- Fetch lesson safely ---
    lesson = get_object_or_404(
        Lesson.objects.select_related('subject', 'school', 'timetable__term__academic_year'),
        id=lesson_id,
        teacher=request.user.staffprofile
    )

    # --- Fetch students enrolled in lesson's subject or fallback to all active students in the school ---
    students = Enrollment.objects.filter(
        school=lesson.school,
        status='active'
    ).select_related('student__user')

    # Optional: filter by subject if you want strict matching
    subject_students = students.filter(subject=lesson.subject)
    if subject_students.exists():
        students = subject_students

    # --- Fetch existing attendance records for this lesson date ---
    attendance_dict = {
        att.enrollment_id: att
        for att in Attendance.objects.filter(
            enrollment__in=students,
            date=lesson.lesson_date
        )
    }

    term = getattr(lesson.timetable, 'term', None) if getattr(lesson, 'timetable', None) else None
    academic_year = getattr(term, 'academic_year', None) if term else None

    valid_statuses = {code for code, _ in Attendance.ATTENDANCE_STATUS_CHOICES}

    if request.method == 'POST':
        saved_count = 0
        for enrollment in students:
            field_name = f'status_{enrollment.id}'
            selected_status = request.POST.get(field_name)

            # Validate selected status
            if selected_status not in valid_statuses:
                selected_status = 'P'  # default to Present if invalid/missing

            Attendance.objects.update_or_create(
                enrollment=enrollment,
                date=lesson.lesson_date,
                defaults={
                    'status': selected_status,
                    'marked_by': request.user.staffprofile,
                    'term': term,
                    'academic_year': academic_year,
                }
            )
            saved_count += 1

        messages.success(request, f"Attendance for {saved_count} students has been recorded.")
        return redirect('school:teacher-attendance-mark', lesson_id=lesson.id)

    return render(request, 'school/teacher/attendance_mark.html', {
        'lesson': lesson,
        'students': students,
        'attendance_dict': attendance_dict,
        'school': lesson.school,
    })
@login_required
def attendance_edit(request, attendance_id):
    user = request.user
    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "Access denied.")
        return redirect('userauths:sign-in')

    school = get_user_school(user)
    if not school:
        messages.error(request, "No school profile found.")
        return redirect('school:dashboard')

    attendance = get_object_or_404(Attendance, id=attendance_id, enrollment__school=school)

    if request.method == 'POST':
        form = AttendanceForm(request.POST, instance=attendance)
        if form.is_valid():
            form.save()
            messages.success(request, f'Attendance for {attendance.enrollment.student} updated successfully.')
            return redirect('school:attendance-dashboard')
    else:
        form = AttendanceForm(instance=attendance)

    return render(request, 'school/attendance_edit.html', {
        'form': form,
        'attendance': attendance,
        'school': school,
    })

@login_required
def attendance_delete(request, attendance_id):
    user = request.user
    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "Access denied.")
        return redirect('userauths:sign-in')
    school = get_user_school(user)
    if not school:
        messages.error(request, "No school profile found.")
        return redirect('school:dashboard')
    attendance = get_object_or_404(Attendance, id=attendance_id, enrollment__school=school)
    if request.method == 'POST':
        attendance.delete()
        messages.success(request, f'Attendance for {attendance.enrollment.student} deleted successfully.')
    return redirect('school:attendance-dashboard')
# Summary view
@login_required
def attendance_summary(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    # Generate summary data
    period_start = request.GET.get('start')
    period_end = request.GET.get('end')
    if not period_start or not period_end:
        period_start = timezone.now().date().replace(day=1)
        period_end = timezone.now().date()
    
    grade_id = request.GET.get('grade')
    attendances = Attendance.objects.filter(
        enrollment__lesson__lesson_date__range=[period_start, period_end],
        enrollment__school=school,
    )
    if grade_id:
        attendances = attendances.filter(enrollment__student__grade_level_id=grade_id)

    summary = list(attendances.values('status').annotate(count=Count('status')).order_by('-count'))
    status_map = {row['status']: row['count'] for row in summary}
    total_records = sum(status_map.values())
    present_pct = round((status_map.get('P', 0) / total_records * 100), 1) if total_records else 0
    grades = Grade.objects.filter(school=school, is_active=True).order_by('name')

    context = {
        'summary': summary,
        'period_start': period_start,
        'period_end': period_end,
        'present_pct': present_pct,
        'school': school,
        'grades': grades,
    }
    return render(request, 'school/attendance_summary.html', context)


@login_required
def teacher_attendance(request):
    try:
        teacher = request.user.staffprofile
        school = teacher.school
    except AttributeError:
        messages.error(request, "Access denied.")
        return redirect('userauths:teacher-dashboard')

    # ───────────── DATE FILTER ─────────────
    start = request.GET.get('start')
    end = request.GET.get('end')

    if start and end:
        try:
            start = datetime.strptime(start, "%Y-%m-%d").date()
            end = datetime.strptime(end, "%Y-%m-%d").date()
        except ValueError:
            start = timezone.now().date().replace(day=1)
            end = timezone.now().date()
    else:
        start = timezone.now().date().replace(day=1)
        end = timezone.now().date()

    # ───────────── FILTERS ─────────────
    stream_id  = request.GET.get('stream', '')
    grade_id   = request.GET.get('grade', '')
    subject_id = request.GET.get('subject', '')
    student_search = request.GET.get('student_search', '').strip()

    # ───────────── LESSONS FOR TEACHER ─────────────
    lessons = Lesson.objects.filter(
        lesson_date__range=[start, end],
        teacher=teacher,
        is_canceled=False
    ).select_related('stream', 'subject')

    if stream_id:
        lessons = lessons.filter(stream_id=stream_id)
    if grade_id:
        lessons = lessons.filter(stream__grade_id=grade_id)
    if subject_id:
        lessons = lessons.filter(subject_id=subject_id)

    # ───────────── ATTENDANCE QUERY ─────────────
    attendances_qs = Attendance.objects.filter(
        enrollment__lesson__in=lessons,
        enrollment__status='active'
    ).select_related(
        'enrollment__student__user',
        'enrollment__lesson__stream',
        'enrollment__lesson__subject',
        'marked_by'
    ).order_by('-date')

    if student_search:
        attendances_qs = attendances_qs.filter(
            Q(enrollment__student__user__first_name__icontains=student_search) |
            Q(enrollment__student__user__last_name__icontains=student_search) |
            Q(enrollment__student__student_id__icontains=student_search)
        )

    # ───────────── SUMMARY ─────────────
    total_records = attendances_qs.count()
    summary_qs = attendances_qs.values('status').annotate(count=Count('status'))
    summary = {row['status']: row['count'] for row in summary_qs}

    present = summary.get('P', 0)
    absent   = summary.get('UA', 0) + summary.get('EA', 0)
    tardy    = summary.get('ET', 0) + summary.get('UT', 0)
    other    = total_records - present - absent - tardy

    present_pct = round((present / total_records * 100) if total_records else 0, 1)
    absent_pct  = round((absent  / total_records * 100) if total_records else 0, 1)
    tardy_pct   = round((tardy   / total_records * 100) if total_records else 0, 1)

    # Full per-status breakdown with labels + colours
    STATUS_META = {
        'P':  ('Present',             'success'),
        'ET': ('Early Tardy',         'warning'),
        'UT': ('Unexcused Tardy',     'warning'),
        'EA': ('Excused Absent',      'info'),
        'UA': ('Unexcused Absent',    'danger'),
        'IB': ('In Building',         'secondary'),
        '18': ('Suspension',          'dark'),
        '20': ('Expulsion',           'dark'),
    }
    status_breakdown = [
        {
            'code':  code,
            'label': meta[0],
            'color': meta[1],
            'count': summary.get(code, 0),
            'pct':   round(summary.get(code, 0) / total_records * 100, 1) if total_records else 0,
        }
        for code, meta in STATUS_META.items()
        if summary.get(code, 0) > 0
    ]

    # ───────────── PER-STUDENT SUMMARY ─────────────
    from django.db.models import Case, When, IntegerField
    student_summary = (
        attendances_qs
        .values(
            'enrollment__student_id',
            'enrollment__student__user__first_name',
            'enrollment__student__user__last_name',
            'enrollment__student__student_id',
        )
        .annotate(
            total=Count('id'),
            present_count=Count(Case(When(status='P',                      then=1), output_field=IntegerField())),
            tardy_count  =Count(Case(When(status__in=['ET','UT'],          then=1), output_field=IntegerField())),
            absent_count =Count(Case(When(status__in=['UA','EA'],          then=1), output_field=IntegerField())),
            other_count  =Count(Case(When(status__in=['IB','18','20'],     then=1), output_field=IntegerField())),
        )
        .order_by('enrollment__student__user__first_name')
    )
    for row in student_summary:
        row['present_pct'] = round(row['present_count'] / row['total'] * 100, 1) if row['total'] else 0

    # ───────────── PAGINATION ─────────────
    paginator = Paginator(attendances_qs, 50)
    attendances_page = paginator.get_page(request.GET.get('page'))

    # ───────────── FILTER DROPDOWNS ─────────────
    streams  = Streams.objects.filter(school=school).order_by('name')
    grades   = Grade.objects.filter(school=school).order_by('name')
    # Only subjects this teacher has taught in the period
    subjects = Subject.objects.filter(
        lessons__teacher=teacher,
        lessons__lesson_date__range=[start, end]
    ).distinct().order_by('name')

    context = {
        'attendances': attendances_page,
        'school': school,

        # Filters
        'period_start': start,
        'period_end': end,
        'streams': streams,
        'grades': grades,
        'subjects': subjects,
        'selected_stream':  stream_id,
        'selected_grade':   grade_id,
        'selected_subject': subject_id,
        'student_search':   student_search,

        # Summary
        'summary': summary,
        'status_breakdown': status_breakdown,
        'student_summary': student_summary,
        'total_records': total_records,
        'present':     present,
        'absent':      absent,
        'tardy':       tardy,
        'other':       other,
        'present_pct': present_pct,
        'absent_pct':  absent_pct,
        'tardy_pct':   tardy_pct,
    }

    return render(request, 'school/teacher/teacher_attendance_dashboard.html', context)


def send_parent_discipline_notification(parent, student, discipline, lesson):
    """
    Customize this based on your notification system (email, SMS, in-app, FCM, etc.)
    """
    subject = f"Discipline Alert: {student.user.get_full_name()}"
    message = (
        f"Dear Parent,\n\n"
        f"Your child {student.user.get_full_name()} was marked with "
        f"{discipline.get_incident_type_display()} on {discipline.date}.\n\n"
        f"Reason: {discipline.description}\n"
        f"Severity: {discipline.get_severity_display()}\n"
        f"Action Taken: {discipline.action_taken or 'None recorded'}\n\n"
        f"Teacher: {discipline.teacher.user.get_full_name()}\n"
        f"Subject: {lesson.subject.name}\n\n"
        f"Please contact the school if needed."
    )

    # Example using Django email
    try:
        send_mail(
            subject=subject,
            message=message,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[parent.email],  # assuming parent has email
            fail_silently=False,
        )
    except Exception as e:
        logger.error(f"Failed to send discipline email to {parent.email}: {e}")

# Constants
DISCIPLINE_CODES = ['IB', '18', '20']
ATTENDANCE_CODES = ['P', 'ET', 'UT', 'EA', 'UA']
@login_required
def teacher_attendance_mark(request, lesson_id):
    lesson = get_object_or_404(
        Lesson,
        id=lesson_id,
        teacher=request.user.staffprofile
    )

    teacher = request.user.staffprofile
    can_override = teacher.position in ['principal', 'deputy_principal']

    enrollments = Enrollment.objects.filter(
        lesson=lesson,
        status='active'
    ).select_related('student', 'student__user')

    # Prefetch today's attendance
    today_attendance_qs = Attendance.objects.filter(
        enrollment__in=enrollments,
        date=lesson.lesson_date
    ).select_related('enrollment')

    today_attendance_map = {a.enrollment_id: a for a in today_attendance_qs}

    # Current statuses
    current_statuses = {eid: att.status for eid, att in today_attendance_map.items()}

    # Status choices
    status_choices = [
        {'code': 'P',  'label': 'Present', 'color': 'success'},
        {'code': 'ET', 'label': 'Excused Tardy', 'color': 'warning'},
        {'code': 'UT', 'label': 'Unexcused Tardy', 'color': 'warning'},
        {'code': 'EA', 'label': 'Excused Absence', 'color': 'danger'},
        {'code': 'UA', 'label': 'Unexcused Absence', 'color': 'danger'},
        {'code': 'IB', 'label': 'Inappropriate Behavior', 'color': 'info'},
        {'code': '18', 'label': 'Suspension', 'color': 'dark'},
        {'code': '20', 'label': 'Expulsion', 'color': 'dark'},
    ]

    # Hide 18 & 20 from teachers without override
    if not can_override:
        status_choices = [s for s in status_choices if s['code'] not in ['18', '20']]

    # Determine if student is expelled/suspended
    disabled_attendance = {}
    forced_status_map = {}

    for e in enrollments:
        student = e.student
        if student.expelled:
            disabled_attendance[e.id] = True
            forced_status_map[e.id] = '20'
        elif student.suspended:
            disabled_attendance[e.id] = True
            forced_status_map[e.id] = '18'
        else:
            disabled_attendance[e.id] = False

    # POST - save attendance
    if request.method == 'POST' and request.headers.get('x-requested-with') == 'XMLHttpRequest':
        with transaction.atomic():
            updated_attendances = []

            for e in enrollments:
                # Force expelled/suspended status
                if e.id in forced_status_map:
                    status = forced_status_map[e.id]
                else:
                    status = request.POST.get(f'status_{e.id}')
                    # Prevent teacher from setting 18/20
                    if not can_override and status in ['18', '20']:
                        previous = today_attendance_map.get(e.id)
                        status = previous.status if previous else 'P'

                    # Validate
                    valid_codes = [s['code'] for s in status_choices]
                    if status not in valid_codes:
                        status = 'P'

                att, _ = Attendance.objects.update_or_create(
                    enrollment=e,
                    date=lesson.lesson_date,
                    defaults={
                        'status': status,
                        'term': lesson.timetable.term,
                        'academic_year': lesson.timetable.term.year,
                        'marked_by': teacher
                    }
                )
                updated_attendances.append(att)

        # Notifications for first/last slot
        is_first = is_first_slot(lesson)
        is_last  = is_last_slot(lesson)

        if is_first:
            for att in updated_attendances:
                if att.status in ['P', 'ET', 'UT', 'EA', 'UA']:
                    notify_first_lesson(att)

        if is_last:
            students_notified = set()
            for att in updated_attendances:
                student = att.enrollment.student
                if student.id not in students_notified:
                    notify_last_lesson(student, lesson.lesson_date)
                    students_notified.add(student.id)

        return JsonResponse({'success': True})

    # Term trends
    term_attendance = Attendance.objects.filter(
        enrollment__student__in=[e.student for e in enrollments],
        term=lesson.timetable.term
    ).values('enrollment__student_id', 'status')

    trend_map = defaultdict(lambda: {'P':0,'ET':0,'UT':0,'EA':0,'UA':0,'total':0})

    for row in term_attendance:
        sid = row['enrollment__student_id']
        status = row['status']
        if status in trend_map[sid]:
            trend_map[sid][status] += 1
        trend_map[sid]['total'] += 1

    # Convert to %
    for sid in trend_map:
        total = trend_map[sid]['total'] or 1
        for code in ['P','ET','UT','EA','UA']:
            trend_map[sid][code] = round((trend_map[sid][code] / total) * 100, 0)

    return render(request, 'school/teacher/attendance_mark.html', {
        'lesson': lesson,
        'enrollments': enrollments,
        'status_choices': status_choices,
        'current_statuses': current_statuses,
        'attendance_locked': False,  # always editable
        'can_override': can_override,
        'ATTENDANCE_CODES': ['P','ET','UT','EA','UA'],
        'DISCIPLINE_CODES': ['IB','18','20'],
        'disabled_attendance': disabled_attendance,
        'forced_status_map': forced_status_map,
        'INCIDENT_TYPE_CHOICES': INCIDENT_TYPE_CHOICES,
        'SEVERITY_CHOICES': DisciplineRecord._meta.get_field('severity').choices,
        'trend_map': dict(trend_map),
    })

@login_required
def teacher_attendance_smart(request, lesson_id):
    """
    Returns smart attendance statuses via AJAX.
    Only updates radio buttons; does NOT save anything.
    Scan logic:
        - Scan before 10 mins after lesson start: P
        - Scan within lesson time but after 10 mins: UT
        - No scan: UA
    Trends are ignored for now.
    """
    lesson = get_object_or_404(Lesson, id=lesson_id, teacher=request.user.staffprofile)

    enrollments = Enrollment.objects.filter(
        lesson=lesson,
        status='active'
    ).select_related('student', 'student__user')

    smart_statuses = {}

    # Combine lesson date with start/end times (timezone-aware)
    import pytz
    tz = pytz.timezone('Africa/Nairobi')
    lesson_start_dt = tz.localize(timezone.datetime.combine(lesson.lesson_date, lesson.time_slot.start_time))
    lesson_end_dt   = tz.localize(timezone.datetime.combine(lesson.lesson_date, lesson.time_slot.end_time))
    buffer_end = lesson_start_dt + timedelta(minutes=10)

    for e in enrollments:
        # Fetch scans for this student on lesson date
        student_scans = GradeAttendance.objects.filter(
            student=e.student,
            recorded_at__date=lesson.lesson_date
        ).order_by('recorded_at')

        if student_scans.exists():
            first_scan = student_scans.first().recorded_at

            if first_scan <= buffer_end:
                status = 'P'  # Present
            elif first_scan <= lesson_end_dt:
                status = 'UT'  # Unexcused Tardy
            else:
                status = 'UA'  # Scan after lesson end → unexcused absence
        else:
            status = 'UA'  # No scan → unexcused absence

        smart_statuses[e.id] = status

    return JsonResponse({
        'success': True,
        'statuses': smart_statuses
    })


@login_required
@require_POST
def teacher_discipline_create_ajax(request):
    teacher = request.user.staffprofile
    
    # ─── Extract & validate inputs ─────────────────────────────────────
    try:
        student_id     = request.POST['student_id']
        code           = request.POST.get('code')
        description    = request.POST['description'].strip()
        incident_type  = request.POST['incident_type']
        severity       = request.POST['severity']
        action         = request.POST.get('action', '').strip()

        if not description:
            return JsonResponse({'success': False, 'error': 'Description is required'}, status=400)

        # Discipline code authorization check
        if code in ['18', '20'] and teacher.position not in ['principal', 'deputy_principal']:
            return JsonResponse({'success': False, 'error': 'Only principal or deputy can issue suspension/expulsion'}, status=403)

    except KeyError as e:
        return JsonResponse({'success': False, 'error': f'Missing field: {str(e)}'}, status=400)

    # ─── Validate student exists and belongs to the school ─────────────
    try:
        student = Student.objects.get(id=student_id, school=teacher.school)
    except Student.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'Invalid or unauthorized student'}, status=403)

    # ─── Create record inside transaction ──────────────────────────────
    try:
        with transaction.atomic():
            record = DisciplineRecord.objects.create(
                student=student,               # better than student_id directly
                teacher=teacher,
                school=teacher.school,
                description=description,
                incident_type=incident_type,
                severity=severity,
                action_taken=action,
                reported_by=request.user,
                # Optional: link to the lesson if you want traceability
                # lesson_id=request.POST.get('lesson_id'),
            )
    except ValidationError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=400)
    except Exception as e:
        # In production: log this error
        return JsonResponse({'success': False, 'error': 'Failed to create record'}, status=500)

    return JsonResponse({
        'success': True,
        'id': record.id,
        'message': 'Discipline record created successfully'
    })

@login_required
def teacher_attendance_edit(request, attendance_id):
    attendance = get_object_or_404(Attendance, id=attendance_id, lesson__teacher=request.user.staffprofile)
    if request.method == 'POST':
        form = AttendanceForm(request.POST, instance=attendance)
        if form.is_valid():
            form.save()
            messages.success(request, f'Attendance for {attendance.enrollment.student} updated successfully.')
        else:
            messages.error(request, "Error updating attendance. Check the form.")
    return redirect('school:teacher-attendance')

@login_required
def teacher_attendance_delete(request, attendance_id):
    attendance = get_object_or_404(Attendance, id=attendance_id, lesson__teacher=request.user.staffprofile)
    if request.method == 'POST':
        attendance.delete()
        messages.success(request, f'Attendance for {attendance.enrollment.student} deleted successfully.')
    return redirect('school:teacher-attendance')

@login_required
def teacher_attendance_summary(request):
    try:
        teacher = request.user.staffprofile
        school = teacher.school
    except AttributeError:
        messages.error(request, "Access denied: Admin privileges required.")
        return redirect('userauths:teacher-dashboard')

    # ─── DATE FILTER ─────────────────────────
    period_start = request.GET.get('start')
    period_end = request.GET.get('end')

    if period_start and period_end:
        try:
            period_start = datetime.strptime(period_start, "%Y-%m-%d").date()
            period_end = datetime.strptime(period_end, "%Y-%m-%d").date()
        except ValueError:
            period_start = timezone.now().date().replace(day=1)
            period_end = timezone.now().date()
    else:
        period_start = timezone.now().date().replace(day=1)
        period_end = timezone.now().date()

    # ─── LESSONS ─────────────────────────
    lessons = Lesson.objects.filter(
        lesson_date__range=[period_start, period_end],
        teacher=teacher
    )

    # ─── ATTENDANCE ─────────────────────────
    attendances = Attendance.objects.filter(
        enrollment__lesson__in=lessons
    )

    total_records = attendances.count()

    # ─── SUMMARY ─────────────────────────
    summary_qs = attendances.values('status').annotate(count=Count('status'))
    summary = {row['status']: row['count'] for row in summary_qs}

    present_count = summary.get('P', 0)

    present_pct = (present_count / total_records * 100) if total_records else 0

    context = {
        'summary': summary,
        'period_start': period_start,
        'period_end': period_end,
        'present_pct': round(present_pct, 2),
        'school': school,
        'total_records': total_records,
    }

    return render(request, 'school/teacher/attendance_summary.html', context)

#teacher discipline

@login_required
def teacher_discipline(request):
    try:
        school = request.user.staffprofile.school
    except AttributeError:
        messages.error(request, "Access denied: Admin privileges required.")
        return redirect('userauths:teacher-dashboard')
    
    records = DisciplineRecord.objects.filter(school=school, teacher=request.user.staffprofile).select_related('student')
    
    query = request.GET.get('q', '')
    if query:
        records = records.filter(
            Q(student__user__first_name__icontains=query) | Q(description__icontains=query)
        )
    
    paginator = Paginator(records, 10)
    page_number = request.GET.get('page')
    records_page = paginator.get_page(page_number)
    form = DisciplineRecordForm(school=school)
    context = {
        'records': records_page,
        'school': school,
        'form': form,
        'query': query,
    }
    return render(request, 'school/teacher/discipline.html', context)

@login_required
def teacher_discipline_create(request):
    try:
        school = request.user.staffprofile.school
    except AttributeError:
        messages.error(request, "Access denied: Admin privileges required.")
        return redirect('userauths:teacher-dashboard')
    
    if request.method == 'POST':
        form = DisciplineRecordForm(request.POST, school=school)
        if form.is_valid():
            record = form.save(commit=False)
            record.school = school
            record.teacher = request.user.staffprofile
            record.reported_by = request.user
            record.save()

            # Notify all parents via SMS + email, track delivery status
            parents = record.student.parents.all()
            notif_title = f"Discipline Alert: {record.student.user.get_full_name()}"
            notif_msg = (
                f"Dear Parent, an incident of {record.get_incident_type_display()} "
                f"has been recorded for {record.student.user.get_full_name()} on {record.date}. "
                f"Severity: {record.get_severity_display()}. "
                f"Action: {record.action_taken or 'None recorded'}. "
                f"Please contact the school if needed."
            )
            for parent in parents:
                if not parent.user:
                    continue
                notif = Notification(
                    recipient=parent.user,
                    title=notif_title,
                    message=notif_msg,
                    related_discipline=record,
                    school=school,
                )
                phone = getattr(parent, 'phone', None) or getattr(parent.user, 'phone_number', None)
                if phone:
                    ok = _send_sms_via_eujim(str(phone), notif_msg)
                    notif.sms_sent = ok
                    if not ok:
                        notif.sms_error = "SMS delivery failed"
                if parent.user.email:
                    ok = send_email(parent.user.email, notif_title, notif_msg)
                    notif.email_sent = ok
                    if not ok:
                        notif.email_error = "Email delivery failed"
                notif.save()

            messages.success(request, f'Discipline record for {record.student} created and parents notified.')
            return redirect('school:teacher-discipline')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{form.fields[field].label}: {error}")

    return redirect('school:teacher-discipline')

@login_required
def teacher_discipline_delete(request, record_id):
    try:
        school = request.user.staffprofile.school
    except AttributeError:
        messages.error(request, "Access denied: Admin privileges required.")
        return redirect('userauths:teacher-dashboard')
    
    record = get_object_or_404(DisciplineRecord, id=record_id, school=school, teacher=request.user.staffprofile)
    if request.method == 'POST':
        record.delete()
        messages.success(request, f'Discipline record for {record.student} deleted.')
    return redirect('school:teacher-discipline')

@login_required
def teacher_discipline_edit(request, record_id):
    try:
        school = request.user.staffprofile.school
    except AttributeError:
        messages.error(request, "Access denied: Admin privileges required.")
        return redirect('userauths:teacher-dashboard')
    
    record = get_object_or_404(DisciplineRecord, id=record_id, school=school, teacher=request.user.staffprofile)
    if request.method == 'POST':
        form = DisciplineRecordForm(request.POST, instance=record, school=school)
        if form.is_valid():
            form.save()
            messages.success(request, f'Discipline record for {record.student} updated.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{form.fields[field].label}: {error}")
    return redirect('school:teacher-discipline')

@login_required
def school_discipline(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    query = request.GET.get('q', '')
    records_qs = (
        DisciplineRecord.objects
        .filter(school=school)
        .select_related('student__user', 'teacher__user')
        .prefetch_related('notification_set')
        .order_by('-date')
    )

    if query:
        records_qs = records_qs.filter(
            Q(student__user__first_name__icontains=query) |
            Q(student__user__last_name__icontains=query) |
            Q(description__icontains=query)
        )

    paginator = Paginator(records_qs, 10)  # 10 per page
    page_number = request.GET.get('page')
    records = paginator.get_page(page_number)

    form = DisciplineRecordForm(school=school)

    context = {
        'records': records,
        'form': form,
        'school': school,
        'query': query,
    }
    return render(request, 'school/discipline.html', context)


@login_required
def discipline_create(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    if request.method == 'POST':
        form = DisciplineRecordForm(request.POST, school=school)
        if form.is_valid():
            record = form.save(commit=False)
            record.school = school
            record.teacher = request.user.staffprofile
            record.reported_by = request.user
            record.save()
            
            # Notify all parents with SMS + email, track delivery status
            parents = record.student.parents.all()
            notif_title = f"Discipline Alert: {record.student.user.get_full_name()}"
            notif_msg = (
                f"Dear Parent, an incident of {record.get_incident_type_display()} "
                f"has been recorded for {record.student.user.get_full_name()} on {record.date}. "
                f"Severity: {record.get_severity_display()}. "
                f"Action: {record.action_taken or 'None recorded'}. "
                f"Please contact the school if needed."
            )
            for parent in parents:
                if not parent.user:
                    continue
                notif = Notification(
                    recipient=parent.user,
                    title=notif_title,
                    message=notif_msg,
                    related_discipline=record,
                    school=school,
                )
                # SMS
                phone = getattr(parent, 'phone', None) or getattr(parent.user, 'phone_number', None)
                if phone:
                    ok = _send_sms_via_eujim(str(phone), notif_msg)
                    notif.sms_sent = ok
                    if not ok:
                        notif.sms_error = "SMS delivery failed"
                # Email
                if parent.user.email:
                    ok = send_email(parent.user.email, notif_title, notif_msg)
                    notif.email_sent = ok
                    if not ok:
                        notif.email_error = "Email delivery failed"
                notif.save()

            messages.success(request, f'Discipline record for {record.student} created and parents notified.')
            return redirect('school:school-discipline')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{form.fields[field].label}: {error}")
    return redirect('school:school-discipline')

# Similar edit/delete, with validation for severity/frequency...

@login_required
def discipline_edit(request, pk):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    record = get_object_or_404(DisciplineRecord, id=pk, school=school)
    if request.method == 'POST':
        form = DisciplineRecordForm(request.POST, instance=record, school=school)
        if form.is_valid():
            form.save()
            messages.success(request, f'Discipline record for {record.student} updated.')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{form.fields[field].label}: {error}")
    return redirect('school:school-discipline')


@login_required
def discipline_delete(request, pk):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    record = get_object_or_404(DisciplineRecord, id=pk, school=school)
    if request.method == 'POST':
        record.delete()
        messages.success(request, f'Discipline record for {record.student} deleted.')
    return redirect('school:school-discipline')


@login_required
def school_notifications(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    notifications = Notification.objects.filter(school=school).order_by('-sent_at')
    # Mark as read for current user if applicable
    if request.user.is_authenticated:
        notifications.filter(recipient=request.user, is_read=False).update(is_read=True)
    
    form = NotificationForm()
    context = {
        'notifications': notifications,
        'form': form,
        'school': school,
    }
    return render(request, 'school/notification.html', context)

@login_required
def notification_send(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    if request.method == 'POST':
        form = NotificationForm(request.POST, school=school)
        if form.is_valid():
            notification = form.save(commit=False)
            notification.school = school
            notification.save()
            # Actual send via email/SMS (integrate with external service)
            if notification.recipient.email:
                send_mail(
                    notification.title,
                    notification.message,
                    settings.DEFAULT_FROM_EMAIL,
                    [notification.recipient.email],
                )
            messages.success(request, 'Notification sent successfully.')
            return redirect('school:school-notifications')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{form.fields[field].label}: {error}")
    return redirect('school:school-notifications')

@login_required
def school_reports(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    # Generate on-the-fly or from SummaryReport
    report_type = request.GET.get('type', 'school')
    period_start = request.GET.get('start', timezone.now().date().replace(day=1))
    period_end = request.GET.get('end', timezone.now().date())
    
    # Example: Attendance report
    if report_type == 'attendance':
        data = Attendance.objects.filter(
            enrollment__lesson__lesson_date__range=[period_start, period_end],
            enrollment__school=school,
        ).values('status').annotate(count=Count('status'))
        # Save as SummaryReport
        SummaryReport.objects.update_or_create(
            school=school, report_type='school', period_start=period_start, period_end=period_end,
            defaults={'data': dict(data), 'generated_by': request.user.staffprofile}
        )
    
    context = {
        'report_type': report_type,
        'period_start': period_start,
        'period_end': period_end,
        'school': school,
        # Pass data...
    }
    return render(request, 'school/report.html', context)

@login_required
def export_report(request):
    """Filtered attendance CSV — delegates to export_attendance_csv."""
    return export_attendance_csv(request)

@login_required
def export_pdf_report(request):
    """Filtered attendance PDF — delegates to export_attendance_pdf."""
    return export_attendance_pdf(request)

@login_required
def school_student_assignments(request):
    try:
        school = request.user.staffprofile.school
    except AttributeError:
        messages.error(request, "Access denied: Admin privileges required.")
        return redirect('school:dashboard')
    
    query = request.GET.get('q', '')
    assignments = Assignment.objects.filter(school=school).select_related('subject')
    if query:
        assignments = assignments.filter(Q(title__icontains=query) | Q(subject__name__icontains=query))
    
    form = AssignmentForm(school=school)
    context = {
        'assignments': assignments,
        'form': form,
        'school': school,
        'query': query,
    }
    return render(request, 'school/assignment.html', context)

# Create/edit assignment similar...
@login_required
def assignment_create(request):
    try:
        school = request.user.staffprofile.school
    except AttributeError:
        messages.error(request, "Access denied: Admin privileges required.")
        return redirect('school:school-assignments')
    
    if request.method == 'POST':
        form = AssignmentForm(request.POST, school=school)
        if form.is_valid():
            assignment = form.save(commit=False)
            assignment.school = school
            assignment.teacher = request.user.staffprofile
            assignment.save()
            messages.success(request, f'Assignment "{assignment.title}" created successfully.')
            return redirect('school:school-assignments')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{form.fields[field].label}: {error}")
    return redirect('school:school-assignments')
@login_required
def assignment_edit(request, pk):
    try:
        school = request.user.staffprofile.school
    except AttributeError:
        messages.error(request, "Access denied: Admin privileges required.")
        return redirect('school:school-assignments')
    
    assignment = get_object_or_404(Assignment, pk=pk, school=school)
    if request.method == 'POST':
        form = AssignmentForm(request.POST, instance=assignment, school=school)
        if form.is_valid():
            form.save()
            messages.success(request, f'Assignment "{assignment.title}" updated successfully.')
            return redirect('school:school-assignments')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{form.fields[field].label}: {error}")
    return redirect('school:school-assignments')

@login_required
def assignment_delete(request, pk):
    try:
        school = request.user.staffprofile.school
    except AttributeError:
        messages.error(request, "Access denied: Admin privileges required.")
        return redirect('school:school-assignments')
    
    assignment = get_object_or_404(Assignment, pk=pk, school=school)
    if assignment.submissions.exists():
        messages.error(request, "Cannot delete this assignment — students have already submitted work.")
        return redirect('school:school-assignments')
    assignment.delete()
    messages.success(request, f'Assignment "{assignment.title}" deleted successfully.')
    return redirect('school:school-assignments')
# Delete assignment


@login_required
def school_students_submissions(request, pk):
    try:
        school = request.user.staffprofile.school
    except AttributeError:
        messages.error(request, "Access denied: Admin privileges required.")
        return redirect('userauths:teacher-dashboard')

    assignment = get_object_or_404(Assignment, pk=pk, school=school)
    query = request.GET.get('q', '')
    submissions = Submission.objects.filter(
        school=school, assignment=assignment
    ).select_related('enrollment__student__user', 'assignment')
    if query:
        submissions = submissions.filter(
            Q(enrollment__student__user__first_name__icontains=query) |
            Q(enrollment__student__user__last_name__icontains=query)
        )

    context = {
        'submissions': submissions,
        'assignment': assignment,
        'school': school,
        'query': query,
    }
    return render(request, 'school/submission.html', context)

# Grade submission (update score/feedback)
@login_required
def submission_grade(request, pk):
    try:
        school = request.user.staffprofile.school
    except AttributeError:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    submission = get_object_or_404(Submission, pk=pk, school=school)
    if request.method == 'POST':
        try:
            raw = request.POST.get('score', '')
            submission.score = int(raw) if raw.strip() else None
        except (ValueError, TypeError):
            messages.error(request, 'Score must be a whole number.')
            return redirect('school:assignment-submissions', pk=submission.assignment_id)
        submission.feedback = request.POST.get('feedback', '')
        submission.save()
        messages.success(request, 'Submission graded.')
    return redirect('school:assignment-submissions', pk=submission.assignment_id)


#student submission

@login_required
def submission_create(request):
    try:
        student = request.user.student
    except AttributeError:
        messages.error(request, "Access denied: Student privileges required.")
        return redirect('school:dashboard')
    
    query = request.GET.get('q', '')
    submissions = Submission.objects.filter(enrollment__student=student).select_related('assignment')
    if query:
        submissions = submissions.filter(Q(assignment__title__icontains=query))
    
    context = {
        'submissions': submissions,
        'query': query,
    }
    return render(request, 'school/student_submissions.html', context)

@login_required
def submission_edit(request, pk):
    try:
        student = request.user.student
    except AttributeError:
        messages.error(request, "Access denied: Student privileges required.")
        return redirect('school:dashboard')
    
    submission = get_object_or_404(Submission, pk=pk, enrollment__student=student)
    if request.method == 'POST':
        submission.file = request.FILES.get('file', submission.file)
        submission.save()
        messages.success(request, 'Submission updated successfully.')
        return redirect('school:student-submissions')
    return redirect('school:student-submissions')


@login_required
def submission_delete(request, pk):
    try:
        student = request.user.student
    except AttributeError:
        messages.error(request, "Access denied: Student privileges required.")
        return redirect('school:dashboard')
    
    submission = get_object_or_404(Submission, pk=pk, enrollment__student=student)
    if request.method == 'POST':
        submission.delete()
        messages.success(request, 'Submission deleted successfully.')
    return redirect('school:student-submissions')


@login_required
def student_submissions(request):
    try:
        student = request.user.student
    except AttributeError:
        messages.error(request, "Access denied: Student privileges required.")
        return redirect('school:dashboard')

    query = request.GET.get('q', '')
    submissions = Submission.objects.filter(enrollment__student=student).select_related('assignment')
    if query:
        submissions = submissions.filter(Q(assignment__title__icontains=query))

    context = {
        'submissions': submissions,
        'query': query,
    }
    return render(request, 'school/student_submissions.html', context)


@login_required
def student_assignments_portal(request):
    """Student: view assignments, download files, upload submissions, see scores & remarks."""
    if not getattr(request.user, 'is_student', False):
        return redirect('userauths:sign-in')

    try:
        student = request.user.student
    except Exception:
        return redirect('userauths:sign-in')

    # Student's active subject enrollments
    subject_ids = SubjectEnrollment.objects.filter(
        student=student, is_active=True
    ).values_list('subject_id', flat=True)

    # All assignments for those subjects
    q = request.GET.get('q', '')
    status_filter = request.GET.get('status', '')
    subject_filter = request.GET.get('subject', '')

    assignments_qs = Assignment.objects.filter(
        school=student.school,
        subject_id__in=subject_ids,
    ).select_related('subject').prefetch_related('submissions').order_by('due_date')

    if q:
        assignments_qs = assignments_qs.filter(title__icontains=q)
    if subject_filter:
        assignments_qs = assignments_qs.filter(subject_id=subject_filter)

    # Annotate with student's submission
    from django.db.models import Prefetch
    student_submissions_prefetch = Prefetch(
        'submissions',
        queryset=Submission.objects.filter(enrollment__student=student).select_related('enrollment'),
        to_attr='my_submissions'
    )
    assignments_qs = assignments_qs.prefetch_related(student_submissions_prefetch)

    # Build enriched list
    today = timezone.now()
    assignment_data = []
    for asgn in assignments_qs:
        my_sub = asgn.my_submissions[0] if asgn.my_submissions else None
        if status_filter == 'submitted' and not my_sub:
            continue
        if status_filter == 'pending' and my_sub:
            continue
        if status_filter == 'overdue' and (my_sub or asgn.due_date >= today):
            continue
        assignment_data.append({
            'assignment': asgn,
            'my_submission': my_sub,
            'is_overdue': asgn.due_date < today and not my_sub,
        })

    # Handle submission upload
    if request.method == 'POST':
        asgn_id = request.POST.get('assignment_id')
        try:
            asgn = Assignment.objects.get(pk=asgn_id, school=student.school)
            # Get or create active enrollment for this subject
            enrollment = Enrollment.objects.filter(
                student=student, lesson__subject=asgn.subject, status='active'
            ).first()
            if not enrollment:
                messages.error(request, "You are not enrolled in this subject.")
            else:
                sub_file = request.FILES.get('file_submission')
                existing = Submission.objects.filter(enrollment=enrollment, assignment=asgn).first()
                if existing:
                    if sub_file:
                        existing.file_submission = sub_file
                        existing.save()
                    messages.success(request, "Submission updated.")
                else:
                    Submission.objects.create(
                        enrollment=enrollment,
                        assignment=asgn,
                        school=student.school,
                        file_submission=sub_file,
                    )
                    messages.success(request, "Assignment submitted successfully.")
        except Assignment.DoesNotExist:
            messages.error(request, "Assignment not found.")
        return redirect('school:student-assignments')

    subjects = Subject.objects.filter(id__in=subject_ids)
    paginator = Paginator(assignment_data, 15)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'school/student/assignments.html', {
        'assignment_data': page,
        'student': student,
        'subjects': subjects,
        'q': q,
        'status_filter': status_filter,
        'subject_filter': subject_filter,
    })
@login_required
def school_fees(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    query = request.GET.get('q', '')
    payments = Payment.objects.filter(school=school).select_related('student')
    if query:
        payments = payments.filter(Q(student__user__first_name__icontains=query) | Q(payment_type__icontains=query))
    
    form = PaymentForm(school=school)
    context = {
        'payments': payments,
        'form': form,
        'school': school,
        'query': query,
    }
    return render(request, 'school/fee.html', context)

# Integrate M-Pesa for payments...
@login_required
def process_mpesa_payment(request):
    # Example: Trigger STK push
    # Use Daraja API (assume configured)
    # For now, placeholder
    messages.success(request, 'Payment processed via M-Pesa.')
    return redirect('school:school-fees')

@login_required
def school_finance(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    subscriptions = SchoolSubscription.objects.filter(school=school)
    invoices = Invoice.objects.filter(school=school)
    
    context = {
        'subscriptions': subscriptions,
        'invoices': invoices,
        'school': school,
    }
    return render(request, 'school/finance.html', context)

@login_required
def school_subscriptions(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    subscription = school.platform_subscription
    form = SchoolSubscriptionForm(instance=subscription)
    
    if request.method == 'POST':
        form = SchoolSubscriptionForm(request.POST, instance=subscription)
        if form.is_valid():
            form.save()
            messages.success(request, 'Subscription updated.')
            return redirect('school:school-subscriptions')
    
    context = {
        'subscription': subscription,
        'form': form,
        'plans': SubscriptionPlan.objects.filter(is_active=True),
        'school': school,
    }
    return render(request, 'school/subscription.html', context)

# Generate invoice
@login_required
def generate_invoice(request):
    school = request.user.school_admin_profile
    subscription = school.platform_subscription
    if subscription:
        cost = subscription.calculate_current_cost()
        invoice = Invoice.objects.create(
            school=school,
            subscription=subscription,
            amount_due=cost,
            line_items=[{'description': 'Subscription', 'total': cost}]  # Simplify
        )
        messages.success(request, f'Invoice {invoice.invoice_id} generated.')
    return redirect('school:school-finance')

@login_required
def school_calendar(request):
    # Integrate with fullcalendar.js or similar in template
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    # Fetch events (use Session, Lesson, etc. as events)
    events = []  # JSON for JS
    import json
    context = {
        'events_json': json.dumps(events),  # rendered via |escapejs in template — no mark_safe needed
        'school': school,
    }
    return render(request, 'school/calendar.html', context)

@login_required
def school_events(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    # Events could be a new model or use existing (e.g., Sessions)
    # Placeholder: List sessions as events
    events = Session.objects.filter(school=school)
    context = {'events': events, 'school': school}
    return render(request, 'school/event.html', context)

# ------------------------------- SETTINGS & PERMISSIONS -------------------------------
@login_required
def school_settings(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)
    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    if request.method == 'POST':
        action = request.POST.get('action')

        if action == 'school_info':
            school.name = request.POST.get('name', school.name).strip()
            school.contact_email = request.POST.get('contact_email', school.contact_email).strip()
            school.contact_phone = request.POST.get('contact_phone', school.contact_phone).strip()
            if 'logo' in request.FILES:
                school.logo = request.FILES['logo']
            school.save()
            messages.success(request, 'School information updated.')

        elif action == 'admin_account':
            admin_user = school.school_admin
            new_email = request.POST.get('admin_email', '').strip()
            new_phone = request.POST.get('admin_phone', '').strip()
            if new_email and new_email != admin_user.email:
                from django.contrib.auth import get_user_model
                _User = get_user_model()
                if _User.objects.filter(email=new_email).exclude(pk=admin_user.pk).exists():
                    messages.error(request, 'That email is already in use.')
                else:
                    admin_user.email = new_email
                    admin_user.save(update_fields=['email'])
                    messages.success(request, 'Admin email updated.')
            if new_phone and new_phone != admin_user.phone_number:
                admin_user.phone_number = new_phone
                admin_user.save(update_fields=['phone_number'])
                messages.success(request, 'Admin phone updated.')

        elif action == 'change_password':
            from django.contrib.auth import update_session_auth_hash
            current  = request.POST.get('current_password', '')
            new_pw   = request.POST.get('new_password', '')
            confirm  = request.POST.get('confirm_password', '')
            if not user.check_password(current):
                messages.error(request, 'Current password is incorrect.')
            elif len(new_pw) < 8:
                messages.error(request, 'New password must be at least 8 characters.')
            elif new_pw != confirm:
                messages.error(request, 'New passwords do not match.')
            else:
                user.set_password(new_pw)
                user.save()
                update_session_auth_hash(request, user)
                messages.success(request, 'Password changed successfully.')

        return redirect('school:school-settings')

    context = {'school': school}
    return render(request, 'school/settings.html', context)

@login_required
def permissions_logs(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    # Log actions via middleware or signals; placeholder query
    logs = []  # From audit log model if added
    context = {'logs': logs, 'school': school}
    return render(request, 'school/permissions.html', context)  # Assume template

@login_required
def contact_messages(request):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    
    messages_list = ContactMessage.objects.all().order_by('-created_at')
    context = {'messages_list': messages_list, 'school': school}
    return render(request, 'school/contact.html', context)

# Callbacks for M-Pesa (STK push response, payment confirmation)
def mpesa_stk_callback(request):
    # Parse XML/JSON from M-Pesa
    # Update MpesaStkPushRequestResponse
    return JsonResponse({'ResultCode': 0, 'ResultDesc': 'Success'})

def mpesa_payment_callback(request):
    """Safaricom Daraja STK push callback for school fee payments."""
    import json as _json

    SAFARICOM_IPS = {
        '196.201.214.200', '196.201.214.206', '196.201.213.114',
        '196.201.214.207', '196.201.214.208', '196.201.213.44',
        '196.201.212.127', '196.201.212.138', '196.201.212.129',
        '196.201.212.136', '196.201.212.74',  '196.201.212.69',
    }
    if not settings.DEBUG:
        forwarded = request.META.get('HTTP_X_FORWARDED_FOR', '')
        caller_ip = forwarded.split(',')[0].strip() if forwarded else request.META.get('REMOTE_ADDR', '')
        if caller_ip not in SAFARICOM_IPS:
            logger.warning("M-Pesa school callback rejected from IP: %s", caller_ip)
            return JsonResponse({'ResultCode': 1, 'ResultDesc': 'Rejected'}, status=403)

    try:
        data = _json.loads(request.body)
        stk = data.get('Body', {}).get('stkCallback', {})
        result_code  = stk.get('ResultCode')
        checkout_id  = stk.get('CheckoutRequestID', '').strip()

        if not checkout_id:
            return JsonResponse({'ResultCode': 0, 'ResultDesc': 'Accepted'})

        mpesa_pay = MpesaPayment.objects.filter(
            checkout_request_id=checkout_id, status='pending'
        ).select_for_update().first()

        if not mpesa_pay:
            logger.info("M-Pesa school callback: no pending record for %s", checkout_id)
            return JsonResponse({'ResultCode': 0, 'ResultDesc': 'Accepted'})

        if result_code == 0:
            items = stk.get('CallbackMetadata', {}).get('Item', [])
            receipt = next((i['Value'] for i in items if i.get('Name') == 'MpesaReceiptNumber'), '')
            mpesa_pay.status = 'paid'
            mpesa_pay.mpesa_receipt_number = receipt
            mpesa_pay.save(update_fields=['status', 'mpesa_receipt_number'])
            # Update the linked FeeInvoice
            if hasattr(mpesa_pay, 'fee_invoice') and mpesa_pay.fee_invoice:
                inv = mpesa_pay.fee_invoice
                inv.amount_paid = (inv.amount_paid or 0) + (mpesa_pay.amount or 0)
                inv.status = 'paid' if inv.amount_paid >= inv.amount_required else 'partial'
                inv.save(update_fields=['amount_paid', 'status'])
        else:
            mpesa_pay.status = 'failed'
            mpesa_pay.save(update_fields=['status'])

    except Exception as exc:
        logger.error("M-Pesa school callback error: %s", exc)

    return JsonResponse({'ResultCode': 0, 'ResultDesc': 'Accepted'})

@login_required
def grade_streams_view(request, grade_id):

    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    grade = get_object_or_404(Grade, id=grade_id, school=school)

    streams = grade.streams.filter(is_active=True).order_by("name")

    data = [
        {
            "id": s.id,
            "name": s.name
        }
        for s in streams
    ]

    return JsonResponse(data, safe=False)

@login_required
def create_stream(request, grade_id):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    grade = get_object_or_404(Grade, id=grade_id, school=school)

    if request.method == "POST":

        form = StreamForm(request.POST)

        if form.is_valid():
            stream = form.save(commit=False)
            stream.grade = grade
            stream.school = school
            stream.save()

            messages.success(request, "Stream added successfully.")

        else:
            messages.error(request, "Failed to create stream.")

    return redirect("school:school-grades")

@login_required
def edit_stream(request, stream_id):

    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    stream = get_object_or_404(Streams, id=stream_id, school=school)

    if request.method == "POST":

        form = StreamForm(request.POST, instance=stream)

        if form.is_valid():
            form.save()
            messages.success(request, "Stream updated successfully.")

        else:
            messages.error(request, "Failed to update stream.")

    return redirect("school:school-grades")
@login_required
def delete_stream(request, stream_id):

    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    stream = get_object_or_404(Streams, id=stream_id, school=school)

    if request.method == "POST":

        stream.is_active = False
        stream.save()

        messages.success(request, "Stream deleted successfully.")

    return redirect("school:school-grades")


@login_required
def create_parent_student(request):
    # Ensure user is a school admin
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    if request.method == "POST":
        form = ParentStudentCreationForm(request.POST, request.FILES, school=school)
        if form.is_valid():
            try:
                # Atomic creation to ensure both User and Parent/Student are saved
                with transaction.atomic():
                    obj, temp_password = form.save(commit=True)

                member_type = 'Parent' if form.cleaned_data.get('member_type') == 'parent' else 'Student'
                messages.success(request, f"{member_type} created successfully.")

                # Optional: show temp password if email sending is disabled
                if form.cleaned_data.get('send_email'):
                    messages.info(request, f"A welcome email has been sent. Temporary password: {temp_password}")

                return redirect('school:school-users')

            except Exception as e:
                messages.error(request, f"Failed to create member: {str(e)}")
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = ParentStudentCreationForm(school=school)

    context = {
        'parent_student_form': form,
        'school': school,
    }
    return render(request, 'school/staff.html', context)

@login_required
def assign_class_teacher_for_teacher(request, teacher_id):
    user = request.user

    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "You do not have permission to access this dashboard.")
        return redirect("userauths:sign-in")

    school = get_user_school(user)

    if not school:
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    # Get teacher to assign
    teacher = get_object_or_404(StaffProfile, id=teacher_id, school=school)

    if request.method == 'POST':
        stream_id = request.POST.get('stream')
        stream = get_object_or_404(Streams, id=stream_id, school=school)

        assignment, created = TeacherStreamAssignment.objects.get_or_create(
            teacher=teacher,
            stream=stream,
            school=school
        )

        if created:
            messages.success(request, f"{teacher.user.get_full_name()} assigned to {stream.name} successfully.")
        else:
            messages.info(request, f"{teacher.user.get_full_name()} is already assigned to {stream.name}.")

        return redirect('school:assign-class-teacher-for-teacher', teacher_id=teacher.id)

    streams = Streams.objects.filter(school=school)
    assignments = TeacherStreamAssignment.objects.filter(teacher=teacher, school=school)

    return render(request, 'school/staff.html', {
        'teacher': teacher,
        'streams': streams,
        'assignments': assignments
    })



@login_required
def ajax_subjects(request):
    school_id = request.GET.get("school_id")
    grade_id = request.GET.get("grade_id")
    qs = Subject.objects.all()
    if school_id:
        qs = qs.filter(school_id=school_id)
    if grade_id:
        qs = qs.filter(grade__id=grade_id)
    subjects = qs.values("id","name").order_by("name")
    return JsonResponse(list(subjects), safe=False)

def is_policymaker(user):
    return getattr(user, "is_policy_maker", False)

# ───────────── AJAX FILTERS ─────────────
@login_required
def ajax_subcounties(request):
    county_id = request.GET.get("county_id")
    if not county_id:
        subcounties = SubCounty.objects.all().values("id","name").order_by("name")
    else:
        subcounties = SubCounty.objects.filter(county_id=county_id).values("id","name").order_by("name")
    return JsonResponse(list(subcounties), safe=False)

@login_required
def ajax_schools(request):
    subcounty_id = request.GET.get("subcounty_id")
    if not subcounty_id:
        schools = School.objects.filter(is_active=True).values("id","name").order_by("name")
    else:
        schools = School.objects.filter(sub_county_id=subcounty_id, is_active=True).values("id","name").order_by("name")
    return JsonResponse(list(schools), safe=False)

@login_required
def ajax_grades(request):
    school_id = request.GET.get("school_id")
    if not school_id:
        grades = Grade.objects.all().values("id","name").order_by("name")
    else:
        grades = Grade.objects.filter(school_id=school_id).values("id","name").order_by("name")
    return JsonResponse(list(grades), safe=False)


@login_required
def ajax_students_search(request):
    """AJAX: search students by name or ID for finance modal."""
    q = request.GET.get('q', '').strip()
    school_id = request.GET.get('school', '')
    school = get_user_school(request.user)
    if not school:
        return JsonResponse([], safe=False)

    qs = Student.objects.filter(school=school).select_related('user')
    if q:
        qs = qs.filter(
            Q(user__first_name__icontains=q) |
            Q(user__last_name__icontains=q) |
            Q(student_id__icontains=q)
        )
    data = [{'id': s.pk, 'name': s.user.get_full_name(), 'student_id': s.student_id} for s in qs[:20]]
    return JsonResponse(data, safe=False)


@login_required
@user_passes_test(is_policymaker)
def policymaker_dashboard(request):
    import json as _json
    from django.db.models.functions import TruncMonth

    active_view  = request.GET.get("view", "")
    county_id    = request.GET.get("county")
    subcounty_id = request.GET.get("subcounty")
    school_id    = request.GET.get("school")
    grade_id     = request.GET.get("grade")
    subject_id   = request.GET.get("subject")
    term_id      = request.GET.get("term")
    classification = request.GET.get("classification")

    # ───────── SCHOOL FILTER ─────────
    schools = School.objects.filter(is_active=True)
    if county_id:
        schools = schools.filter(county_id=county_id)
    if subcounty_id:
        schools = schools.filter(sub_county_id=subcounty_id)
    if school_id:
        schools = schools.filter(id=school_id)
    if classification:
        schools = schools.filter(school_classification=classification)

    # ───────── ATTENDANCE BASE QUERY ─────────
    attendance = (
        Attendance.objects
        .filter(enrollment__school__in=schools)
        .select_related(
            "enrollment__lesson__teacher__user",
            "enrollment__lesson__subject",
            "enrollment__lesson__stream__grade",
            "enrollment__school__county",
        )
    )
    if grade_id:
        attendance = attendance.filter(enrollment__lesson__stream__grade_id=grade_id)
    if subject_id:
        attendance = attendance.filter(enrollment__lesson__subject_id=subject_id)
    if term_id:
        attendance = attendance.filter(term_id=term_id)

    _att_expr = ExpressionWrapper(F("present") * 100.0 / F("total"), output_field=FloatField())

    # ───────── SCHOOL RANKING ─────────
    _school_rank_qs = (
        attendance.values("enrollment__school__id", "enrollment__school__name",
                          "enrollment__school__school_classification",
                          "enrollment__school__county__name")
        .annotate(total=Count("id"), present=Count("id", filter=Q(status="P")))
        .annotate(present_rate=ExpressionWrapper(F("present") * 100.0 / F("total"), output_field=FloatField()))
        .order_by("-present_rate")
    )
    school_rankings = _school_rank_qs if active_view == "schools" else _school_rank_qs[:20]

    # ───────── COUNTY RANKING ─────────
    county_rankings = (
        attendance.values("enrollment__school__county__name")
        .annotate(total=Count("id"), present=Count("id", filter=Q(status="P")))
        .annotate(performance=ExpressionWrapper(F("present") * 100.0 / F("total"), output_field=FloatField()))
        .order_by("-performance")
    )

    # ───────── TEACHER PERFORMANCE ─────────
    _teacher_rank_qs = (
        attendance.values(
            "enrollment__lesson__teacher__user__first_name",
            "enrollment__lesson__teacher__user__last_name",
            "enrollment__school__name",
            "enrollment__school__county__name",
        )
        .annotate(total=Count("id"), present=Count("id", filter=Q(status="P")))
        .annotate(performance=ExpressionWrapper(F("present") * 100.0 / F("total"), output_field=FloatField()))
        .order_by("-performance")
    )
    teacher_rankings = _teacher_rank_qs if active_view == "teachers" else _teacher_rank_qs[:15]

    # ───────── GRADE PERFORMANCE ─────────
    grade_rankings = (
        attendance.values("enrollment__lesson__stream__grade__name")
        .annotate(total=Count("id"), present=Count("id", filter=Q(status="P")))
        .annotate(performance=ExpressionWrapper(F("present") * 100.0 / F("total"), output_field=FloatField()))
        .order_by("-performance")
    )

    # ───────── SUBJECT PERFORMANCE ─────────
    subject_rankings = (
        attendance.values("enrollment__lesson__subject__name")
        .annotate(total=Count("id"), present=Count("id", filter=Q(status="P")))
        .annotate(performance=ExpressionWrapper(F("present") * 100.0 / F("total"), output_field=FloatField()))
        .order_by("-performance")
    )

    # ───────── NATIONAL / FILTERED RATE ─────────
    totals = attendance.aggregate(
        total=Count("id"),
        present=Count("id", filter=Q(status="P")),
        absent_unexcused=Count("id", filter=Q(status="UA")),
        tardy=Count("id", filter=Q(status__in=["ET","UT"])),
    )
    total_records = totals["total"] or 1
    national_rate = round((totals["present"] or 0) * 100.0 / total_records, 1)

    # ───────── TREND (monthly last 6 months) ─────────
    six_months_ago = timezone.localdate() - timedelta(days=182)
    trend_qs = (
        attendance.filter(date__gte=six_months_ago)
        .annotate(month=TruncMonth("date"))
        .values("month")
        .annotate(total=Count("id"), present=Count("id", filter=Q(status="P")))
        .order_by("month")
    )
    trend_labels = [r["month"].strftime("%b %Y") for r in trend_qs if r["month"]]
    trend_rates  = [
        round(r["present"] * 100.0 / r["total"], 1) if r["total"] else 0
        for r in trend_qs if r["month"]
    ]

    # ───────── DISCIPLINE BREAKDOWN ─────────
    _disc_base = DisciplineRecord.objects.filter(school__in=schools)
    discipline_breakdown = (
        _disc_base.values("incident_type", "severity")
        .annotate(count=Count("id"))
        .order_by("-count")
    )
    discipline_by_type = (
        _disc_base.values("incident_type")
        .annotate(count=Count("id"))
        .order_by("-count")
    )
    discipline_by_severity = (
        _disc_base.values("severity")
        .annotate(count=Count("id"))
        .order_by("-count")
    )
    discipline_per_school = (
        _disc_base.values("school__name", "school__county__name")
        .annotate(total=Count("id"), open=Count("id", filter=Q(resolved=False)))
        .order_by("-total")[:30]
    )

    # ───────── GENDER BREAKDOWN ─────────
    gender_qs = (
        Student.objects.filter(school__in=schools)
        .values("gender")
        .annotate(count=Count("id"))
    )
    gender_labels = [g["gender"].upper() or "Unknown" for g in gender_qs]
    gender_counts = [g["count"] for g in gender_qs]

    # ───────── SCHOOL TYPE BREAKDOWN ─────────
    school_type_qs = (
        schools.values("school_classification")
        .annotate(count=Count("id"))
        .order_by("-count")
    )

    # ───────── KPI SUMMARY ─────────
    total_schools   = schools.count()
    total_students  = Student.objects.filter(school__in=schools).count()
    total_teachers  = StaffProfile.objects.filter(school__in=schools, position="teacher").count()
    total_lessons   = Lesson.objects.filter(timetable__school__in=schools).count()
    total_discipline = DisciplineRecord.objects.filter(school__in=schools).count()

    # sub-counties for filter (scoped to county if selected)
    sub_counties = SubCounty.objects.filter(county_id=county_id) if county_id else SubCounty.objects.all()
    all_terms = Term.objects.filter(school__in=schools).distinct().order_by("-start_date")[:10]

    context = {
        # KPIs
        "total_schools":    total_schools,
        "total_students":   total_students,
        "total_teachers":   total_teachers,
        "total_lessons":    total_lessons,
        "total_discipline": total_discipline,
        "national_rate":    national_rate,
        "totals":           totals,
        # Rankings
        "school_rankings":  school_rankings,
        "county_rankings":  county_rankings,
        "teacher_rankings": teacher_rankings,
        "grade_rankings":   grade_rankings,
        "subject_rankings": subject_rankings,
        # Breakdowns
        "discipline_by_type":  discipline_by_type,
        "school_type_qs":      school_type_qs,
        # Charts (JSON)
        "trend_labels_json": _json.dumps(trend_labels),
        "trend_rates_json":  _json.dumps(trend_rates),
        "gender_labels_json": _json.dumps(gender_labels),
        "gender_counts_json": _json.dumps(gender_counts),
        # Filters
        "counties":        County.objects.all().order_by("name"),
        "sub_counties":    sub_counties.order_by("name"),
        "all_terms":       all_terms,
        # Selected filter values (for persistence)
        "sel_county":        county_id,
        "sel_subcounty":     subcounty_id,
        "sel_school":        school_id,
        "sel_grade":         grade_id,
        "sel_term":          term_id,
        "sel_classification": classification,
        "classification_choices": School._meta.get_field("school_classification").choices or [],
        # View switching
        "active_view": active_view,
        # Discipline focused data
        "discipline_by_severity": discipline_by_severity,
        "discipline_per_school":  discipline_per_school,
    }

    return render(request, "school/policy/dashboard.html", context)


#FIle Upload


import random
import string

def generate_password(length=8):
    return ''.join(random.choices(string.ascii_letters + string.digits, k=length))


def safe_email(value, fallback):
    return value.strip().lower() if value else fallback

def generate_time_slots(school, staff_user):
    start = datetime.datetime.strptime("08:00", "%H:%M")
    end = datetime.datetime.strptime("16:00", "%H:%M")

    slots = []
    current = start
    count = 0

    while current < end:
        duration = 45
        description = "Lesson"

        if count == 2 and count == 5:
            duration = 20
            description = "Break"
        elif count == 8:
            duration = 45
            description = "Lunch"

        next_time = current + timedelta(minutes=duration)

        slot, _ = TimeSlot.objects.get_or_create(
            school=school,
            start_time=current.time(),
            end_time=next_time.time(),
            defaults={
                "description": description,
                "created_by": staff_user
            }
        )

        if description == "Lesson":
            slots.append(slot)

        current = next_time
        count += 1

    return slots

def populate_student_lesson_enrollments(school, grade=None, stream=None, term=None):
    """
    Create Enrollment records for all lessons in a term based on student's subject enrollments.
    """
    import logging
    logger = logging.getLogger(__name__)

    # ── Students
    students = Student.objects.filter(school=school, is_active=True)
    if grade:
        students = students.filter(grade_level=grade)
    if stream:
        students = students.filter(stream=stream)

    # ── Lessons within term
    lessons_qs = Lesson.objects.filter(
        timetable__school=school,
        is_canceled=False
    )
    if stream:
        lessons_qs = lessons_qs.filter(stream=stream)
    if term:
        lessons_qs = lessons_qs.filter(
            lesson_date__gte=term.start_date,
            lesson_date__lte=term.end_date
        )

    # Map lessons by subject
    lessons_map = defaultdict(list)
    for lesson in lessons_qs.select_related('subject', 'stream'):
        lessons_map[lesson.subject_id].append(lesson)

    # Prefetch student's subject enrollments
    students = students.prefetch_related('subject_enrollments__subject')

    enrollments_to_create = []

    for student in students:
        try:
            subject_ids = list(student.subject_enrollments.values_list('subject_id', flat=True))
            if not subject_ids:
                logger.warning(f"Student {student.student_id} has no subject enrollments")
                continue

            # For each subject, get all lessons for this term & stream
            for sid in subject_ids:
                lessons = lessons_map.get(sid, [])
                for lesson in lessons:
                    enrollments_to_create.append(
                        Enrollment(
                            student=student,
                            lesson=lesson,
                            school=school,
                            status='active'
                        )
                    )

        except Exception as e:
            logger.exception(f"Failed to process enrollments for student {student.student_id}")

    if enrollments_to_create:
        Enrollment.objects.bulk_create(enrollments_to_create, ignore_conflicts=True)

    # Debug print
    total_students = students.count()
    total_lessons = lessons_qs.count()
    total_enrollments = len(enrollments_to_create)
    logger.info(f"[SUMMARY] Students={total_students} Lessons={total_lessons} Enrollments={total_enrollments}")



SUBJECT_CODE_MAP = {
    "MAT": "Mathematics",
    "CMAT": "Core Maths",
    "EMAT": "Essential Maths",
    "ENG": "English",
    "KISW": "Kiswahili",
    "CHEM": "Chemistry",
    "BIO": "Biology",
    "PHY": "Physics",
    "CRE": "Christian Religious Education",
    "GEO": "Geography",
    "HIST": "History",
    "BST": "Business Studies",
    "AGRI": "Agriculture",
    "MUDA": "Music & Dance",
    "LIT": "Literature",
    "FAS": "Fasihi",
    "TF": "Theatre and Film",
    "GENSCI": "General Science",
    "CSL": "Community Service Learning",
    "HSCI" : "Home Science",
    "ICT": "Information & Communication Technology",
    "PPI": "Pastoral Programme of Instruction",
    "GRPS": "Guidance & Pastoral Studies",
    "PE": "Physical Education",
    "LS": "Life Skills",
    "COMP": "Computer Studies",
}


# ================= HELPERS =================

def normalize_phone(phone):
    """Normalize Kenyan phone numbers to consistent format: 07XXXXXXXX"""
    if not phone:
        return ""
    
    phone = str(phone).strip()
    
    # Remove all non-digit characters
    phone = ''.join(filter(str.isdigit, phone))
    
    if not phone:
        return ""
    
    # Convert to 10-digit format starting with 0
    if phone.startswith("254"):
        phone = phone[3:]          # remove 254
    elif phone.startswith("+254"):
        phone = phone[4:]          # remove +254
    
    if len(phone) == 9 and not phone.startswith("0"):
        phone = "0" + phone        # 712345678 → 0712345678
    
    elif len(phone) == 10 and phone.startswith("0"):
        pass                       # already good (0712345678)
    
    else:
        # Invalid length → return as is (will fail lookup gracefully)
        return phone
    
    return phone

def generate_lesson_dates(term, weekday):
    weekday_map = {
        "monday": 0,
        "tuesday": 1,
        "wednesday": 2,
        "thursday": 3,
        "friday": 4,
        "saturday": 5,
        "sunday": 6,
    }

    target = weekday_map[weekday.lower()]
    current = term.start_date

    while current.weekday() != target:
        current += timedelta(days=1)

    dates = []
    while current <= term.end_date:
        dates.append(current)
        current += timedelta(days=7)

    return dates



def normalize_weekday(day_str: str) -> str | None:
    day_str = day_str.strip().lower()
    mapping = {
        'mon': 'monday', 'tue': 'tuesday', 'wed': 'wednesday',
        'thu': 'thursday', 'thur': 'thursday', 'fri': 'friday',
        'sat': 'saturday', 'sun': 'sunday'
    }
    for short, full in mapping.items():
        if day_str.startswith(short):
            return full
    return day_str if day_str in {'monday','tuesday','wednesday','thursday','friday','saturday','sunday'} else None


def norm_time(t):
    """Robust time parser handling AM/PM, HH:MM:SS, H:MM, etc."""
    if pd.isna(t) or not str(t).strip():
        return None

    t = str(t).strip().upper().replace(' ', '')
    am_pm = ''
    if 'AM' in t:
        am_pm = 'AM'
        t = t.replace('AM', '')
    elif 'PM' in t:
        am_pm = 'PM'
        t = t.replace('PM', '')

    t = t.replace('.', ':').replace(';', ':')

    try:
        if ':' in t:
            parts = t.split(':')
            h = parts[0].zfill(2)
            m = parts[1][:2].zfill(2) if len(parts) > 1 else '00'
            time_str = f"{h}:{m}"

            h_int = int(h)
            if am_pm == 'PM' and h_int != 12:
                h_int += 12
            elif am_pm == 'AM' and h_int == 12:
                h_int = 0

            return f"{h_int:02d}:{m}"

        if len(t) == 4 and t.isdigit():
            h = t[:2]
            m = t[2:]
            return f"{h}:{m}"

    except Exception:
        pass

    return None


def _run_excel_processing(df, category, school, user, results):
    """
    Pure data-processing logic extracted from universal_excel_upload.
    Mutates `results` dict in-place. Called from background thread.
    """
    staff = StaffProfile.objects.filter(user=user, school=school).first()
    active_term = Term.objects.filter(school=school, is_active=True).first()

    with transaction.atomic():

        # ====================== STUDENTS ======================
        if category == "students":
            if not active_term:
                results["fatal"] = "No active term found"
                return

            for idx, row in df.iterrows():
                try:
                    with transaction.atomic():
                        admin_no = str(row.get("admin_no", "")).strip()
                        if not admin_no:
                            continue  # blank row — skip silently

                        student_email = f"{admin_no}@student.school.local"
                        student_password = generate_password()
                        first_name = str(row.get("first_name", "")).strip()
                        last_name = str(row.get("last_name", "")).strip()
                        gender = str(row.get("gender", "")).strip().upper()

                        user_obj, user_created = User.objects.get_or_create(
                            email=student_email,
                            defaults={
                                "phone_number": admin_no,
                                "first_name": first_name,
                                "last_name": last_name,
                                "is_student": True,
                            }
                        )
                        if user_created:
                            user_obj.set_password(student_password)
                            user_obj.save()

                        grade_name = str(row.get("grade", "")).strip()
                        stream_name = str(row.get("stream", "")).strip()
                        grade = Grade.objects.filter(name__iexact=grade_name, school=school).first()
                        stream = Streams.objects.filter(name__iexact=stream_name, school=school).first()

                        if not grade:
                            raise ValueError(f"Grade '{grade_name}' not found for this school")
                        if not stream:
                            raise ValueError(f"Stream '{stream_name}' not found in {grade_name}")

                        pathway_obj = None
                        pathway_raw = str(row.get("pathway", "")).strip()
                        if pathway_raw and grade:
                            normalized_csv = pathway_raw.lower().replace(".", "").replace("&", "and").replace(" ", "")
                            for p in Pathway.objects.filter(school=school, grade=grade):
                                normalized_db = p.name.strip().lower().replace(".", "").replace("&", "and").replace(" ", "")
                                if normalized_db == normalized_csv:
                                    pathway_obj = p
                                    break

                        student, created = Student.objects.update_or_create(
                            user=user_obj,
                            student_id=admin_no,
                            school=school,
                            defaults={
                                "date_of_birth": row.get("date_of_birth"),
                                "gender": gender,
                                "enrollment_date": timezone.now().date(),
                                "grade_level": grade,
                                "stream": stream,
                                "pathway": pathway_obj,
                            }
                        )

                        subject_codes = [
                            c.strip().upper()
                            for c in str(row.get("subjects", "")).split(",")
                            if c.strip()
                        ]
                        for code in subject_codes:
                            # Only enroll in subjects the school has already activated from catalog
                            subject = Subject.objects.filter(
                                school=school, is_active=True
                            ).filter(
                                Q(code__iexact=code) | Q(name__iexact=code)
                            ).first()
                            if not subject:
                                results["warnings"].append({
                                    "row": idx + 2,
                                    "message": (
                                        f"Subject '{code}' not found in school's activated subjects. "
                                        f"Go to Subjects → 'From Catalog' to activate it first."
                                    ),
                                })
                                continue
                            SubjectEnrollment.objects.get_or_create(student=student, subject=subject)

                        parent_phone = str(row.get("parent_phone", "")).strip()
                        if parent_phone:
                            parent_email = str(row.get("parent_email", "")).strip() or f"{parent_phone}@parent.school.local"
                            parent_first_name = str(row.get("parent_first_name", "")).strip()
                            parent_last_name = str(row.get("parent_last_name", "")).strip()
                            parent_password = generate_password()

                            parent_user, parent_user_created = User.objects.get_or_create(
                                email=parent_email,
                                defaults={
                                    "phone_number": parent_phone,
                                    "first_name": parent_first_name,
                                    "last_name": parent_last_name,
                                    "is_parent": True,
                                }
                            )
                            if parent_user_created:
                                parent_user.set_password(parent_password)
                                parent_user.save()

                            parent_obj, _ = Parent.objects.get_or_create(
                                phone=parent_phone,
                                defaults={
                                    "user": parent_user,
                                    "parent_id": parent_phone,
                                    "school": school,
                                }
                            )
                            if not parent_obj.user:
                                parent_obj.user = parent_user
                                parent_obj.save()

                            student.parents.add(parent_obj)

                        results["created"] += 1

                except Exception as e:
                    logger.exception(f"Student row {idx + 2} failed")
                    results["errors"].append({"row": idx + 2, "error": str(e)})

        # ====================== TEACHERS ======================
        elif category == "teachers":
            required = ["phone", "email", "first_name", "last_name", "subjects", "streams", "class_teacher"]
            missing = [c for c in required if c not in df.columns]
            if missing:
                results["fatal"] = f"Missing columns: {missing}"
                return

            for idx, row in df.iterrows():
                try:
                    if not str(row.get("phone", "")).strip():
                        continue  # blank row — skip silently
                    phone = normalize_phone(row["phone"])
                    email = safe_email(row["email"], f"{phone}@gmail.com")
                    password = generate_password()

                    user_obj, created = User.objects.get_or_create(
                        email=email,
                        defaults={
                            "phone_number": phone,
                            "first_name": str(row["first_name"]).strip(),
                            "last_name": str(row["last_name"]).strip(),
                            "is_teacher": True,
                        }
                    )
                    if created:
                        user_obj.set_password(password)
                        user_obj.save()

                    staff_id = f"TCHR{str(user_obj.id).zfill(5)}"
                    staff_profile, _ = StaffProfile.objects.get_or_create(
                        user=user_obj,
                        school=school,
                        defaults={"staff_id": staff_id, "position": "teacher"},
                    )

                    for code in str(row["subjects"]).replace(" ", "").split(","):
                        code = code.strip().upper()
                        if code:
                            subj = Subject.objects.filter(
                                school=school, is_active=True
                            ).filter(
                                Q(code__iexact=code) | Q(name__iexact=code)
                            ).first()
                            if subj:
                                staff_profile.subjects.add(subj)
                            else:
                                results["warnings"].append({
                                    "row": idx + 2,
                                    "message": (
                                        f"Subject '{code}' not found in school's activated subjects "
                                        f"— skipped for teacher {email}."
                                    ),
                                })

                    if str(row.get("class_teacher", "")).upper() == "YES":
                        stream_names = [s.strip() for s in str(row["streams"]).replace(" ", "").split(",") if s.strip()]
                        streams = Streams.objects.filter(name__in=stream_names, school=school)
                        for stream in streams:
                            TeacherStreamAssignment.objects.get_or_create(
                                teacher=staff_profile, stream=stream, school=school
                            )

                    results["created"] += 1

                except Exception as e:
                    logger.exception(f"Teacher row {idx + 2} failed")
                    results["errors"].append({"row": idx + 2, "error": str(e)})

        # ====================== GRADES ======================
        elif category == "grades":
            required = ["grade_name", "code", "lessons_per_term", "capacity", "stream_name"]
            missing = [c for c in required if c not in df.columns]
            if missing:
                results["fatal"] = f"Missing columns: {missing}"
                return

            for idx, row in df.iterrows():
                try:
                    grade_code = str(row["code"]).strip().upper()
                    grade_name = str(row["grade_name"]).strip()
                    raw_streams = str(row["stream_name"]).strip()

                    if not raw_streams:
                        raise ValueError("stream_name is required")

                    stream_names = [s.strip().upper() for s in raw_streams.split(",") if s.strip()]

                    grade, _ = Grade.objects.update_or_create(
                        school=school,
                        code=grade_code,
                        defaults={
                            "name": grade_name,
                            "lessons_per_term": int(row["lessons_per_term"]),
                            "capacity": int(row["capacity"]),
                            "description": str(row.get("description", "")),
                            "is_active": True,
                        }
                    )

                    stream_capacity = int(row.get("stream_capacity") or grade.capacity)
                    for stream_name in stream_names:
                        Streams.objects.update_or_create(
                            school=school,
                            grade=grade,
                            name=stream_name,
                            defaults={"capacity": stream_capacity, "is_active": True},
                        )

                    results["created"] += 1

                except Exception as e:
                    results["errors"].append({"row": idx + 2, "error": str(e)})

        # ====================== TIME SLOTS ======================
        elif category == "time_slots":
            for idx, row in df.iterrows():
                try:
                    start_time = safe_parse_time(row["start_time"])
                    end_time = safe_parse_time(row["end_time"])
                    if not start_time or not end_time:
                        raise ValueError("Invalid time format")

                    TimeSlot.objects.update_or_create(
                        school=school,
                        start_time=start_time,
                        end_time=end_time,
                        defaults={"description": str(row.get("description", "")), "updated_by": staff}
                    )
                    results["created"] += 1
                except Exception as e:
                    results["errors"].append({"row": idx + 2, "error": str(e)})

        # ====================== PARENTS ======================
        elif category == "parents":
            phone_col = next((c for c in ["mobile", "phone", "phone_number", "contact", "telephone"] if c in df.columns), None)
            required = ["first_name", "last_name", "admin_no"]
            missing = [c for c in required if c not in df.columns]
            if not phone_col:
                missing.append("mobile (or phone / phone_number)")
            if missing:
                results["fatal"] = f"Missing columns: {missing}"
                return

            for idx, row in df.iterrows():
                row_num = idx + 2
                try:
                    with transaction.atomic():
                        phone = normalize_phone(row[phone_col])
                        admin_no = str(row["admin_no"]).strip()
                        email = f"{phone}@parent.school.local"
                        password = generate_password()

                        user_obj = User.objects.filter(phone_number=phone).first()
                        if not user_obj:
                            user_obj = User.objects.create_user(
                                email=email,
                                phone_number=phone,
                                first_name=str(row["first_name"]).strip(),
                                last_name=str(row["last_name"]).strip(),
                                is_parent=True,
                                password=password,
                            )

                        parent, _ = Parent.objects.get_or_create(
                            user=user_obj,
                            defaults={
                                "parent_id": f"PARENT{user_obj.id}",
                                "phone": phone,
                                "school": school,
                            }
                        )

                        student = Student.objects.filter(student_id=admin_no, school=school).first()
                        if student:
                            student.parents.add(parent)
                        else:
                            results["warnings"].append({
                                "row": row_num,
                                "message": f"Student with admin_no '{admin_no}' not found"
                            })

                        results["created"] += 1

                except Exception as e:
                    logger.exception(f"Parent row {row_num} failed")
                    results["errors"].append({"row": row_num, "error": str(e)})

        # ====================== TIMETABLE ======================
        elif category == "timetable":
            required = ["grade", "stream", "term", "year", "subject_code", "room", "weekdays", "start_time", "end_time"]
            missing = [c for c in required if c not in df.columns]
            if missing:
                results["fatal"] = f"Missing columns: {', '.join(missing)}"
                return

            first = df.iloc[0]
            term_name = str(first.get("term", "")).strip()
            excel_year_value = int(str(first.get("year")).strip())

            term = Term.objects.filter(name=term_name, school=school, is_active=True).first()
            if not term:
                results["fatal"] = f"No active term named '{term_name}' found for this school. Please activate the term first."
                return

            grade_name = str(first.get("grade", "")).strip()
            stream_name = str(first.get("stream", "")).strip()
            grade = Grade.objects.get(name=grade_name, school=school)
            stream = Streams.objects.get(name=stream_name, school=school, grade=grade)

            timetable, _ = Timetable.objects.get_or_create(
                school=school, grade=grade, stream=stream, term=term, year=excel_year_value,
                defaults={"start_date": term.start_date, "end_date": term.end_date}
            )

            subject_map = {s.code.upper(): s for s in Subject.objects.filter(school=school)}
            staff_map = {}
            for s in StaffProfile.objects.filter(school=school).select_related('user'):
                if s.user and s.user.phone_number:
                    norm_phone = normalize_phone(s.user.phone_number)
                    staff_map[norm_phone] = s
                    if norm_phone.startswith("0"):
                        staff_map[norm_phone[1:]] = s

            slot_cache = {
                (slot.start_time.strftime('%H:%M'), slot.end_time.strftime('%H:%M')): slot
                for slot in TimeSlot.objects.filter(school=school)
                if slot.start_time and slot.end_time
            }

            lessons_to_create = []
            seen_patterns = set()
            weekday_counts = defaultdict(int)

            for idx, row in df.iterrows():
                row_num = idx + 2
                try:
                    subj_code = str(row["subject_code"]).strip().upper()
                    room = str(row.get("room", "")).strip() or "N/A"
                    subject = subject_map.get(subj_code)
                    if not subject:
                        raise ValueError(f"Subject '{subj_code}' not found")

                    teacher = None
                    teacher_phone_raw = None
                    for col in ["teacher_phone", "teacher", "phone", "teacherphone", "mobile", "tutor"]:
                        if col in df.columns:
                            teacher_phone_raw = row.get(col)
                            if teacher_phone_raw and str(teacher_phone_raw).strip():
                                break

                    if teacher_phone_raw:
                        phone = normalize_phone(teacher_phone_raw)
                        if phone:
                            teacher = staff_map.get(phone)
                            if not teacher:
                                results["warnings"].append({
                                    "row": row_num,
                                    "message": f"Teacher phone '{phone}' not found in StaffProfile"
                                })

                    raw_day = str(row["weekdays"]).strip().lower()
                    wd = normalize_weekday(raw_day)
                    if not wd:
                        raise ValueError(f"Invalid weekday: '{raw_day}'")

                    start_norm = norm_time(row["start_time"])
                    end_norm = norm_time(row["end_time"])
                    if not start_norm or not end_norm:
                        raise ValueError(f"Invalid time format")

                    time_slot = slot_cache.get((start_norm, end_norm))
                    if not time_slot:
                        results["warnings"].append({
                            "row": row_num,
                            "message": f"Time slot {start_norm}–{end_norm} not found"
                        })
                        continue

                    pattern_key = (wd, time_slot.id, subject.id, teacher.id if teacher else None, room)
                    if pattern_key in seen_patterns:
                        continue
                    seen_patterns.add(pattern_key)

                    current_date = term.start_date
                    while current_date <= term.end_date:
                        if current_date.strftime("%A").lower() == wd:
                            break
                        current_date += timedelta(days=1)

                    while current_date <= term.end_date:
                        lessons_to_create.append(
                            Lesson(
                                timetable=timetable, subject=subject, stream=stream,
                                teacher=teacher, day_of_week=wd,
                                time_slot=time_slot, room=room, lesson_date=current_date,
                            )
                        )
                        weekday_counts[wd] += 1
                        current_date += timedelta(days=7)

                except Exception as e:
                    logger.exception(f"[ROW {row_num}] Failed")
                    results["errors"].append({"row": row_num, "error": str(e)})

            Lesson.objects.filter(timetable=timetable).delete()
            if lessons_to_create:
                Lesson.objects.bulk_create(lessons_to_create, ignore_conflicts=True)
                results["created"] = len(lessons_to_create)

            transaction.on_commit(
                lambda: populate_student_lesson_enrollments(school, grade, stream, term)
            )

            results["summary"] = {
                "grade": grade_name, "stream": stream_name, "term": term_name,
                "year": excel_year_value, "lessons_generated": len(lessons_to_create),
                "weekdays": dict(weekday_counts),
            }
        else:
            results["fatal"] = "Invalid category"


def _process_upload_in_background(upload_id, user_id, school_id, category):
    """Thread target: process saved Excel file, update Upload record, email uploader."""
    from django.db import connection as _conn, close_old_connections
    close_old_connections()

    try:
        upload = Upload.objects.get(pk=upload_id)
        upload.status = 'processing'
        upload.save(update_fields=['status'])

        user   = User.objects.get(pk=user_id)
        school = School.objects.get(pk=school_id)

        try:
            df = pd.read_excel(upload.file.path).fillna("")
            df.columns = df.columns.str.strip().str.lower().str.replace(r'[^a-z0-9_]', '_', regex=True)
        except Exception as e:
            upload.status = 'failed'
            upload.result_json = {"fatal": f"Cannot read file: {e}", "created": 0, "errors": [], "warnings": []}
            upload.completed_at = timezone.now()
            upload.save(update_fields=['status', 'result_json', 'completed_at'])
            _send_upload_email(user, upload, category)
            return

        results = {"created": 0, "errors": [], "warnings": []}
        try:
            _run_excel_processing(df, category, school, user, results)
            upload.status = 'done' if not results.get("fatal") else 'failed'
        except Exception as e:
            logger.exception(f"[BG UPLOAD #{upload_id}] Unhandled error")
            results["fatal"] = str(e)
            upload.status = 'failed'

        upload.result_json = results
        upload.completed_at = timezone.now()
        upload.save(update_fields=['status', 'result_json', 'completed_at'])
        _send_upload_email(user, upload, category)

    except Exception as e:
        logger.exception(f"[BG UPLOAD #{upload_id}] Thread-level failure: {e}")
    finally:
        _conn.close()


def _send_upload_email(user, upload, category):
    """Send completion email to the uploader."""
    result  = upload.result_json or {}
    created = result.get("created", 0)
    errors  = result.get("errors", [])
    warnings = result.get("warnings", [])
    fatal   = result.get("fatal", "")
    status  = upload.status

    if status == 'done':
        subject = f"[Kiswate] Upload Complete — {created} {category} record(s) imported"
        colour  = "#065f46"
        icon    = "✅"
        headline = f"{created} record(s) imported successfully"
    else:
        subject = f"[Kiswate] Upload Failed — {category}"
        colour  = "#991b1b"
        icon    = "❌"
        headline = fatal or "The upload could not be completed"

    error_lines = "".join(
        f"<li>Row {e['row']}: {e['error']}</li>" for e in errors[:20]
    )
    warn_lines = "".join(
        f"<li>Row {w.get('row','?')}: {w.get('message','')}</li>" for w in warnings[:10]
    )

    html = f"""
    <div style="font-family:Arial,sans-serif;max-width:600px;margin:0 auto;padding:24px;">
      <div style="background:{colour};border-radius:10px;padding:20px 24px;color:#fff;margin-bottom:20px;">
        <h2 style="margin:0;font-size:20px;">{icon} {headline}</h2>
        <p style="margin:6px 0 0;opacity:.85;font-size:14px;">
          Category: <strong>{category.replace('_',' ').title()}</strong> &nbsp;·&nbsp;
          Upload #{upload.pk}
        </p>
      </div>

      <table style="width:100%;border-collapse:collapse;font-size:14px;margin-bottom:20px;">
        <tr style="background:#f9fafb;">
          <td style="padding:8px 12px;font-weight:600;border:1px solid #e5e7eb;">Records imported</td>
          <td style="padding:8px 12px;border:1px solid #e5e7eb;">{created}</td>
        </tr>
        <tr>
          <td style="padding:8px 12px;font-weight:600;border:1px solid #e5e7eb;">Errors</td>
          <td style="padding:8px 12px;border:1px solid #e5e7eb;color:{'#991b1b' if errors else '#065f46'};">{len(errors)}</td>
        </tr>
        <tr style="background:#f9fafb;">
          <td style="padding:8px 12px;font-weight:600;border:1px solid #e5e7eb;">Warnings</td>
          <td style="padding:8px 12px;border:1px solid #e5e7eb;">{len(warnings)}</td>
        </tr>
        <tr>
          <td style="padding:8px 12px;font-weight:600;border:1px solid #e5e7eb;">Completed at</td>
          <td style="padding:8px 12px;border:1px solid #e5e7eb;">{upload.completed_at.strftime('%d %b %Y %H:%M') if upload.completed_at else '—'}</td>
        </tr>
      </table>

      {'<h3 style="color:#991b1b;font-size:15px;">Row Errors (first 20)</h3><ul style="font-size:13px;color:#374151;">' + error_lines + '</ul>' if errors else ''}
      {'<h3 style="color:#92400e;font-size:15px;">Warnings (first 10)</h3><ul style="font-size:13px;color:#374151;">' + warn_lines + '</ul>' if warnings else ''}

      <p style="font-size:12px;color:#9ca3af;margin-top:24px;border-top:1px solid #e5e7eb;padding-top:12px;">
        Kiswate Digital · This is an automated message, please do not reply.
      </p>
    </div>
    """

    try:
        send_mail(
            subject=subject,
            message=f"Upload #{upload.pk} ({category}): {created} imported, {len(errors)} errors.",
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[user.email],
            html_message=html,
            fail_silently=True,
        )
        logger.info(f"[UPLOAD #{upload.pk}] Completion email sent to {user.email}")
    except Exception as e:
        logger.error(f"[UPLOAD #{upload.pk}] Failed to send email: {e}")


@require_POST
@login_required
def universal_excel_upload(request):
    """
    Accepts the file, saves it to the Upload model, and immediately returns
    {"status": "queued"}. Heavy processing runs in a background thread and
    sends an email to the uploader on completion.
    """
    excel_file = request.FILES.get("file")
    category = request.POST.get("category")

    if not excel_file or not category:
        return JsonResponse({"error": "File and category are required"}, status=400)

    try:
        _validate_excel_upload(excel_file)
    except ValueError as e:
        return JsonResponse({"error": str(e)}, status=400)

    if category == "subjects":
        return JsonResponse({
            "error": (
                "Schools cannot upload subjects directly. "
                "Go to Subjects → 'From Catalog' to activate subjects your school will offer."
            )
        }, status=400)

    user = request.user
    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        return JsonResponse({"error": "Permission denied."}, status=403)

    school = get_user_school(user)
    if not school:
        return JsonResponse({"error": "No school linked to your account."}, status=400)

    # Save file immediately; processing happens in background thread
    try:
        upload_record = Upload.objects.create(
            file=excel_file,
            school=school,
            uploaded_by=user,
            category=category,
            status='pending',
        )
        logger.info(f"File saved: {upload_record.file.name} | ID: {upload_record.id}")
    except Exception as e:
        logger.error(f"Failed to save uploaded file: {e}")
        return JsonResponse({"error": f"Failed to save file: {e}"}, status=500)

    # Kick off background thread
    t = threading.Thread(
        target=_process_upload_in_background,
        args=(upload_record.id, user.id, school.id, category),
        daemon=True,
    )
    t.start()

    return JsonResponse({
        "status": "queued",
        "upload_id": upload_record.id,
        "message": (
            f"Your {category.replace('_',' ')} file has been received and is being processed. "
            f"You will receive an email at {user.email} when it completes."
        ),
    })


@login_required
def upload_job_status(request, pk):
    """Poll endpoint: returns current status of an upload job."""
    upload = get_object_or_404(Upload, pk=pk, uploaded_by=request.user)
    result = upload.result_json or {}
    return JsonResponse({
        "upload_id": upload.pk,
        "status": upload.status,
        "category": upload.category,
        "created": result.get("created", 0),
        "error_count": len(result.get("errors", [])),
        "warning_count": len(result.get("warnings", [])),
        "errors": result.get("errors", [])[:20],
        "warnings": result.get("warnings", [])[:10],
        "fatal": result.get("fatal", ""),
        "summary": result.get("summary", {}),
        "completed_at": upload.completed_at.isoformat() if upload.completed_at else None,
    })



@login_required
def upload_excel_page(request):
    """Renders the Excel upload page."""
    user = request.user
    school = get_user_school(user)
    return render(request, 'school/file_upload.html', {'school': school})


# =============================================================================
# SUBJECT CATALOG — Kiswate platform admin manages the global subject catalog.
# Principals then activate subjects from it for their school.
# =============================================================================

@login_required
def catalog_subject_list(request):
    """Kiswate admin: view and manage the platform subject catalog."""
    if not (request.user.is_kiswate_user or request.user.is_kiswate_admin or request.user.is_superuser):
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    query = request.GET.get('q', '').strip()
    qs = SubjectCatalog.objects.annotate(school_count=Count('school_subjects')).order_by('name')
    if query:
        qs = qs.filter(Q(name__icontains=query) | Q(code__icontains=query))

    paginator = Paginator(qs, 25)
    catalog_list = paginator.get_page(request.GET.get('page'))
    form = SubjectCatalogForm()
    upload_form = BulkSubjectUploadForm()

    is_kiswate = request.user.is_kiswate_user or request.user.is_kiswate_admin or request.user.is_superuser
    return render(request, 'school/catalog/subjects.html', {
        'catalog_list': catalog_list,
        'form': form,
        'upload_form': upload_form,
        'query': query,
        'base_template': 'Dashboard/base.html' if is_kiswate else 'school/base.html',
    })


@login_required
def catalog_subject_create(request):
    if not (request.user.is_kiswate_user or request.user.is_kiswate_admin or request.user.is_superuser):
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    if request.method == 'POST':
        form = SubjectCatalogForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f'"{form.cleaned_data["name"]}" added to catalog.')
        else:
            for field, errs in form.errors.items():
                for e in errs:
                    messages.error(request, f"{field}: {e}")
    return redirect('school:catalog-subject-list')


@login_required
def catalog_subject_edit(request, pk):
    if not (request.user.is_kiswate_user or request.user.is_kiswate_admin or request.user.is_superuser):
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    entry = get_object_or_404(SubjectCatalog, pk=pk)
    if request.method == 'POST':
        form = SubjectCatalogForm(request.POST, instance=entry)
        if form.is_valid():
            form.save()
            messages.success(request, f'"{entry.name}" updated.')
        else:
            for field, errs in form.errors.items():
                for e in errs:
                    messages.error(request, f"{field}: {e}")
    return redirect('school:catalog-subject-list')


@login_required
def catalog_subject_delete(request, pk):
    if not (request.user.is_kiswate_user or request.user.is_kiswate_admin or request.user.is_superuser):
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    entry = get_object_or_404(SubjectCatalog, pk=pk)
    if request.method == 'POST':
        name = entry.name
        entry.delete()
        messages.success(request, f'"{name}" removed from catalog.')
    return redirect('school:catalog-subject-list')


@login_required
def catalog_subject_toggle(request, pk):
    """Quick AJAX toggle for is_active on a catalog entry."""
    if not (request.user.is_kiswate_user or request.user.is_kiswate_admin or request.user.is_superuser):
        return JsonResponse({'error': 'Access denied'}, status=403)
    entry = get_object_or_404(SubjectCatalog, pk=pk)
    entry.is_active = not entry.is_active
    entry.save(update_fields=['is_active'])
    return JsonResponse({'is_active': entry.is_active})


# =============================================================================
# SUBJECT SELECTION — Principal activates catalog subjects for their school.
# =============================================================================

@login_required
def subject_activate_from_catalog(request):
    """Principal/admin: pick catalog subjects to activate for this school."""
    user = request.user
    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    school = get_user_school(user)
    if not school:
        messages.error(request, "No school profile found.")
        return redirect('school:dashboard')

    active_term = Term.objects.filter(school=school, is_active=True).first()
    default_start = active_term.start_date if active_term else timezone.localdate()

    form = SubjectActivationForm(school=school, initial={'start_date': default_start})

    if request.method == 'POST':
        form = SubjectActivationForm(request.POST, school=school)
        if form.is_valid():
            created = form.activate(school)
            if created:
                messages.success(
                    request,
                    f'{len(created)} subject(s) activated: '
                    + ', '.join(s.name for s in created)
                    + '. Set the grade(s) for each subject below.'
                )
            else:
                messages.info(request, 'Selected subjects were already active for your school.')
            return redirect('school:school-subjects')
        else:
            messages.error(request, 'Please correct the errors below.')

    already_active = Subject.objects.filter(school=school, catalog_ref__isnull=False).select_related('catalog_ref')
    catalog_groups = {}
    for entry in SubjectCatalog.objects.filter(is_active=True).order_by('curriculum', 'name'):
        catalog_groups.setdefault(entry.get_curriculum_display(), []).append(entry)

    return render(request, 'school/subject_selection.html', {
        'form': form,
        'school': school,
        'catalog_groups': catalog_groups,
        'already_active': already_active,
    })


# =============================================================================
# BULK SUBJECT UPLOAD — Kiswate admin uploads catalog subjects via Excel
# =============================================================================

@login_required
def catalog_subject_bulk_upload(request):
    """Parse an uploaded Excel file and create SubjectCatalog entries."""
    if not (request.user.is_kiswate_user or request.user.is_kiswate_admin or request.user.is_superuser):
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    if request.method != 'POST':
        return redirect('school:catalog-subject-list')

    form = BulkSubjectUploadForm(request.POST, request.FILES)
    if not form.is_valid():
        for field, errs in form.errors.items():
            for e in errs:
                messages.error(request, f"{e}")
        return redirect('school:catalog-subject-list')

    file = form.cleaned_data['file']
    try:
        df = pd.read_excel(file, dtype=str)
    except Exception as e:
        messages.error(request, f"Could not read file: {e}")
        return redirect('school:catalog-subject-list')

    required_cols = {'name', 'code'}
    missing = required_cols - set(c.lower().strip() for c in df.columns)
    if missing:
        messages.error(request, f"Missing required columns: {', '.join(missing)}")
        return redirect('school:catalog-subject-list')

    df.columns = [c.lower().strip() for c in df.columns]

    valid_curricula = {'cbc', 'kcse', 'kcpe', 'igcse', 'other'}
    created_count = updated_count = skipped_count = 0

    for idx, row in df.iterrows():
        name = str(row.get('name', '')).strip()
        code = str(row.get('code', '')).strip()
        if not name or not code or name == 'nan' or code == 'nan':
            skipped_count += 1
            continue

        curriculum = str(row.get('curriculum', 'cbc')).strip().lower()
        if curriculum not in valid_curricula:
            curriculum = 'cbc'

        def _bool(val, default=True):
            s = str(val).strip().lower()
            if s in ('true', '1', 'yes'):
                return True
            if s in ('false', '0', 'no'):
                return False
            return default

        defaults = {
            'curriculum': curriculum,
            'is_core': _bool(row.get('is_core', 'TRUE')),
            'is_elective': _bool(row.get('is_elective', 'FALSE'), default=False),
            'sessions_per_week_default': max(1, int(float(row.get('sessions_per_week_default', 2) or 2))),
            'description': str(row.get('description', '')).strip() if row.get('description') else '',
            'is_active': True,
        }

        try:
            obj, created = SubjectCatalog.objects.update_or_create(
                code=code,
                defaults={**defaults, 'name': name},
            )
            if created:
                created_count += 1
            else:
                updated_count += 1
        except Exception:
            skipped_count += 1

    messages.success(
        request,
        f"Import complete: {created_count} added, {updated_count} updated, {skipped_count} skipped."
    )
    return redirect('school:catalog-subject-list')


# =============================================================================
# PARENT SELF-SERVICE PORTAL
# =============================================================================

def _get_parent(user):
    try:
        return user.parent
    except Exception:
        return None


@login_required
def parent_dashboard(request):
    """Rich parent dashboard with per-child stats for parents with 1+ children."""
    if not getattr(request.user, 'is_parent', False):
        messages.error(request, "This section is for parents only.")
        return redirect('userauths:sign-in')

    parent = _get_parent(request.user)
    if not parent:
        messages.error(request, "Parent profile not found.")
        return redirect('userauths:sign-in')

    children = parent.children.select_related('user', 'grade_level', 'stream', 'school').all()
    unread_count = request.user.notifications.filter(is_read=False).count()
    open_complaints = Complaint.objects.filter(parent=parent, status='open').count()

    today = timezone.now().date()
    next_week = today + timedelta(days=7)
    total_balance = Decimal('0.00')
    child_data = []

    for child in children:
        # Attendance stats
        att_qs = Attendance.objects.filter(enrollment__student=child)
        total_att = att_qs.count()
        present_att = att_qs.filter(status='P').count()
        att_rate = int((present_att / total_att * 100) if total_att > 0 else 0)

        # Fee balance
        fee_balance = FeeInvoice.objects.filter(
            student=child
        ).exclude(status='paid').aggregate(
            bal=Sum(F('amount_required') - F('amount_paid'))
        )['bal'] or Decimal('0.00')
        total_balance += fee_balance

        # Discipline records
        discipline_count = DisciplineRecord.objects.filter(student=child, resolved=False).count()

        # Upcoming lessons (next 7 days)
        upcoming_lessons = Lesson.objects.filter(
            stream=child.stream,
            lesson_date__gte=today,
            lesson_date__lte=next_week,
        ).select_related('subject', 'teacher__user', 'time_slot').order_by('lesson_date', 'time_slot__start_time')[:8] if child.stream else []

        # Announcements for child's school/grade (active only)
        _now = timezone.now()
        announcements = Announcement.objects.filter(
            school=child.school,
            audience__in=['all', 'parents'],
        ).filter(
            Q(grade=child.grade_level) | Q(grade__isnull=True)
        ).filter(
            Q(expires_at__isnull=True) | Q(expires_at__gt=_now)
        ).order_by('-is_pinned', '-created_at')[:4]

        # Active term
        active_term = Term.objects.filter(school=child.school, is_active=True).first()

        # Student status
        if child.expelled:
            status_color, status_label = 'danger', 'Expelled'
        elif child.suspended:
            status_color, status_label = 'warning', 'Suspended'
        elif child.is_active:
            status_color, status_label = 'success', 'Active'
        else:
            status_color, status_label = 'secondary', 'Inactive'

        child_data.append({
            'student': child,
            'school': child.school,
            'att_rate': att_rate,
            'present_att': present_att,
            'total_att': total_att,
            'fee_balance': fee_balance,
            'discipline_count': discipline_count,
            'upcoming_lessons': upcoming_lessons,
            'announcements': announcements,
            'active_term': active_term,
            'status_color': status_color,
            'status_label': status_label,
        })

    return render(request, 'school/parent/dashboard.html', {
        'parent': parent,
        'child_data': child_data,
        'total_balance': total_balance,
        'unread_count': unread_count,
        'open_complaints': open_complaints,
        'today': today,
    })


@login_required
def parent_portal(request):
    return redirect('school:parent-dashboard')


@login_required
def parent_notifications(request):
    """Parent: view and manage notifications."""
    if not getattr(request.user, 'is_parent', False):
        messages.error(request, "Access denied.")
        return redirect('userauths:sign-in')

    if request.method == 'POST' and request.POST.get('action') == 'mark_all_read':
        request.user.notifications.filter(is_read=False).update(is_read=True)
        messages.success(request, "All notifications marked as read.")
        return redirect('school:parent-notifications')

    if request.method == 'POST' and request.POST.get('action') == 'mark_read':
        nid = request.POST.get('notification_id')
        request.user.notifications.filter(pk=nid).update(is_read=True)
        return JsonResponse({'ok': True})

    qs = request.user.notifications.select_related(
        'related_attendance__enrollment__lesson__subject',
        'related_discipline__student__user',
    ).all()

    # Filters
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    read_status = request.GET.get('read_status')

    if date_from:
        qs = qs.filter(sent_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(sent_at__date__lte=date_to)
    if read_status == 'unread':
        qs = qs.filter(is_read=False)
    elif read_status == 'read':
        qs = qs.filter(is_read=True)

    paginator = Paginator(qs, 20)
    page = paginator.get_page(request.GET.get('page'))
    unread_count = request.user.notifications.filter(is_read=False).count()

    return render(request, 'school/parent/notifications.html', {
        'notifications': page,
        'unread_count': unread_count,
    })


@login_required
def parent_complaints(request):
    """Parent: submit and track complaints."""
    if not getattr(request.user, 'is_parent', False):
        messages.error(request, "Access denied.")
        return redirect('userauths:sign-in')

    parent = _get_parent(request.user)
    if not parent:
        messages.error(request, "Parent profile not found.")
        return redirect('userauths:sign-in')

    form = ComplaintForm(parent=parent)

    if request.method == 'POST':
        form = ComplaintForm(request.POST, parent=parent)
        if form.is_valid():
            complaint = form.save(commit=False)
            complaint.parent = parent
            complaint.school = parent.school
            if complaint.is_anonymous:
                complaint.parent = parent  # still store for audit; display hides name
            complaint.save()
            messages.success(request, "Complaint submitted. The school will respond shortly.")
            return redirect('school:parent-complaints')
        else:
            messages.error(request, "Please correct the errors below.")

    complaints = Complaint.objects.filter(parent=parent).select_related('student__user', 'responded_by__user')
    paginator = Paginator(complaints, 15)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'school/parent/complaints.html', {
        'form': form,
        'complaints': page,
        'parent': parent,
    })


@login_required
def parent_assignments(request):
    """Parent: view assignments and submission status for all children."""
    if not getattr(request.user, 'is_parent', False):
        messages.error(request, "Access denied.")
        return redirect('userauths:sign-in')

    parent = _get_parent(request.user)
    if not parent:
        messages.error(request, "Parent profile not found.")
        return redirect('userauths:sign-in')

    children = parent.children.select_related('user', 'grade_level').all()
    selected_child_id = request.GET.get('child')

    if selected_child_id:
        children_qs = children.filter(pk=selected_child_id)
    else:
        children_qs = children

    assignment_data = []
    for child in children_qs:
        child_subject_ids = child.subject_enrollments.filter(is_active=True).values_list('subject_id', flat=True)
        child_subjects = Subject.objects.filter(id__in=child_subject_ids)

        assignments = Assignment.objects.filter(
            school=child.school,
            subject__in=child_subjects,
        ).select_related('subject').order_by('-due_date')[:20]

        rows = []
        for asn in assignments:
            submission = Submission.objects.filter(
                enrollment__student=child,
                assignment=asn,
            ).first()
            rows.append({'assignment': asn, 'submission': submission})

        assignment_data.append({'student': child, 'rows': rows})

    return render(request, 'school/parent/assignments.html', {
        'assignment_data': assignment_data,
        'children': children,
        'selected_child_id': selected_child_id,
        'parent': parent,
    })


@login_required
def parent_exam_results(request):
    """Parent: view published exam results for their children."""
    if not getattr(request.user, 'is_parent', False):
        messages.error(request, "Access denied.")
        return redirect('userauths:sign-in')

    parent = _get_parent(request.user)
    if not parent:
        messages.error(request, "Parent profile not found.")
        return redirect('userauths:sign-in')

    children = parent.children.select_related('user', 'grade_level', 'stream').all()
    selected_child_id = request.GET.get('child')

    results_data = []
    for child in children:
        if selected_child_id and str(child.pk) != selected_child_id:
            continue

        sessions = ExamSession.objects.filter(
            school=child.school, grade=child.grade_level, is_published=True
        ).order_by('-year', '-created_at')

        child_sessions = []
        for session in sessions:
            results = ExamResult.objects.filter(
                session=session, student=child
            ).select_related('subject').order_by('subject__name')

            stream_students = Student.objects.filter(school=child.school, stream=child.stream, is_active=True).select_related('user')
            stream_ranking = _compute_student_totals(session, stream_students, stream=child.stream)
            stream_pos = next((r['position'] for r in stream_ranking if r['student'].pk == child.pk), '–')

            overall_total = sum(r.total for r in results if r.total is not None)
            overall_pct = round((overall_total / (session.total_marks * len(results))) * 100, 1) if results and session.total_marks else None

            child_sessions.append({
                'session': session,
                'results': results,
                'stream_pos': stream_pos,
                'overall_total': round(overall_total, 2),
                'overall_pct': overall_pct,
                'overall_band': cbc_grade_band(overall_pct) if overall_pct is not None else '–',
            })

        results_data.append({'student': child, 'sessions': child_sessions})

    return render(request, 'school/parent/exam_results.html', {
        'results_data': results_data,
        'children': children,
        'selected_child_id': selected_child_id,
        'parent': parent,
    })


@login_required
def parent_fee_updates(request):
    """Parent: view fee invoices and outstanding balances for children."""
    if not getattr(request.user, 'is_parent', False):
        messages.error(request, "Access denied.")
        return redirect('userauths:sign-in')

    parent = _get_parent(request.user)
    if not parent:
        messages.error(request, "Parent profile not found.")
        return redirect('userauths:sign-in')

    children = parent.children.select_related('user', 'grade_level', 'stream').all()
    selected_child_id = request.GET.get('child')
    term_id = request.GET.get('term')
    terms = Term.objects.filter(school=parent.school).order_by('-start_date') if parent.school else Term.objects.none()

    fee_data = []
    for child in children:
        if selected_child_id and str(child.pk) != selected_child_id:
            continue
        invoices = FeeInvoice.objects.filter(school=child.school, student=child).select_related('term').order_by('-created_at')
        if term_id:
            invoices = invoices.filter(term_id=term_id)
        total_required = invoices.aggregate(t=Sum('amount_required'))['t'] or Decimal('0')
        total_paid = invoices.aggregate(t=Sum('amount_paid'))['t'] or Decimal('0')
        total_balance = total_required - total_paid
        pending_count = invoices.filter(status__in=['pending', 'partial']).count()
        fee_data.append({
            'student': child,
            'invoices': invoices,
            'total_required': total_required,
            'total_paid': total_paid,
            'total_balance': total_balance,
            'pending_count': pending_count,
        })

    return render(request, 'school/parent/fees.html', {
        'fee_data': fee_data,
        'children': children,
        'parent': parent,
        'terms': terms,
        'selected_child_id': selected_child_id,
        'term_id': term_id,
    })


@login_required
def parent_attendance(request):
    """Parent: view attendance records for children."""
    if not getattr(request.user, 'is_parent', False):
        messages.error(request, "Access denied.")
        return redirect('userauths:sign-in')

    parent = _get_parent(request.user)
    if not parent:
        messages.error(request, "Parent profile not found.")
        return redirect('userauths:sign-in')

    children = parent.children.select_related('user', 'grade_level').all()
    selected_child_id = request.GET.get('child')

    attendance_data = []
    for child in children:
        if selected_child_id and str(child.pk) != selected_child_id:
            continue

        base_records = Attendance.objects.filter(enrollment__student=child)
        total = base_records.count()
        present = base_records.filter(status='P').count()
        rate = round((present / total * 100) if total else 0, 1)

        records = base_records.select_related(
            'enrollment__lesson__subject',
            'enrollment__lesson__time_slot',
        ).order_by('-date')[:50]

        attendance_data.append({
            'student': child,
            'records': records,
            'total': total,
            'present': present,
            'rate': rate,
        })

    return render(request, 'school/parent/attendance.html', {
        'attendance_data': attendance_data,
        'children': children,
        'selected_child_id': selected_child_id,
        'parent': parent,
    })


# =============================================================================
# SCHOOL ADMIN: VIEW COMPLAINTS
# =============================================================================

@login_required
def admin_complaints_list(request):
    """Principal/admin view of all parent complaints."""
    user = request.user
    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    school = get_user_school(user)
    if not school:
        messages.error(request, "No school profile found.")
        return redirect('school:dashboard')

    status_filter = request.GET.get('status', '')
    qs = Complaint.objects.filter(school=school).select_related(
        'parent__user', 'student__user', 'responded_by__user'
    ).order_by('-created_at')
    if status_filter:
        qs = qs.filter(status=status_filter)

    if request.method == 'POST':
        complaint_id = request.POST.get('complaint_id')
        response_text = request.POST.get('response', '').strip()
        new_status = request.POST.get('status', 'resolved')
        complaint = get_object_or_404(Complaint, pk=complaint_id, school=school)
        if response_text:
            try:
                staff = user.staffprofile
            except Exception:
                staff = None
            complaint.response = response_text
            complaint.status = new_status
            complaint.responded_by = staff
            complaint.responded_at = timezone.now()
            complaint.save()
            AuditLog.objects.create(
                school=school, model_name='Complaint', object_id=complaint.pk,
                action='update', actor=user,
                description=f"Responded to complaint: {complaint.subject}",
            )
            messages.success(request, "Response saved.")
        return redirect('school:admin-complaints')

    paginator = Paginator(qs, 20)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'school/complaints.html', {
        'complaints': page,
        'status_filter': status_filter,
        'school': school,
        'status_choices': Complaint.STATUS_CHOICES,
    })


# =============================================================================
# BULK NOTIFICATION — school admins and Kiswate admins send SMS/email/in-app
# =============================================================================

def _resolve_recipients(audience, school, grade=None, stream=None):
    """
    Return list of (user, phone) tuples for the given audience.
    phone may be None if the user has no phone.
    """
    recipients = []

    def _phone(u, parent_obj=None):
        if parent_obj and parent_obj.phone:
            return parent_obj.phone
        return getattr(u, 'phone_number', None) or ''

    if audience in ('all_parents', 'all', 'grade_parents', 'stream_parents'):
        qs = Parent.objects.filter(school=school).select_related('user')
        if audience == 'grade_parents':
            qs = qs.filter(children__grade_level=grade).distinct()
        elif audience == 'stream_parents':
            qs = qs.filter(children__stream=stream).distinct()
        for p in qs:
            recipients.append((p.user, _phone(p.user, p)))

    if audience in ('all_students', 'all', 'grade_students', 'stream_students'):
        qs = Student.objects.filter(school=school, is_active=True).select_related('user')
        if audience == 'grade_students':
            qs = qs.filter(grade_level=grade)
        elif audience == 'stream_students':
            qs = qs.filter(stream=stream)
        for s in qs:
            recipients.append((s.user, _phone(s.user)))

    if audience in ('all_staff', 'all'):
        for sp in StaffProfile.objects.filter(school=school).select_related('user'):
            recipients.append((sp.user, _phone(sp.user)))

    # Deduplicate by user pk (a user might appear in multiple groups in 'all')
    seen = set()
    unique = []
    for u, ph in recipients:
        if u.pk not in seen:
            seen.add(u.pk)
            unique.append((u, ph))
    return unique


def _sms_configured():
    return bool(getattr(settings, 'SMS_API_KEY', ''))


@login_required
def bulk_notify(request):
    """Send bulk notifications via in-app, email, and/or SMS."""
    user = request.user
    is_kiswate = (
        getattr(user, 'is_kiswate_user', False) or
        getattr(user, 'is_kiswate_admin', False) or
        user.is_superuser
    )
    is_school_admin = (
        getattr(user, 'is_admin', False) or
        getattr(user, 'is_principal', False) or
        getattr(user, 'is_deputy_principal', False)
    )
    if not (is_kiswate or is_school_admin):
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    school = get_user_school(user)

    # Kiswate admins may select school via query param
    if is_kiswate and not school:
        school_id = request.GET.get('school_id') or request.POST.get('school_id')
        if school_id:
            from school.models import School as SchoolModel
            school = SchoolModel.objects.filter(pk=school_id).first()

    if not school:
        messages.error(request, "No school context found. Append ?school_id=<id> to the URL.")
        return redirect('school:dashboard')

    form = BulkNotificationForm(school=school)
    result = None

    if request.method == 'POST':
        form = BulkNotificationForm(request.POST, school=school)
        if form.is_valid():
            cd = form.cleaned_data
            title    = cd['title']
            message  = cd['message']
            audience = cd['audience']
            grade    = cd.get('grade')
            stream   = cd.get('stream')
            do_inapp = cd['channel_inapp']
            do_email = cd['channel_email']
            do_sms   = cd['channel_sms']

            recipients = _resolve_recipients(audience, school, grade, stream)

            inapp_sent = email_sent = sms_sent = 0
            email_failed = sms_failed = 0

            for recipient_user, phone in recipients:
                if do_inapp:
                    Notification.objects.create(
                        recipient=recipient_user,
                        title=title,
                        message=message,
                        school=school,
                    )
                    inapp_sent += 1

                if do_email and recipient_user.email:
                    ok = send_email(recipient_user.email, title, message)
                    if ok:
                        email_sent += 1
                    else:
                        email_failed += 1

                if do_sms and phone:
                    sms_text = f"{title}\n{message}"[:320]
                    ok = _send_sms_via_eujim(phone, sms_text)
                    if ok:
                        sms_sent += 1
                    else:
                        sms_failed += 1

            AuditLog.objects.create(
                school=school, model_name='BulkNotification', object_id=0,
                action='create', actor=user,
                description=(
                    f"Bulk notify '{title}' → audience={audience}, "
                    f"inapp={inapp_sent}, email={email_sent}, sms={sms_sent}"
                ),
            )

            result = {
                'total':        len(recipients),
                'inapp_sent':   inapp_sent,
                'email_sent':   email_sent,
                'email_failed': email_failed,
                'sms_sent':     sms_sent,
                'sms_failed':   sms_failed,
                'title':        title,
                'audience':     dict(BulkNotificationForm.AUDIENCE_CHOICES).get(audience, audience),
            }
            if do_sms and not _sms_configured():
                messages.warning(request, "SMS credentials are not configured — SMS was skipped. Set SMS_API_KEY, SMS_PARTNERID, SMS_SHORTCODE in settings.")

    return render(request, 'school/bulk_notify.html', {
        'form': form,
        'school': school,
        'result': result,
        'sms_configured': _sms_configured(),
    })


@login_required
def student_notifications(request):
    """Student: view in-app notifications."""
    if not getattr(request.user, 'is_student', False):
        messages.error(request, "This section is for students only.")
        return redirect('userauths:sign-in')

    if request.method == 'POST' and request.POST.get('action') == 'mark_all_read':
        request.user.notifications.filter(is_read=False).update(is_read=True)
        messages.success(request, "All notifications marked as read.")
        return redirect('school:student-notifications')

    notifications = request.user.notifications.select_related(
        'related_attendance__enrollment__lesson__subject',
        'related_discipline',
    ).all()

    paginator = Paginator(notifications, 20)
    page = paginator.get_page(request.GET.get('page'))
    unread_count = request.user.notifications.filter(is_read=False).count()

    return render(request, 'school/student/notifications.html', {
        'notifications': page,
        'unread_count': unread_count,
    })


# =============================================================================
# ANNOUNCEMENTS
# =============================================================================

@login_required
def announcement_list(request):
    """Admin/principal creates and views announcements."""
    user = request.user
    school = get_user_school(user)
    if not school:
        return redirect('school:dashboard')

    announcements = Announcement.objects.filter(school=school).select_related('created_by__user', 'grade')
    audience_filter = request.GET.get('audience', '')
    status_filter = request.GET.get('status', '')
    if audience_filter:
        announcements = announcements.filter(audience=audience_filter)
    now = timezone.now()
    if status_filter == 'active':
        announcements = announcements.filter(Q(expires_at__isnull=True) | Q(expires_at__gt=now))
    elif status_filter == 'expired':
        announcements = announcements.filter(expires_at__lte=now)

    context = {
        'announcements': announcements,
        'school': school,
        'audience_filter': audience_filter,
        'status_filter': status_filter,
    }
    return render(request, 'school/announcements/list.html', context)


@login_required
def announcement_create(request):
    user = request.user
    school = get_user_school(user)
    if not school:
        return redirect('school:dashboard')

    try:
        staff = user.staffprofile
    except Exception:
        messages.error(request, "Staff profile required.")
        return redirect('school:dashboard')

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        body = request.POST.get('body', '').strip()
        audience = request.POST.get('audience', 'all')
        grade_id = request.POST.get('grade')
        is_pinned = request.POST.get('is_pinned') == 'on'
        expires_at = request.POST.get('expires_at') or None

        if title and body:
            ann = Announcement.objects.create(
                school=school,
                title=title,
                body=body,
                audience=audience,
                grade_id=grade_id if grade_id else None,
                is_pinned=is_pinned,
                created_by=staff,
                expires_at=expires_at,
            )
            messages.success(request, "Announcement published.")
            return redirect('school:announcements')
        else:
            messages.error(request, "Title and body are required.")

    grades = Grade.objects.filter(school=school, is_active=True)
    return render(request, 'school/announcements/create.html', {
        'school': school,
        'grades': grades,
        'audiences': Announcement.AUDIENCE_CHOICES,
    })


@login_required
def announcement_delete(request, pk):
    school = get_user_school(request.user)
    ann = get_object_or_404(Announcement, pk=pk, school=school)
    if request.method == 'POST':
        ann.delete()
        messages.success(request, "Announcement deleted.")
    return redirect('school:announcements')


@login_required
def announcement_disable(request, pk):
    """Expire an announcement immediately so it stops showing on dashboards."""
    school = get_user_school(request.user)
    ann = get_object_or_404(Announcement, pk=pk, school=school)
    if request.method == 'POST':
        now = timezone.now()
        if ann.expires_at and ann.expires_at <= now:
            # Re-enable: clear expiry
            ann.expires_at = None
            ann.save(update_fields=['expires_at'])
            messages.success(request, "Announcement re-enabled.")
        else:
            # Disable: expire now
            ann.expires_at = now
            ann.save(update_fields=['expires_at'])
            messages.success(request, "Announcement disabled — it will no longer appear on dashboards.")
    return redirect('school:announcements')


@login_required
def parent_announcements(request):
    """Parent: view school announcements."""
    if not getattr(request.user, 'is_parent', False):
        return redirect('userauths:sign-in')

    parent = _get_parent(request.user)
    if not parent:
        return redirect('userauths:sign-in')

    children = parent.children.select_related('school').all()
    school_ids = children.values_list('school_id', flat=True).distinct()

    announcements = Announcement.objects.filter(
        school_id__in=school_ids,
        audience__in=['all', 'parents']
    ).filter(
        Q(expires_at__isnull=True) | Q(expires_at__gt=timezone.now())
    ).select_related('school', 'created_by__user').order_by('-is_pinned', '-created_at')

    paginator = Paginator(announcements, 10)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'school/announcements/parent_view.html', {
        'announcements': page,
    })


@login_required
def student_announcements(request):
    """Student: view school announcements."""
    if not getattr(request.user, 'is_student', False):
        return redirect('userauths:sign-in')

    try:
        student = request.user.student
    except Exception:
        return redirect('userauths:sign-in')

    announcements = Announcement.objects.filter(
        school=student.school,
        audience__in=['all', 'students']
    ).filter(
        Q(expires_at__isnull=True) | Q(expires_at__gt=timezone.now())
    ).select_related('created_by__user').order_by('-is_pinned', '-created_at')

    if student.grade_level:
        announcements = announcements.filter(
            Q(grade=student.grade_level) | Q(grade__isnull=True)
        )

    paginator = Paginator(announcements, 10)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'school/announcements/student_view.html', {
        'announcements': page,
        'student': student,
    })


# =============================================================================
# FINANCE DASHBOARD
# =============================================================================

@login_required
def finance_dashboard(request):
    """Modern finance dashboard with student fee balances, payments, and receipt generation."""
    user = request.user
    if not (user.is_admin or user.is_principal or user.is_deputy_principal):
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    school = get_user_school(user)
    if not school:
        return redirect('school:dashboard')

    # Filters
    q = request.GET.get('q', '').strip()
    grade_filter = request.GET.get('grade', '')
    status_filter = request.GET.get('status', '')
    term_filter = request.GET.get('term', '')

    # Student fee invoices
    inv_qs = FeeInvoice.objects.filter(school=school).select_related(
        'student__user', 'student__grade_level', 'student__stream', 'term'
    )

    if q:
        inv_qs = inv_qs.filter(
            Q(student__user__first_name__icontains=q) |
            Q(student__user__last_name__icontains=q) |
            Q(student__student_id__icontains=q)
        )
    if grade_filter:
        inv_qs = inv_qs.filter(student__grade_level_id=grade_filter)
    if status_filter:
        inv_qs = inv_qs.filter(status=status_filter)
    if term_filter:
        inv_qs = inv_qs.filter(term_id=term_filter)

    # Aggregate stats
    from django.db.models import Sum, Count
    agg = inv_qs.aggregate(
        total_expected=Sum('amount_required'),
        total_collected=Sum('amount_paid'),
        total_invoices=Count('id'),
    )
    total_expected = agg['total_expected'] or 0
    total_collected = agg['total_collected'] or 0
    total_balance = total_expected - total_collected
    collection_rate = round((total_collected / total_expected * 100) if total_expected else 0, 1)

    # Status breakdown
    status_counts = inv_qs.values('status').annotate(n=Count('id'))
    status_map = {s['status']: s['n'] for s in status_counts}

    # Paginate
    paginator = Paginator(inv_qs.order_by('-created_at'), 20)
    page = paginator.get_page(request.GET.get('page'))

    # Recent payments
    recent_payments = Payment.objects.filter(school=school).select_related(
        'student__user'
    ).order_by('-paid_at', '-id')[:10]

    grades = Grade.objects.filter(school=school, is_active=True)
    terms = Term.objects.filter(school=school).order_by('-start_date')

    context = {
        'school': school,
        'invoices': page,
        'total_expected': total_expected,
        'total_collected': total_collected,
        'total_balance': total_balance,
        'collection_rate': collection_rate,
        'status_map': status_map,
        'recent_payments': recent_payments,
        'grades': grades,
        'terms': terms,
        'q': q,
        'grade_filter': grade_filter,
        'status_filter': status_filter,
        'term_filter': term_filter,
    }
    return render(request, 'school/finance_dashboard.html', context)


@login_required
def finance_create_invoice(request):
    """Create a fee invoice for a student."""
    user = request.user
    school = get_user_school(user)
    if not school:
        return redirect('school:dashboard')

    if request.method == 'POST':
        student_id = request.POST.get('student_id')
        amount = request.POST.get('amount')
        description = request.POST.get('description', 'School Fees')
        term_id = request.POST.get('term')
        due_date = request.POST.get('due_date') or None

        try:
            student = Student.objects.get(pk=student_id, school=school)
            FeeInvoice.objects.create(
                student=student,
                school=school,
                term_id=term_id if term_id else None,
                amount_required=amount,
                description=description,
                due_date=due_date,
                created_by=user,
            )
            messages.success(request, f"Invoice created for {student.user.get_full_name()}.")
        except Exception as e:
            messages.error(request, f"Error: {e}")

    return redirect('school:finance-dashboard')


@login_required
def finance_record_payment(request, invoice_id):
    """Record a payment (cash, cheque, or M-Pesa) against a fee invoice."""
    user = request.user
    school = get_user_school(user)
    invoice = get_object_or_404(FeeInvoice, pk=invoice_id, school=school)

    if request.method == 'POST':
        amount_str = request.POST.get('amount', '0')
        method = request.POST.get('method', 'cash')
        cheque_number = request.POST.get('cheque_number', '')
        notes = request.POST.get('notes', '')

        try:
            amount = Decimal(amount_str)
            invoice.amount_paid += amount
            invoice.payment_method = method
            invoice.cheque_number = cheque_number
            invoice.notes = notes
            invoice.save()

            # Also create a Payment record for audit trail
            Payment.objects.create(
                student=invoice.student,
                school=school,
                amount=amount,
                payment_type='fees',
                status='paid',
                paid_at=timezone.now(),
                description=f"Payment for: {invoice.description} | Method: {method}",
                transaction_id=f"INV-{invoice.pk}-{uuid.uuid4().hex[:6].upper()}",
            )
            messages.success(request, f"Payment of KES {amount} recorded.")
        except Exception as e:
            messages.error(request, f"Error recording payment: {e}")

    return redirect('school:finance-dashboard')


@login_required
def finance_receipt(request, invoice_id):
    """Generate/view receipt for a fee invoice."""
    school = get_user_school(request.user)
    invoice = get_object_or_404(FeeInvoice, pk=invoice_id, school=school)
    return render(request, 'school/finance_receipt.html', {
        'invoice': invoice,
        'school': school,
    })


@login_required
def finance_stk_push(request, invoice_id):
    """Trigger M-Pesa STK push for a fee invoice."""
    school = get_user_school(request.user)
    invoice = get_object_or_404(FeeInvoice, pk=invoice_id, school=school)

    if request.method == 'POST':
        phone = request.POST.get('phone', '')
        # In production: call Daraja STK Push API here
        messages.info(request, f"STK Push initiated to {phone} for KES {invoice.balance}. Check your phone.")

    return redirect('school:finance-dashboard')

@login_required
def student_exam_results(request):
    """Student: view their own published exam results."""
    if not getattr(request.user, 'is_student', False):
        return redirect('userauths:sign-in')
    try:
        student = request.user.student
    except Exception:
        return redirect('userauths:sign-in')

    sessions = ExamSession.objects.filter(
        school=student.school, grade=student.grade_level, is_published=True
    ).order_by('-year', '-created_at')

    sessions_data = []
    for session in sessions:
        results = ExamResult.objects.filter(session=session, student=student).select_related('subject').order_by('subject__name')
        stream_students = Student.objects.filter(school=student.school, stream=student.stream, is_active=True).select_related('user')
        stream_ranking = _compute_student_totals(session, stream_students, stream=student.stream)
        stream_pos = next((r['position'] for r in stream_ranking if r['student'].pk == student.pk), '–')
        stream_total = len([r for r in stream_ranking if r['total'] is not None])

        overall_total = sum(r.total for r in results if r.total is not None)
        overall_pct = round((overall_total / (session.total_marks * len(results))) * 100, 1) if results and session.total_marks else None

        sessions_data.append({
            'session': session,
            'results': results,
            'stream_pos': stream_pos,
            'stream_total': stream_total,
            'overall_total': round(overall_total, 2),
            'overall_pct': overall_pct,
            'overall_band': cbc_grade_band(overall_pct) if overall_pct is not None else '–',
        })

    return render(request, 'school/exams/student_results.html', {
        'student': student,
        'school': student.school,
        'sessions_data': sessions_data,
    })


# =============================================================================
# FINANCE MODULE — FEE STRUCTURES, BULK INVOICING, STATEMENTS, REPORTS
# =============================================================================

@login_required
def fee_structure_list(request):
    user = request.user
    school = get_user_school(user)
    if not school or not _can_access_finance(user):
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    term_id = request.GET.get('term')
    grade_id = request.GET.get('grade')

    qs = FeeStructure.objects.filter(school=school).select_related('grade', 'stream', 'term')
    if term_id:
        qs = qs.filter(term_id=term_id)
    if grade_id:
        qs = qs.filter(grade_id=grade_id)

    terms = Term.objects.filter(school=school).order_by('-start_date')
    grades = Grade.objects.filter(school=school, is_active=True)
    paginator = Paginator(qs, 20)
    page = paginator.get_page(request.GET.get('page'))

    return render(request, 'school/finance/fee_structure_list.html', {
        'school': school, 'page_obj': page,
        'terms': terms, 'grades': grades,
        'term_id': term_id, 'grade_id': grade_id,
    })


@login_required
def fee_type_create_ajax(request):
    """AJAX endpoint: create a new FeeType for this school, return JSON {id, name}."""
    from django.http import JsonResponse
    if request.method != 'POST':
        return JsonResponse({'error': 'POST required'}, status=405)
    user = request.user
    school = get_user_school(user)
    if not school or not _can_access_finance(user):
        return JsonResponse({'error': 'Access denied'}, status=403)
    name = request.POST.get('name', '').strip()
    if not name:
        return JsonResponse({'error': 'Name is required'}, status=400)
    ft, created = FeeType.objects.get_or_create(school=school, name=name)
    return JsonResponse({'id': ft.pk, 'name': ft.name, 'created': created})


def fee_structure_create(request):
    user = request.user
    school = get_user_school(user)
    if not school or not _can_access_finance(user):
        messages.error(request, "Access denied.")
        return redirect('school:finance-dashboard')

    grades = Grade.objects.filter(school=school, is_active=True)
    streams = Streams.objects.filter(school=school)
    terms = Term.objects.filter(school=school).order_by('-start_date')
    years = AcademicYear.objects.filter(school=school).order_by('-start_date')
    fee_types = FeeType.objects.filter(school=school)

    if request.method == 'POST':
        grade_id   = request.POST.get('grade')
        stream_id  = request.POST.get('stream') or None
        term_id    = request.POST.get('term')
        year_id    = request.POST.get('academic_year') or None
        ft_ids     = request.POST.getlist('fee_type_id')
        descs      = request.POST.getlist('description')
        amounts    = request.POST.getlist('amount')

        errors = []
        if not grade_id:
            errors.append('Grade is required.')
        if not term_id:
            errors.append('Term is required.')
        if not ft_ids:
            errors.append('Add at least one fee item.')

        if not errors:
            try:
                grade  = Grade.objects.get(pk=grade_id, school=school)
                term   = Term.objects.get(pk=term_id, school=school)
                stream = Streams.objects.get(pk=stream_id, school=school) if stream_id else None
                acad   = AcademicYear.objects.get(pk=year_id, school=school) if year_id else None
            except Exception:
                errors.append('Invalid grade, stream, or term.')

        if not errors:
            try:
                staff = user.staffprofile
            except Exception:
                staff = None

            created_count = 0
            skipped_count = 0
            for ft_id, desc, amt in zip(ft_ids, descs, amounts):
                try:
                    ft = FeeType.objects.get(pk=ft_id, school=school)
                    amount_val = Decimal(amt)
                    if amount_val < 0:
                        raise ValueError
                    _, was_created = FeeStructure.objects.get_or_create(
                        school=school, grade=grade, stream=stream,
                        term=term, fee_type=ft,
                        defaults={'description': desc or ft.name, 'amount': amount_val,
                                  'academic_year': acad, 'created_by': staff}
                    )
                    if was_created:
                        created_count += 1
                    else:
                        skipped_count += 1
                except Exception:
                    errors.append(f'Invalid amount for fee type ID {ft_id}.')
                    break

            if not errors:
                msg = f'{created_count} fee structure(s) created.'
                if skipped_count:
                    msg += f' {skipped_count} skipped (already exist).'
                messages.success(request, msg)
                return redirect('school:fee-structure-list')

        for e in errors:
            messages.error(request, e)

    ctx = {
        'school': school,
        'grades': grades,
        'streams': streams,
        'terms': terms,
        'years': years,
        'fee_types': fee_types,
    }
    return render(request, 'school/finance/fee_structure_form.html', ctx)


@login_required
def fee_structure_edit(request, pk):
    user = request.user
    school = get_user_school(user)
    if not school or not _can_access_finance(user):
        messages.error(request, "Access denied.")
        return redirect('school:finance-dashboard')

    fs = get_object_or_404(FeeStructure, pk=pk, school=school)
    form = FeeStructureForm(school, request.POST or None, instance=fs)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, 'Fee structure updated.')
        return redirect('school:fee-structure-list')
    return render(request, 'school/finance/fee_structure_form.html', {'form': form, 'school': school, 'edit': True})


@login_required
def fee_structure_delete(request, pk):
    user = request.user
    school = get_user_school(user)
    if not school or not _can_access_finance(user):
        messages.error(request, "Access denied.")
        return redirect('school:finance-dashboard')

    fs = get_object_or_404(FeeStructure, pk=pk, school=school)
    if request.method == 'POST':
        fs.delete()
        messages.success(request, 'Fee structure deleted.')
        return redirect('school:fee-structure-list')
    return render(request, 'school/finance/fee_structure_confirm_delete.html', {'fs': fs, 'school': school})


@login_required
def bulk_generate_invoices(request, term_id):
    """Generate FeeInvoice records for all students matching each selected FeeStructure."""
    user = request.user
    school = get_user_school(user)
    if not school or not _can_access_finance(user):
        messages.error(request, "Access denied.")
        return redirect('school:finance-dashboard')

    term = get_object_or_404(Term, pk=term_id, school=school)
    form = BulkInvoiceGenerateForm(school, term, request.POST or None)

    if request.method == 'POST' and form.is_valid():
        fee_structures = form.cleaned_data['fee_structures']
        due_date = form.cleaned_data.get('due_date')
        overwrite = form.cleaned_data.get('overwrite_existing', False)
        created = skipped = 0

        for fs in fee_structures:
            # Get students matching this fee structure's grade (and stream if specified)
            student_qs = Student.objects.filter(school=school, grade_level=fs.grade, is_active=True)
            if fs.stream:
                student_qs = student_qs.filter(stream=fs.stream)

            for student in student_qs:
                exists = FeeInvoice.objects.filter(
                    school=school, student=student, term=term, description=fs.description
                ).exists()
                if exists and not overwrite:
                    skipped += 1
                    continue
                if exists and overwrite:
                    FeeInvoice.objects.filter(
                        school=school, student=student, term=term, description=fs.description
                    ).delete()

                FeeInvoice.objects.create(
                    school=school, student=student, term=term,
                    academic_year=fs.academic_year,
                    description=fs.description,
                    amount_required=fs.amount,
                    due_date=due_date,
                    created_by=user,
                )
                created += 1

        messages.success(request, f'{created} invoice(s) generated, {skipped} skipped (already existed).')
        return redirect('school:finance-dashboard')

    structures = FeeStructure.objects.filter(school=school, term=term, is_active=True).select_related('grade', 'stream')
    return render(request, 'school/finance/bulk_generate.html', {
        'form': form, 'school': school, 'term': term, 'structures': structures,
    })


@login_required
def fee_payment_upload(request):
    """Upload Excel file to record payments in bulk."""
    import openpyxl
    user = request.user
    school = get_user_school(user)
    if not school or not _can_access_finance(user):
        messages.error(request, "Access denied.")
        return redirect('school:finance-dashboard')

    form = FeePaymentUploadForm(school, request.POST or None, request.FILES or None)
    if request.method == 'POST' and form.is_valid():
        term = form.cleaned_data['term']
        f = form.cleaned_data['file']

        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        headers = [str(c.value).strip().lower() if c.value else '' for c in ws[1]]

        def col(name):
            return headers.index(name) if name in headers else None

        idx_sid = col('student_id')
        idx_amt = col('amount')
        idx_method = col('payment_method')
        idx_notes = col('notes')

        if idx_sid is None or idx_amt is None:
            messages.error(request, 'Columns "student_id" and "amount" are required.')
            return redirect('school:fee-payment-upload')

        saved = skipped = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            sid = str(row[idx_sid]).strip() if row[idx_sid] else ''
            if not sid:
                continue
            try:
                student = Student.objects.get(student_id=sid, school=school)
            except Student.DoesNotExist:
                skipped += 1
                continue

            try:
                amount = Decimal(str(row[idx_amt]))
            except Exception:
                skipped += 1
                continue

            method = str(row[idx_method]).strip().lower() if idx_method is not None and row[idx_method] else 'cash'
            notes = str(row[idx_notes]).strip() if idx_notes is not None and row[idx_notes] else ''

            # Find the latest unpaid/partial invoice for this student in this term
            invoice = FeeInvoice.objects.filter(
                school=school, student=student, term=term,
                status__in=['pending', 'partial']
            ).order_by('created_at').first()

            if invoice:
                invoice.amount_paid = min(invoice.amount_required, invoice.amount_paid + amount)
                invoice.payment_method = method if method in ['cash', 'mpesa', 'cheque', 'bank'] else 'cash'
                if notes:
                    invoice.notes = (invoice.notes + '\n' + notes).strip()
                invoice.save()
            else:
                # Create a generic payment record even without an invoice
                FeeInvoice.objects.create(
                    school=school, student=student, term=term,
                    description='Fee Payment (Uploaded)',
                    amount_required=amount, amount_paid=amount,
                    payment_method=method if method in ['cash', 'mpesa', 'cheque', 'bank'] else 'cash',
                    notes=notes, created_by=user,
                )

            # Audit payment record
            Payment.objects.create(
                student=student, school=school,
                amount=amount, payment_type='fees', status='paid',
                transaction_id=f"UPL-{uuid.uuid4().hex[:10].upper()}",
                description=f"Bulk upload – {term.name}",
            )
            saved += 1

        messages.success(request, f'Upload complete: {saved} payments recorded, {skipped} rows skipped.')
        return redirect('school:finance-dashboard')

    return render(request, 'school/finance/payment_upload.html', {'form': form, 'school': school})


@login_required
def student_fee_statement(request, student_pk):
    """Per-student fee statement — all invoices across terms."""
    user = request.user
    school = get_user_school(user)

    # Access: finance staff OR the student themselves OR parent of student
    is_finance = _can_access_finance(user)
    is_own = getattr(user, 'is_student', False) and hasattr(user, 'student') and user.student.pk == student_pk
    is_parent_of = False
    if getattr(user, 'is_parent', False):
        try:
            is_parent_of = user.parent.children.filter(pk=student_pk).exists()
        except Exception:
            pass

    if not (is_finance or is_own or is_parent_of):
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    student = get_object_or_404(Student, pk=student_pk)
    school = student.school

    term_id = request.GET.get('term')
    invoices = FeeInvoice.objects.filter(school=school, student=student).select_related('term').order_by('-created_at')
    if term_id:
        invoices = invoices.filter(term_id=term_id)

    total_required = invoices.aggregate(t=Sum('amount_required'))['t'] or Decimal('0')
    total_paid = invoices.aggregate(t=Sum('amount_paid'))['t'] or Decimal('0')
    total_balance = total_required - total_paid

    terms = Term.objects.filter(school=school).order_by('-start_date')
    return render(request, 'school/finance/student_statement.html', {
        'student': student, 'school': school, 'invoices': invoices,
        'total_required': total_required, 'total_paid': total_paid, 'total_balance': total_balance,
        'terms': terms, 'term_id': term_id,
    })


@login_required
def student_fee_statement_pdf(request, student_pk):
    """PDF version of a student's fee statement."""
    from io import BytesIO
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
    from reportlab.lib.enums import TA_CENTER, TA_LEFT

    user = request.user
    student = get_object_or_404(Student, pk=student_pk)
    school = student.school

    is_finance = _can_access_finance(user)
    is_own = getattr(user, 'is_student', False) and hasattr(user, 'student') and user.student.pk == student_pk
    is_parent_of = False
    if getattr(user, 'is_parent', False):
        try:
            is_parent_of = user.parent.children.filter(pk=student_pk).exists()
        except Exception:
            pass
    if not (is_finance or is_own or is_parent_of):
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')

    invoices = FeeInvoice.objects.filter(school=school, student=student).select_related('term').order_by('-created_at')
    total_required = invoices.aggregate(t=Sum('amount_required'))['t'] or Decimal('0')
    total_paid = invoices.aggregate(t=Sum('amount_paid'))['t'] or Decimal('0')
    total_balance = total_required - total_paid

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, leftMargin=1.5*cm, rightMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    green = colors.HexColor('#0b7a2a')
    red = colors.HexColor('#bb0a21')
    light_green = colors.HexColor('#e8f5e9')
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle('H1', fontSize=15, fontName='Helvetica-Bold', textColor=green, alignment=TA_CENTER, spaceAfter=2)
    sub = ParagraphStyle('Sub', fontSize=9, fontName='Helvetica', textColor=colors.grey, alignment=TA_CENTER, spaceAfter=4)
    bold9 = ParagraphStyle('B9', fontSize=9, fontName='Helvetica-Bold')

    elements = []
    elements.append(Paragraph(school.name.upper(), h1))
    elements.append(Paragraph(f"{school.address} | {school.contact_phone} | {school.contact_email}", sub))
    elements.append(HRFlowable(width='100%', thickness=2, color=green, spaceAfter=4))
    elements.append(Paragraph("FEE STATEMENT", ParagraphStyle('T', fontSize=12, fontName='Helvetica-Bold', alignment=TA_CENTER, spaceAfter=6)))
    elements.append(HRFlowable(width='100%', thickness=1, color=red, spaceAfter=8))

    info = [
        ['Student:', student.user.get_full_name(), 'Adm No:', student.student_id],
        ['Grade:', student.grade_level.name, 'Stream:', student.stream.name if student.stream else '—'],
        ['Printed:', date.today().strftime('%d/%m/%Y'), '', ''],
    ]
    info_t = Table(info, colWidths=[2.5*cm, 7*cm, 2.5*cm, 6*cm])
    info_t.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(info_t)
    elements.append(Spacer(1, 8))

    hdr = ['Term', 'Description', 'Required (KES)', 'Paid (KES)', 'Balance (KES)', 'Status']
    tdata = [hdr]
    for inv in invoices:
        tdata.append([
            inv.term.name if inv.term else '—',
            inv.description,
            f'{inv.amount_required:,.2f}',
            f'{inv.amount_paid:,.2f}',
            f'{inv.balance:,.2f}',
            inv.get_status_display(),
        ])
    tdata.append(['', 'TOTAL', f'{total_required:,.2f}', f'{total_paid:,.2f}', f'{total_balance:,.2f}', ''])

    t = Table(tdata, colWidths=[3.5*cm, 5*cm, 3*cm, 3*cm, 3*cm, 2.5*cm], repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), green),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), light_green),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, light_green]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    elements.append(t)

    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="fee_statement_{student.student_id}.pdf"'
    return response


@login_required
def finance_payment_statement(request):
    """School-level payment statement for a date range."""
    user = request.user
    school = get_user_school(user)
    if not school or not _can_access_finance(user):
        messages.error(request, "Access denied.")
        return redirect('school:finance-dashboard')

    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    term_id = request.GET.get('term')

    payments = Payment.objects.filter(school=school, status='paid').order_by('-paid_at')
    if from_date:
        payments = payments.filter(paid_at__date__gte=from_date)
    if to_date:
        payments = payments.filter(paid_at__date__lte=to_date)
    if term_id:
        # Filter FeeInvoice payments recorded in that term window
        try:
            term = Term.objects.get(pk=term_id, school=school)
            payments = payments.filter(paid_at__date__gte=term.start_date, paid_at__date__lte=term.end_date)
        except Term.DoesNotExist:
            pass

    total = payments.aggregate(t=Sum('amount'))['t'] or Decimal('0')
    by_method = FeeInvoice.objects.filter(school=school, amount_paid__gt=0)
    if term_id:
        by_method = by_method.filter(term_id=term_id)
    method_totals = by_method.values('payment_method').annotate(total=Sum('amount_paid')).order_by('-total')

    paginator = Paginator(payments.select_related('student__user'), 30)
    page = paginator.get_page(request.GET.get('page'))
    terms = Term.objects.filter(school=school).order_by('-start_date')

    return render(request, 'school/finance/payment_statement.html', {
        'school': school, 'page_obj': page, 'total': total,
        'method_totals': method_totals, 'terms': terms,
        'from_date': from_date, 'to_date': to_date, 'term_id': term_id,
    })


@login_required
def finance_collection_report(request):
    """Collection report grouped by grade and stream for a selected term."""
    user = request.user
    school = get_user_school(user)
    if not school or not _can_access_finance(user):
        messages.error(request, "Access denied.")
        return redirect('school:finance-dashboard')

    term_id = request.GET.get('term')
    terms = Term.objects.filter(school=school).order_by('-start_date')
    term = None
    report = []

    if term_id:
        try:
            term = Term.objects.get(pk=term_id, school=school)
        except Term.DoesNotExist:
            pass

    if term:
        grades = Grade.objects.filter(school=school, is_active=True).prefetch_related('streams')
        for grade in grades:
            streams = Streams.objects.filter(grade=grade, school=school)
            grade_data = {'grade': grade, 'streams': []}
            for stream in streams:
                invoices = FeeInvoice.objects.filter(
                    school=school, term=term,
                    student__stream=stream,
                )
                expected = invoices.aggregate(t=Sum('amount_required'))['t'] or Decimal('0')
                collected = invoices.aggregate(t=Sum('amount_paid'))['t'] or Decimal('0')
                balance = expected - collected
                rate = round((collected / expected * 100), 1) if expected else 0
                grade_data['streams'].append({
                    'stream': stream,
                    'expected': expected, 'collected': collected,
                    'balance': balance, 'rate': rate,
                    'students': invoices.values('student').distinct().count(),
                })
            # Grade totals
            grade_invoices = FeeInvoice.objects.filter(school=school, term=term, student__grade_level=grade)
            grade_data['expected'] = grade_invoices.aggregate(t=Sum('amount_required'))['t'] or Decimal('0')
            grade_data['collected'] = grade_invoices.aggregate(t=Sum('amount_paid'))['t'] or Decimal('0')
            grade_data['balance'] = grade_data['expected'] - grade_data['collected']
            grade_data['rate'] = round((grade_data['collected'] / grade_data['expected'] * 100), 1) if grade_data['expected'] else 0
            report.append(grade_data)

    school_expected = sum(g['expected'] for g in report)
    school_collected = sum(g['collected'] for g in report)
    school_balance = school_expected - school_collected
    school_rate = round((school_collected / school_expected * 100), 1) if school_expected else 0

    return render(request, 'school/finance/collection_report.html', {
        'school': school, 'terms': terms, 'term': term, 'report': report,
        'school_expected': school_expected, 'school_collected': school_collected,
        'school_balance': school_balance, 'school_rate': school_rate,
    })


@login_required
def fee_invoice_delete(request, invoice_id):
    user = request.user
    school = get_user_school(user)
    if not school or not _can_access_finance(user):
        messages.error(request, "Access denied.")
        return redirect('school:finance-dashboard')
    invoice = get_object_or_404(FeeInvoice, pk=invoice_id, school=school)
    if request.method == 'POST':
        invoice.delete()
        messages.success(request, 'Invoice deleted.')
    return redirect('school:finance-dashboard')


# ── Student fee portal ─────────────────────────────────────────────────────────

@login_required
def student_fees_portal(request):
    """Student views their own fee invoices."""
    if not getattr(request.user, 'is_student', False):
        return redirect('userauths:sign-in')
    try:
        student = request.user.student
    except Exception:
        return redirect('userauths:sign-in')

    term_id = request.GET.get('term')
    invoices = FeeInvoice.objects.filter(school=student.school, student=student).select_related('term').order_by('-created_at')
    if term_id:
        invoices = invoices.filter(term_id=term_id)

    total_required = invoices.aggregate(t=Sum('amount_required'))['t'] or Decimal('0')
    total_paid = invoices.aggregate(t=Sum('amount_paid'))['t'] or Decimal('0')
    total_balance = total_required - total_paid
    terms = Term.objects.filter(school=student.school).order_by('-start_date')

    return render(request, 'school/student/fees.html', {
        'student': student, 'school': student.school, 'invoices': invoices,
        'total_required': total_required, 'total_paid': total_paid,
        'total_balance': total_balance, 'terms': terms, 'term_id': term_id,
    })


# ═══════════════════════════════════════════════════════════════════════════════
# GRADE PROMOTION  (principals, deputy principals, class teachers)
# ═══════════════════════════════════════════════════════════════════════════════

def _get_class_teacher_stream(user):
    """Return the Streams assigned to this class teacher, or None."""
    try:
        sp = user.staffprofile
        if sp.position == 'class_teacher':
            assignment = sp.assigned_streams.select_related('stream').first()
            return assignment.stream if assignment else None
    except Exception:
        pass
    return None


def _can_promote(user):
    return user.is_principal or user.is_deputy_principal


@login_required
def grade_promote_view(request):
    """
    Promote students from one grade/stream to another.
    Principals & deputies: can promote any grade.
    Class teachers: scoped to their assigned stream only.
    """
    if not _can_promote(request.user):
        messages.error(request, "You do not have permission to promote students.")
        return redirect('school:dashboard')

    school = get_user_school(request.user)
    if not school:
        return redirect('school:dashboard')

    is_class_teacher = (
        hasattr(request.user, 'staffprofile') and
        request.user.staffprofile.position == 'class_teacher' and
        not (request.user.is_principal or request.user.is_deputy_principal or request.user.is_admin)
    )
    class_teacher_stream = _get_class_teacher_stream(request.user) if is_class_teacher else None

    grades = Grade.objects.filter(school=school, is_active=True).order_by('name')
    streams = Streams.objects.filter(school=school, is_active=True).select_related('grade').order_by('grade__name', 'name')

    preview_students = []
    from_grade_id = request.GET.get('from_grade') or request.POST.get('from_grade')
    from_stream_id = request.GET.get('from_stream') or request.POST.get('from_stream')
    to_grade_id = request.GET.get('to_grade') or request.POST.get('to_grade')
    to_stream_id = request.GET.get('to_stream') or request.POST.get('to_stream')

    # Class teachers are locked to their stream
    if is_class_teacher and class_teacher_stream:
        from_stream_id = str(class_teacher_stream.id)
        from_grade_id = str(class_teacher_stream.grade_id)

    if from_grade_id:
        qs = Student.objects.filter(school=school, grade_level_id=from_grade_id, is_active=True)
        if from_stream_id:
            qs = qs.filter(stream_id=from_stream_id)
        preview_students = qs.select_related('user', 'grade_level', 'stream').order_by('stream__name', 'user__last_name')

    if request.method == 'POST' and 'confirm' in request.POST:
        to_grade = get_object_or_404(Grade, id=to_grade_id, school=school)
        to_stream = None
        if to_stream_id:
            to_stream = get_object_or_404(Streams, id=to_stream_id, school=school)

        selected_ids = request.POST.getlist('student_ids')
        if selected_ids:
            # Promote only the checked students
            qs = Student.objects.filter(id__in=selected_ids, school=school)
        else:
            # Fallback: promote all in the grade/stream (shouldn't happen with the new UI)
            qs = Student.objects.filter(school=school, grade_level_id=from_grade_id, is_active=True)
            if from_stream_id:
                qs = qs.filter(stream_id=from_stream_id)

        count = qs.count()
        qs.update(grade_level=to_grade, stream=to_stream)
        messages.success(request, f"{count} student(s) promoted to {to_grade.name}{' / ' + to_stream.name if to_stream else ''}.")
        return redirect('school:grade-promote')

    return render(request, 'school/grade_promote.html', {
        'school': school,
        'grades': grades,
        'streams': streams,
        'preview_students': preview_students,
        'from_grade_id': from_grade_id,
        'from_stream_id': from_stream_id,
        'to_grade_id': to_grade_id,
        'to_stream_id': to_stream_id,
        'is_class_teacher': is_class_teacher,
        'class_teacher_stream': class_teacher_stream,
    })


# ═══════════════════════════════════════════════════════════════════════════════
# CLASS TEACHER PORTAL
# ═══════════════════════════════════════════════════════════════════════════════

def _require_class_teacher(request):
    """Return (teacher, stream, school) or (None, None, None). Principals/deputies/admins also allowed."""
    user = request.user
    is_school_admin = user.is_principal or user.is_deputy_principal or user.is_admin

    try:
        teacher = user.staffprofile
        school = teacher.school
    except Exception:
        if is_school_admin:
            school = get_user_school(user)
            if not school:
                messages.error(request, "No school found.")
                return None, None, None
            teacher = None
        else:
            messages.error(request, "Staff profile required.")
            return None, None, None

    stream_id = request.GET.get('stream') or request.POST.get('stream')

    # Principals/deputies/admins: any stream in the school via ?stream=<id>
    if is_school_admin:
        stream = Streams.objects.filter(id=stream_id, school=school).first() if stream_id else None
        return teacher, stream, school

    # Role-based check: must have 'class_teacher' role
    if not teacher.is_class_teacher:
        messages.error(request, "Class teacher access only.")
        return None, None, None

    assignments = list(ClassTeacherAssignment.objects.filter(
        teacher=teacher, school=school
    ).select_related('stream__grade'))

    if not assignments:
        messages.error(request, "You have no assigned class. Ask your principal to assign you.")
        return None, None, None

    # If a stream is specified in URL, validate it belongs to this teacher
    if stream_id:
        match = next((a for a in assignments if str(a.stream_id) == str(stream_id)), None)
        if match:
            return teacher, match.stream, school
        messages.error(request, "You are not assigned to that stream.")
        return None, None, None

    # Single assignment — auto-select
    if len(assignments) == 1:
        return teacher, assignments[0].stream, school

    # Multiple assignments — return None stream to trigger stream picker
    return teacher, None, school


@login_required
def manage_class_teachers(request):
    """Principal/deputy: list all class teacher assignments and assign/remove."""
    user = request.user
    if not (user.is_principal or user.is_deputy_principal or user.is_admin):
        messages.error(request, "Access denied.")
        return redirect('school:dashboard')
    school = get_user_school(user)
    if not school:
        return redirect('school:dashboard')

    # Auto-create the 'class_teacher' role for this school if not yet present
    ct_role, _ = Role.objects.get_or_create(
        name='class_teacher', school=school,
        defaults={'description': 'Class teacher with roster and roll-call access'}
    )

    assignments = ClassTeacherAssignment.objects.filter(school=school).select_related(
        'teacher__user', 'stream__grade', 'assigned_by__user'
    )
    # Staff who can be assigned (teachers)
    staff = StaffProfile.objects.filter(school=school).select_related('user').order_by('user__last_name')
    streams = Streams.objects.filter(school=school).select_related('grade').order_by('grade__name', 'name')
    # Already-assigned stream IDs
    assigned_stream_ids = set(assignments.values_list('stream_id', flat=True))

    if request.method == 'POST':
        action = request.POST.get('action')
        if action == 'assign':
            teacher_id = request.POST.get('teacher_id')
            stream_id = request.POST.get('stream_id')
            try:
                teacher_sp = StaffProfile.objects.get(pk=teacher_id, school=school)
                stream = Streams.objects.get(pk=stream_id, school=school)
            except (StaffProfile.DoesNotExist, Streams.DoesNotExist):
                messages.error(request, "Invalid teacher or stream.")
                return redirect('school:manage-class-teachers')

            # Grant the class_teacher role
            teacher_sp.roles.add(ct_role)
            # Create or update the assignment
            ClassTeacherAssignment.objects.update_or_create(
                stream=stream, school=school,
                defaults={'teacher': teacher_sp, 'assigned_by': _safe_staffprofile(user)}
            )
            messages.success(request, f"{teacher_sp.user.get_full_name()} assigned as class teacher for {stream}.")

        elif action == 'remove':
            assignment_id = request.POST.get('assignment_id')
            try:
                assignment = ClassTeacherAssignment.objects.get(pk=assignment_id, school=school)
            except ClassTeacherAssignment.DoesNotExist:
                messages.error(request, "Assignment not found.")
                return redirect('school:manage-class-teachers')
            teacher_sp = assignment.teacher
            assignment.delete()
            # Remove the role only if teacher has no remaining class assignments
            if not ClassTeacherAssignment.objects.filter(teacher=teacher_sp, school=school).exists():
                teacher_sp.roles.remove(ct_role)
            messages.success(request, f"Class teacher assignment removed.")

        return redirect('school:manage-class-teachers')

    return render(request, 'school/class_teacher/manage.html', {
        'assignments': assignments,
        'staff': staff,
        'streams': streams,
        'assigned_stream_ids': assigned_stream_ids,
        'school': school,
    })


def _safe_staffprofile(user):
    try:
        return user.staffprofile
    except Exception:
        return None


@login_required
def class_teacher_roster(request):
    """Class teacher: view all students in their assigned stream."""
    user = request.user
    is_admin_user = user.is_principal or user.is_deputy_principal or user.is_admin
    base_template = 'school/base.html' if is_admin_user else 'school/teacher/base.html'

    teacher, stream, school = _require_class_teacher(request)
    if school is None:
        return redirect('school:dashboard')

    if stream is None:
        # Show stream picker — all streams for admins, only assigned for class teachers
        if is_admin_user:
            streams = Streams.objects.filter(school=school).select_related('grade')
        else:
            streams = [a.stream for a in ClassTeacherAssignment.objects.filter(
                teacher=teacher, school=school
            ).select_related('stream__grade')]
        return render(request, 'school/class_teacher/pick_stream.html', {
            'streams': streams,
            'base_template': base_template,
        })

    students = stream.student_stream.filter(is_active=True).select_related(
        'user', 'grade_level', 'stream'
    ).order_by('user__last_name', 'user__first_name')

    today = timezone.localdate()
    rows = []
    for s in students:
        att_qs = Attendance.objects.filter(enrollment__student=s)
        total = att_qs.count()
        present = att_qs.filter(status='P').count()
        rate = int(present / total * 100) if total else 0
        discipline = DisciplineRecord.objects.filter(student=s, resolved=False).count()
        fee_due = FeeInvoice.objects.filter(student=s).exclude(status='paid').aggregate(
            bal=Sum(F('amount_required') - F('amount_paid'))
        )['bal'] or 0
        rows.append({
            'student': s,
            'att_rate': rate,
            'discipline': discipline,
            'fee_due': fee_due,
        })

    fee_issues_count = sum(1 for r in rows if r['fee_due'] > 0)
    discipline_open_count = sum(1 for r in rows if r['discipline'] > 0)

    return render(request, 'school/class_teacher/roster.html', {
        'teacher': teacher,
        'stream': stream,
        'rows': rows,
        'today': today,
        'fee_issues_count': fee_issues_count,
        'discipline_open_count': discipline_open_count,
        'is_admin_user': is_admin_user,
        'base_template': base_template,
    })


@login_required
def class_teacher_roll_call(request):
    """Class teacher: daily roll call attendance for their stream."""
    user = request.user
    is_admin_user = user.is_principal or user.is_deputy_principal or user.is_admin
    base_template = 'school/base.html' if is_admin_user else 'school/teacher/base.html'

    teacher, stream, school = _require_class_teacher(request)
    if school is None:
        return redirect('school:dashboard')
    if stream is None:
        return redirect('school:class-teacher-roster')

    date_str = request.GET.get('date')
    report_date = date.fromisoformat(date_str) if date_str else timezone.localdate()
    prev_date = report_date - timedelta(days=1)
    next_date = report_date + timedelta(days=1)

    students = stream.student_stream.filter(is_active=True).select_related('user').order_by(
        'user__last_name', 'user__first_name'
    )

    if request.method == 'POST':
        for student in students:
            key = f"status_{student.id}"
            status = request.POST.get(key, 'UA')
            GradeAttendance.objects.update_or_create(
                student=student, stream=stream,
                recorded_at__date=report_date,
                defaults={'status': status, 'recorded_at': timezone.now()},
            )
        messages.success(request, f"Roll call saved for {report_date.strftime('%A, %d %b %Y')}.")
        return redirect(f"{request.path}?date={report_date}")

    # Fetch existing marks for this date
    existing = {
        ga.student_id: ga.status
        for ga in GradeAttendance.objects.filter(
            stream=stream, recorded_at__date=report_date
        )
    }

    rows = [{'student': s, 'status': existing.get(s.id, '')} for s in students]
    STATUS_CHOICES = [
        ('P', 'Present'), ('EA', 'Excused Absent'), ('UA', 'Unexcused Absent'),
        ('ET', 'Excused Tardy'), ('UT', 'Unexcused Tardy'), ('IB', 'In Building'),
    ]
    summary = {s: sum(1 for r in rows if r['status'] == s) for s, _ in STATUS_CHOICES}

    return render(request, 'school/class_teacher/roll_call.html', {
        'teacher': teacher, 'stream': stream,
        'rows': rows, 'report_date': report_date,
        'prev_date': prev_date, 'next_date': next_date,
        'STATUS_CHOICES': STATUS_CHOICES, 'summary': summary,
        'base_template': base_template,
    })


@login_required
def class_teacher_attendance_summary(request):
    """Class teacher: per-student attendance summary over a date range."""
    user = request.user
    is_admin_user = user.is_principal or user.is_deputy_principal or user.is_admin
    base_template = 'school/base.html' if is_admin_user else 'school/teacher/base.html'

    teacher, stream, school = _require_class_teacher(request)
    if school is None:
        return redirect('school:dashboard')
    if stream is None:
        return redirect('school:class-teacher-roster')

    start_str = request.GET.get('start')
    end_str = request.GET.get('end')
    period_start = date.fromisoformat(start_str) if start_str else timezone.localdate().replace(day=1)
    period_end = date.fromisoformat(end_str) if end_str else timezone.localdate()

    students = stream.student_stream.filter(is_active=True).select_related('user').order_by(
        'user__last_name', 'user__first_name'
    )

    rows = []
    STATUS_LABELS = {
        'P': 'Present', 'EA': 'Excused Absent', 'UA': 'Unexcused Absent',
        'ET': 'Excused Tardy', 'UT': 'Unexcused Tardy', 'IB': 'In Building',
        '18': 'Suspended', '20': 'Expelled',
    }
    for student in students:
        att_qs = Attendance.objects.filter(
            enrollment__student=student,
            enrollment__lesson__lesson_date__range=[period_start, period_end],
        )
        total = att_qs.count()
        present = att_qs.filter(status='P').count()
        rate = int(present / total * 100) if total else 0
        breakdown = {row['status']: row['c'] for row in att_qs.values('status').annotate(c=Count('status'))}
        rows.append({'student': student, 'total': total, 'present': present, 'rate': rate, 'breakdown': breakdown})

    return render(request, 'school/class_teacher/attendance_summary.html', {
        'teacher': teacher, 'stream': stream,
        'rows': rows, 'period_start': period_start, 'period_end': period_end,
        'STATUS_LABELS': STATUS_LABELS,
        'base_template': base_template,
    })


@login_required
def class_teacher_subject_summary(request):
    """Class teacher: per-subject attendance rates for their stream."""
    user = request.user
    is_admin_user = user.is_principal or user.is_deputy_principal or user.is_admin
    base_template = 'school/base.html' if is_admin_user else 'school/teacher/base.html'

    teacher, stream, school = _require_class_teacher(request)
    if school is None:
        return redirect('school:dashboard')
    if stream is None:
        return redirect('school:class-teacher-roster')

    start_str = request.GET.get('start')
    end_str = request.GET.get('end')
    period_start = date.fromisoformat(start_str) if start_str else timezone.localdate().replace(day=1)
    period_end = date.fromisoformat(end_str) if end_str else timezone.localdate()

    lessons = Lesson.objects.filter(
        stream=stream,
        lesson_date__range=[period_start, period_end],
    ).select_related('subject', 'teacher__user', 'time_slot')

    from collections import defaultdict
    subject_data = defaultdict(lambda: {'total': 0, 'present': 0, 'lessons': 0, 'subject': None, 'teacher': None})
    for lesson in lessons:
        key = lesson.subject_id
        subject_data[key]['subject'] = lesson.subject
        subject_data[key]['teacher'] = lesson.teacher
        subject_data[key]['lessons'] += 1
        att_qs = Attendance.objects.filter(enrollment__lesson=lesson)
        total = att_qs.count()
        present = att_qs.filter(status='P').count()
        subject_data[key]['total'] += total
        subject_data[key]['present'] += present

    rows = []
    for data in subject_data.values():
        rate = int(data['present'] / data['total'] * 100) if data['total'] else 0
        rows.append({**data, 'rate': rate})
    rows.sort(key=lambda r: r['rate'])

    return render(request, 'school/class_teacher/subject_summary.html', {
        'teacher': teacher, 'stream': stream,
        'rows': rows, 'period_start': period_start, 'period_end': period_end,
        'base_template': base_template,
    })


@login_required
def class_teacher_discipline(request):
    """Class teacher: view and log discipline records for their class."""
    user = request.user
    is_admin_user = user.is_principal or user.is_deputy_principal or user.is_admin
    base_template = 'school/base.html' if is_admin_user else 'school/teacher/base.html'

    teacher, stream, school = _require_class_teacher(request)
    if school is None:
        return redirect('school:dashboard')
    if stream is None:
        return redirect('school:class-teacher-roster')

    students_in_class = stream.student_stream.filter(is_active=True)
    q = request.GET.get('q', '').strip()
    resolved = request.GET.get('resolved', '')

    records = DisciplineRecord.objects.filter(
        student__in=students_in_class
    ).select_related('student__user', 'teacher__user').order_by('-date')

    if q:
        records = records.filter(
            Q(student__user__first_name__icontains=q) |
            Q(student__user__last_name__icontains=q) |
            Q(description__icontains=q)
        )
    if resolved == '1':
        records = records.filter(resolved=True)
    elif resolved == '0':
        records = records.filter(resolved=False)

    page_obj = Paginator(records, 30).get_page(request.GET.get('page'))

    return render(request, 'school/class_teacher/discipline.html', {
        'teacher': teacher, 'stream': stream,
        'page_obj': page_obj, 'q': q, 'resolved': resolved,
        'total_open': DisciplineRecord.objects.filter(student__in=students_in_class, resolved=False).count(),
        'base_template': base_template,
    })


# ═══════════════════════════════════════════════════════════════════════════════
# REPORT EXPORTS
# ═══════════════════════════════════════════════════════════════════════════════

@login_required
def finance_collection_report_csv(request):
    """CSV export of fee collection report for a selected term."""
    user = request.user
    school = get_user_school(user)
    if not school or not _can_access_finance(user):
        messages.error(request, "Access denied.")
        return redirect('school:finance-dashboard')

    term_id = request.GET.get('term')
    if not term_id:
        messages.error(request, "Select a term first, then export.")
        return redirect('school:finance-collection-report')

    try:
        term = Term.objects.get(pk=term_id, school=school)
    except Term.DoesNotExist:
        messages.error(request, "Term not found.")
        return redirect('school:finance-collection-report')

    response = HttpResponse(content_type='text/csv')
    fname = f"collection_report_{term.name.replace(' ', '_')}.csv"
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    writer = csv.writer(response)
    writer.writerow(['Grade', 'Stream', 'Students w/ Invoices',
                     'Expected (KES)', 'Collected (KES)', 'Balance (KES)', 'Rate (%)'])

    grades = Grade.objects.filter(school=school, is_active=True).order_by('name')
    school_exp = school_col = Decimal('0')
    for grade in grades:
        streams = Streams.objects.filter(grade=grade, school=school).order_by('name')
        g_exp = g_col = Decimal('0')
        for stream in streams:
            invoices = FeeInvoice.objects.filter(school=school, term=term, student__stream=stream)
            exp = invoices.aggregate(t=Sum('amount_required'))['t'] or Decimal('0')
            col = invoices.aggregate(t=Sum('amount_paid'))['t'] or Decimal('0')
            bal = exp - col
            rate = round(float(col / exp * 100), 1) if exp else 0
            students = invoices.values('student').distinct().count()
            writer.writerow([grade.name, stream.name, students, exp, col, bal, rate])
            g_exp += exp; g_col += col
        g_bal = g_exp - g_col
        g_rate = round(float(g_col / g_exp * 100), 1) if g_exp else 0
        writer.writerow([f'{grade.name} — SUBTOTAL', '', '', g_exp, g_col, g_bal, g_rate])
        school_exp += g_exp; school_col += g_col

    s_bal = school_exp - school_col
    s_rate = round(float(school_col / school_exp * 100), 1) if school_exp else 0
    writer.writerow(['SCHOOL TOTAL', '', '', school_exp, school_col, s_bal, s_rate])
    return response


@login_required
def finance_payment_statement_csv(request):
    """CSV export of payment statement (mirrors the HTML view filters)."""
    user = request.user
    school = get_user_school(user)
    if not school or not _can_access_finance(user):
        messages.error(request, "Access denied.")
        return redirect('school:finance-dashboard')

    from_date = request.GET.get('from_date')
    to_date   = request.GET.get('to_date')
    term_id   = request.GET.get('term')

    payments = Payment.objects.filter(school=school, status='paid').order_by('-paid_at').select_related(
        'student__user', 'student__grade_level'
    )
    if from_date:
        payments = payments.filter(paid_at__date__gte=from_date)
    if to_date:
        payments = payments.filter(paid_at__date__lte=to_date)
    if term_id:
        try:
            term = Term.objects.get(pk=term_id, school=school)
            payments = payments.filter(paid_at__date__gte=term.start_date, paid_at__date__lte=term.end_date)
        except Term.DoesNotExist:
            pass

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="payment_statement.csv"'
    writer = csv.writer(response)
    writer.writerow(['Student', 'Adm No', 'Grade', 'Type', 'Description', 'Amount (KES)', 'Date', 'Status'])
    for pay in payments:
        student = pay.student
        writer.writerow([
            student.user.get_full_name() if student else '—',
            student.student_id if student else '—',
            student.grade_level.name if student else '—',
            pay.get_payment_type_display() if hasattr(pay, 'get_payment_type_display') else '',
            pay.description or '',
            pay.amount,
            pay.paid_at.strftime('%Y-%m-%d') if pay.paid_at else '—',
            pay.status,
        ])
    return response


@login_required
def exam_ranking_stream_pdf(request, session_pk, stream_pk):
    """PDF export of stream-level exam ranking."""
    user = request.user
    school = get_user_school(user)
    if not (user.is_admin or user.is_principal or user.is_deputy_principal or getattr(user, 'is_teacher', False)):
        messages.error(request, "Access denied.")
        return redirect('school:school-exams')

    session = get_object_or_404(ExamSession, pk=session_pk, school=school)
    stream  = get_object_or_404(Streams, pk=stream_pk, school=school)
    students = Student.objects.filter(school=school, stream=stream, is_active=True).select_related('user')
    ranking  = _compute_student_totals(session, students, stream=stream)

    response = HttpResponse(content_type='application/pdf')
    fname = f"ranking_{stream.name}_{session.name}.pdf".replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'

    from reportlab.platypus import Spacer
    doc = SimpleDocTemplate(response, pagesize=letter, topMargin=40, bottomMargin=36, leftMargin=40, rightMargin=40)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"<b>{school.name}</b>", styles['Title']))
    story.append(Paragraph(f"Stream Ranking &mdash; {session.name} &nbsp;|&nbsp; {stream.grade.name} &ndash; {stream.name}", styles['Heading2']))
    story.append(Paragraph(f"Generated: {timezone.now().strftime('%d %B %Y %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 14))

    header = ['#', 'Student Name', 'Adm No', 'Total Score', 'Subjects Sat']
    rows   = [header]
    for row in ranking:
        rows.append([
            str(row['position']),
            row['student'].user.get_full_name(),
            row['student'].student_id,
            str(round(row['total'], 1)) if row['total'] is not None else '—',
            str(row['subjects_count']),
        ])

    t = Table(rows, colWidths=[28, 220, 80, 80, 72])
    t.setStyle(TableStyle([
        ('BACKGROUND',   (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR',    (0, 0), (-1, 0), colors.white),
        ('FONTNAME',     (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',     (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f4f6f9')]),
        ('BACKGROUND',   (0, 1), (-1, 1), colors.HexColor('#fff3cd')),
        ('GRID',         (0, 0), (-1, -1), 0.25, colors.HexColor('#dee2e6')),
        ('ALIGN',        (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN',        (3, 0), (4, -1), 'CENTER'),
        ('TOPPADDING',   (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    doc.build(story)
    return response


@login_required
def exam_ranking_grade_pdf(request, session_pk):
    """PDF export of full grade-level exam ranking."""
    user = request.user
    school = get_user_school(user)
    if not (user.is_admin or user.is_principal or user.is_deputy_principal or getattr(user, 'is_teacher', False)):
        messages.error(request, "Access denied.")
        return redirect('school:school-exams')

    session  = get_object_or_404(ExamSession, pk=session_pk, school=school)
    students = Student.objects.filter(school=school, grade_level=session.grade, is_active=True).select_related('user', 'stream')
    ranking  = _compute_student_totals(session, students)

    response = HttpResponse(content_type='application/pdf')
    fname = f"grade_ranking_{session.grade.name}_{session.name}.pdf".replace(' ', '_')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'

    from reportlab.platypus import Spacer
    doc = SimpleDocTemplate(response, pagesize=letter, topMargin=40, bottomMargin=36, leftMargin=40, rightMargin=40)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"<b>{school.name}</b>", styles['Title']))
    story.append(Paragraph(f"Grade Ranking &mdash; {session.name} &nbsp;|&nbsp; {session.grade.name}", styles['Heading2']))
    story.append(Paragraph(f"Generated: {timezone.now().strftime('%d %B %Y %H:%M')}", styles['Normal']))
    story.append(Spacer(1, 14))

    header = ['#', 'Student Name', 'Adm No', 'Stream', 'Total Score', 'Subj']
    rows   = [header]
    for row in ranking:
        rows.append([
            str(row['position']),
            row['student'].user.get_full_name(),
            row['student'].student_id,
            row['student'].stream.name if row['student'].stream else '—',
            str(round(row['total'], 1)) if row['total'] is not None else '—',
            str(row['subjects_count']),
        ])

    t = Table(rows, colWidths=[28, 200, 75, 70, 70, 37])
    t.setStyle(TableStyle([
        ('BACKGROUND',   (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
        ('TEXTCOLOR',    (0, 0), (-1, 0), colors.white),
        ('FONTNAME',     (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE',     (0, 0), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f4f6f9')]),
        ('BACKGROUND',   (0, 1), (-1, 1), colors.HexColor('#fff3cd')),
        ('GRID',         (0, 0), (-1, -1), 0.25, colors.HexColor('#dee2e6')),
        ('ALIGN',        (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN',        (4, 0), (5, -1), 'CENTER'),
        ('TOPPADDING',   (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING',(0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    doc.build(story)
    return response
